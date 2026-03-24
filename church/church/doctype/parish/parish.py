# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import now_datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

@frappe.whitelist()
def get_workers_from_members(parish_name):
    """
    Fetch all workers from Member doctype where parish matches
    and is_a_worker is checked
    """
    workers = frappe.get_all('Member',
        filters={
            'parish': parish_name,
            'is_a_worker': 1,
            'member_status': 'Active'
        },
        fields=[
            'full_name',
            'gender',
            'date_of_birth',
            'mobile_phone',
            'email',
            'alternative_phone',
            'designation',
            'date_of_joining'
        ],
        order_by='full_name asc'
    )
    
    return workers

@frappe.whitelist()
def export_workers_to_excel(parish_name, workers):
    """
    Export workers data to a formatted Excel file
    """
    import json
    
    # Parse workers if it's a string
    if isinstance(workers, str):
        workers = json.loads(workers)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Workers"
    
    # Get parish details
    parish = frappe.get_doc('Parish', parish_name)
    
    # Title Section
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"WORKERS REPORT - {parish_name.upper()}"
    title_cell.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Parish Info Section
    ws.merge_cells('A2:B2')
    ws['A2'] = "Parish Pastor:"
    ws['A2'].font = Font(bold=True)
    ws['C2'] = parish.pastor_name or ''
    
    ws.merge_cells('A3:B3')
    ws['A3'] = "Contact:"
    ws['A3'].font = Font(bold=True)
    ws['C3'] = parish.phone or ''
    
    ws.merge_cells('E2:F2')
    ws['E2'] = "Generated On:"
    ws['E2'].font = Font(bold=True)
    ws['G2'] = now_datetime().strftime('%Y-%m-%d %H:%M:%S')
    
    ws.merge_cells('E3:F3')
    ws['E3'] = "Total Workers:"
    ws['E3'].font = Font(bold=True)
    ws['G3'] = len(workers)
    
    # Add spacing
    ws.row_dimensions[4].height = 5
    
    # Headers
    headers = [
        'S/N',
        'Full Name',
        'Gender',
        'Date of Birth',
        'Mobile Phone',
        'Alternative Phone',
        'Email',
        'Department',
        'Date of Joining'
    ]
    
    header_row = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws.row_dimensions[header_row].height = 25
    
    # Data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, worker in enumerate(workers, 1):
        row_num = header_row + idx
        
        # Alternate row colors
        if idx % 2 == 0:
            row_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # S/N
        cell = ws.cell(row=row_num, column=1, value=idx)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = row_fill
        
        # Full Name
        cell = ws.cell(row=row_num, column=2, value=worker.get('full_name', ''))
        cell.border = thin_border
        cell.fill = row_fill
        
        # Gender
        cell = ws.cell(row=row_num, column=3, value=worker.get('gender', ''))
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = row_fill
        
        # Date of Birth
        cell = ws.cell(row=row_num, column=4, value=worker.get('date_of_birth', ''))
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = row_fill
        
        # Mobile Phone
        cell = ws.cell(row=row_num, column=5, value=worker.get('mobile_phone', ''))
        cell.border = thin_border
        cell.fill = row_fill
        
        # Alternative Phone
        cell = ws.cell(row=row_num, column=6, value=worker.get('alternative_phone', ''))
        cell.border = thin_border
        cell.fill = row_fill
        
        # Email
        cell = ws.cell(row=row_num, column=7, value=worker.get('email', ''))
        cell.border = thin_border
        cell.fill = row_fill
        
        # Department
        cell = ws.cell(row=row_num, column=8, value=worker.get('department', ''))
        cell.border = thin_border
        cell.fill = row_fill
        
        # Date of Joining
        cell = ws.cell(row=row_num, column=9, value=worker.get('date_of_joining', ''))
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell.fill = row_fill
    
    # Summary Section
    summary_row = header_row + len(workers) + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Male Count
    male_count = len([w for w in workers if w.get('gender') == 'Male'])
    ws[f'A{summary_row + 1}'] = "Male Workers:"
    ws[f'A{summary_row + 1}'].font = Font(bold=True)
    ws[f'B{summary_row + 1}'] = male_count
    
    # Female Count
    female_count = len([w for w in workers if w.get('gender') == 'Female'])
    ws[f'A{summary_row + 2}'] = "Female Workers:"
    ws[f'A{summary_row + 2}'].font = Font(bold=True)
    ws[f'B{summary_row + 2}'] = female_count
    
    # Total
    ws[f'A{summary_row + 3}'] = "Total Workers:"
    ws[f'A{summary_row + 3}'].font = Font(bold=True)
    ws[f'B{summary_row + 3}'] = len(workers)
    ws[f'B{summary_row + 3}'].font = Font(bold=True)
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # S/N
        'B': 30,  # Full Name
        'C': 12,  # Gender
        'D': 15,  # DOB
        'E': 18,  # Mobile Phone
        'F': 18,  # Alternative Phone
        'G': 30,  # Email
        'H': 20,  # Department
        'I': 15   # Date of Joining
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save file
    file_name = f"Workers_{parish_name.replace(' ', '_')}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(frappe.get_site_path('public', 'files'), file_name)
    
    wb.save(file_path)
    
    # Return file URL
    file_url = f"/files/{file_name}"
    
    return {
        'file_url': file_url,
        'file_name': file_name
    }