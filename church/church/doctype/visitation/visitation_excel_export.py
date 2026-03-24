# -*- coding: utf-8 -*-
"""
Beautiful Excel Export for Visitation Reports
Professional formatting with colors, charts, and styling
"""

from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import getdate, formatdate, now_datetime, nowdate, format_date, format_datetime, get_url

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import io
import base64


@frappe.whitelist()
def export_visitation_to_excel(visitation_name):
    """
    Export Visitation Report to beautiful Excel file
    
    Args:
        visitation_name: Name of the Visitation document
    
    Returns:
        Dict with file content and filename
    """
    
    # Check if openpyxl is installed
    if not HAS_OPENPYXL:
        frappe.throw(_("""
            Excel export requires 'openpyxl' library. 
            Please install it using: bench pip install openpyxl
        """))
    
    try:
        # Get visitation document
        doc = frappe.get_doc("Visitation", visitation_name)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        create_summary_sheet(wb, doc)
        create_team_details_sheet(wb, doc)
        create_report_sheet(wb, doc)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Encode to base64
        file_content = base64.b64encode(excel_file.read()).decode('utf-8')
        
        # Generate filename
        filename = f"Visitation_Report_{doc.branch}_{doc.visitee_full_name}_{formatdate(doc.date_of_visitation, 'yyyy-MM-dd')}.xlsx"
        filename = filename.replace(" ", "_")
        
        return {
            'success': True,
            'file_content': file_content,
            'filename': filename,
            'message': _('Excel file generated successfully')
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), 'Visitation Excel Export Error')
        return {
            'success': False,
            'error': str(e)
        }


def create_summary_sheet(wb, doc):
    """Create beautifully formatted summary overview sheet"""
    
    ws = wb.create_sheet("Visitation Summary", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    
    # Define beautiful color palette
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    subheader_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    section_fill = PatternFill(start_color="FA709A", end_color="FA709A", fill_type="solid")
    highlight_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    info_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=24, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    section_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    label_font = Font(name='Calibri', size=11, bold=True, color="2C3E50")
    value_font = Font(name='Calibri', size=11, color="34495E")
    status_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    current_row = 1
    
    # TITLE SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "⛪ VISITATION REPORT"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = header_gradient
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 45
    current_row += 1
    
    # BRANCH & DATE
    ws.merge_cells(f'A{current_row}:D{current_row}')
    formatted_date = format_date(doc.date_of_visitation, "dd MMMM yyyy") if doc.date_of_visitation else ""
    ws[f'A{current_row}'] = f"{doc.branch} • {formatted_date}"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    ws[f'A{current_row}'].fill = subheader_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Empty row for spacing
    ws.row_dimensions[current_row].height = 10
    current_row += 1
    
    # STATUS INDICATOR
    ws.merge_cells(f'A{current_row}:D{current_row}')
    status_text = f"📊 STATUS: {doc.status.upper()}"
    ws[f'A{current_row}'] = status_text
    ws[f'A{current_row}'].font = status_font
    
    # Color based on status
    status_colors = {
        "Assigned": "3498DB",
        "In Progress": "F39C12",
        "Completed": "2ECC71",
        "Cancelled": "E74C3C",
        "Rescheduled": "9B59B6"
    }
    status_color = status_colors.get(doc.status, "95A5A6")
    ws[f'A{current_row}'].fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 35
    current_row += 1
    
    # Spacing
    current_row += 1
    
    # BASIC INFORMATION SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "📋 BASIC INFORMATION"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = section_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Basic info data
    basic_info = [
        ("Visitation ID:", doc.name, "Type:", doc.type),
        ("Member Visited:", doc.visitee_full_name, "Gender:", doc.gender or "N/A"),
        ("Date of Visit:", formatted_date, "Day:", doc.day or ""),
        ("Location:", doc.location or "N/A", "Branch:", doc.branch),
    ]
    
    for row_data in basic_info:
        # First pair
        ws[f'A{current_row}'] = row_data[0]
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].fill = info_fill
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row}'] = row_data[1]
        ws[f'B{current_row}'].font = value_font
        ws[f'B{current_row}'].fill = highlight_fill
        ws[f'B{current_row}'].border = thin_border
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Second pair
        ws[f'C{current_row}'] = row_data[2]
        ws[f'C{current_row}'].font = label_font
        ws[f'C{current_row}'].fill = info_fill
        ws[f'C{current_row}'].border = thin_border
        ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'D{current_row}'] = row_data[3]
        ws[f'D{current_row}'].font = value_font
        ws[f'D{current_row}'].fill = highlight_fill
        ws[f'D{current_row}'].border = thin_border
        ws[f'D{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    
    # Spacing
    current_row += 1
    
    # CONTACT INFORMATION SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "📞 CONTACT INFORMATION"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = section_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Contact info
    contact_info = [
        ("Email:", doc.email or "N/A", "Phone:", doc.phone or "N/A"),
        ("Alternative Phone:", doc.alternative_phone or "N/A", "", ""),
    ]
    
    for row_data in contact_info:
        ws[f'A{current_row}'] = row_data[0]
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].fill = info_fill
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row}'] = row_data[1]
        ws[f'B{current_row}'].font = value_font
        ws[f'B{current_row}'].fill = highlight_fill
        ws[f'B{current_row}'].border = thin_border
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        if row_data[2]:
            ws[f'C{current_row}'] = row_data[2]
            ws[f'C{current_row}'].font = label_font
            ws[f'C{current_row}'].fill = info_fill
            ws[f'C{current_row}'].border = thin_border
            ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            ws[f'D{current_row}'] = row_data[3]
            ws[f'D{current_row}'].font = value_font
            ws[f'D{current_row}'].fill = highlight_fill
            ws[f'D{current_row}'].border = thin_border
            ws[f'D{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        else:
            ws.merge_cells(f'C{current_row}:D{current_row}')
        
        ws.row_dimensions[current_row].height = 25
        current_row += 1
    
    # Address (full width)
    if doc.address:
        ws[f'A{current_row}'] = "Address:"
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].fill = info_fill
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws[f'B{current_row}'] = doc.address
        ws[f'B{current_row}'].font = value_font
        ws[f'B{current_row}'].fill = highlight_fill
        ws[f'B{current_row}'].border = thin_border
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 40
        current_row += 1
    
    # Spacing
    current_row += 1
    
    # TEAM LEADER SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "👤 TEAM LEADER"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = section_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Team leader info
    ws[f'A{current_row}'] = "Name:"
    ws[f'A{current_row}'].font = label_font
    ws[f'A{current_row}'].fill = info_fill
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = doc.full_name or ""
    ws[f'B{current_row}'].font = Font(name='Calibri', size=12, bold=True, color="667EEA")
    ws[f'B{current_row}'].fill = highlight_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    # Team size
    team_size = len(doc.visitation_team) if doc.visitation_team else 0
    ws[f'A{current_row}'] = "Team Size:"
    ws[f'A{current_row}'].font = label_font
    ws[f'A{current_row}'].fill = info_fill
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'B{current_row}'] = f"{team_size} members"
    ws[f'B{current_row}'].font = value_font
    ws[f'B{current_row}'].fill = highlight_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'C{current_row}'] = "Assigned By:"
    ws[f'C{current_row}'].font = label_font
    ws[f'C{current_row}'].fill = info_fill
    ws[f'C{current_row}'].border = thin_border
    ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'D{current_row}'] = doc.assigned_by or "System"
    ws[f'D{current_row}'].font = value_font
    ws[f'D{current_row}'].fill = highlight_fill
    ws[f'D{current_row}'].border = thin_border
    ws[f'D{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    # Spacing
    current_row += 2
    
    # FOOTER - Generation timestamp
    ws.merge_cells(f'A{current_row}:D{current_row}')
    generation_time = format_datetime(now_datetime(), "dd MMM yyyy HH:mm")
    ws[f'A{current_row}'] = f"📅 Generated on: {generation_time}"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=9, italic=True, color="95A5A6")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Add church branding footer
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "⛪ Ecclesia Church Management System"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True, color="667EEA")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')


def create_team_details_sheet(wb, doc):
    """Create beautifully formatted team details sheet"""
    
    ws = wb.create_sheet("Team Members")
    
    # Define colors
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    column_header_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=10, color="2C3E50")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # TITLE
    ws.merge_cells('A1:F1')
    ws['A1'] = "👥 VISITATION TEAM MEMBERS"
    ws['A1'].font = title_font
    ws['A1'].fill = header_gradient
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 40
    
    # Set column widths
    column_widths = [5, 20, 25, 20, 15, 30]
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    for col, width in zip(columns, column_widths):
        ws.column_dimensions[col].width = width
    
    # Headers
    headers = ["#", "Member ID", "Full Name", "Email", "Phone", "Role/Note"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[2].height = 30
    
    # Team members data
    if doc.visitation_team:
        row_idx = 3
        for idx, member in enumerate(doc.visitation_team, 1):
            fill = alt_row_fill if row_idx % 2 == 0 else white_fill
            
            data = [
                idx,
                member.member_id or "",
                member.full_name or "",
                member.email or "",
                member.phone_no or "",
                member.role or member.note or ""
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.fill = fill
                cell.border = thin_border
                
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            ws.row_dimensions[row_idx].height = 25
            row_idx += 1
    else:
        ws.merge_cells('A3:F3')
        ws['A3'] = "No team members assigned"
        ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="95A5A6")
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 30


def create_report_sheet(wb, doc):
    """Create beautifully formatted visitation report sheet"""
    
    ws = wb.create_sheet("Visitation Report")
    
    # Define colors
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    section_fill = PatternFill(start_color="FA709A", end_color="FA709A", fill_type="solid")
    content_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    section_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    content_font = Font(name='Calibri', size=10, color="2C3E50")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # Set column width
    ws.column_dimensions['A'].width = 100
    
    current_row = 1
    
    # TITLE
    ws[f'A{current_row}'] = "📝 DETAILED VISITATION REPORT"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = header_gradient
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 40
    current_row += 1
    
    # Spacing
    current_row += 1
    
    # SUMMARY SECTION
    if doc.report_summary:
        ws[f'A{current_row}'] = "📋 SUMMARY"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        ws[f'A{current_row}'] = doc.report_summary
        ws[f'A{current_row}'].font = content_font
        ws[f'A{current_row}'].fill = content_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        current_row += 1
    
    # DETAILED REPORT SECTION
    if doc.visitation_report:
        ws[f'A{current_row}'] = "📄 DETAILED REPORT"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Strip HTML tags for Excel
        import re
        clean_report = re.sub('<[^<]+?>', '', doc.visitation_report)
        clean_report = clean_report.replace('&nbsp;', ' ')
        
        ws[f'A{current_row}'] = clean_report
        ws[f'A{current_row}'].font = content_font
        ws[f'A{current_row}'].fill = content_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 120
        current_row += 1
        current_row += 1
    
    # PRAYER REQUESTS SECTION
    if doc.prayer_requests:
        ws[f'A{current_row}'] = "🙏 PRAYER REQUESTS"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        ws[f'A{current_row}'] = doc.prayer_requests
        ws[f'A{current_row}'].font = content_font
        ws[f'A{current_row}'].fill = highlight_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        current_row += 1
    
    # NEEDS IDENTIFIED SECTION
    if doc.needs_identified:
        ws[f'A{current_row}'] = "🎯 NEEDS IDENTIFIED"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        ws[f'A{current_row}'] = doc.needs_identified
        ws[f'A{current_row}'].font = content_font
        ws[f'A{current_row}'].fill = content_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        current_row += 1
    
    # FOLLOW-UP SECTION
    if doc.follow_up_required and doc.follow_up_notes:
        ws[f'A{current_row}'] = "🔄 FOLLOW-UP REQUIRED"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        ws[f'A{current_row}'] = doc.follow_up_notes
        ws[f'A{current_row}'].font = content_font
        ws[f'A{current_row}'].fill = highlight_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 60
        current_row += 1
        current_row += 1
    
    # COMPLETION INFO
    if doc.completed_date:
        ws[f'A{current_row}'] = "✅ VISITATION COMPLETED"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        completion_text = f"Completed on: {format_datetime(doc.completed_date, 'dd MMM yyyy HH:mm')}"
        ws[f'A{current_row}'] = completion_text
        ws[f'A{current_row}'].font = Font(name='Calibri', size=11, bold=True, color="2C3E50")
        ws[f'A{current_row}'].fill = success_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 25


@frappe.whitelist()
def download_excel(visitation_name):
    """
    Download Excel file directly
    Used by client-side to trigger download
    """
    result = export_visitation_to_excel(visitation_name)
    
    if result.get('success'):
        return {
            'success': True,
            'file_url': f'/api/method/church.church.doctype.visitation.visitation_excel_export.get_excel_file?visitation={visitation_name}'
        }
    else:
        frappe.throw(result.get('error', 'Failed to generate Excel file'))


@frappe.whitelist()
def get_excel_file(visitation):
    """
    Return Excel file for download
    """
    result = export_visitation_to_excel(visitation)
    
    if not result.get('success'):
        frappe.throw(result.get('error', 'Failed to generate Excel file'))
    
    # Decode base64
    file_content = base64.b64decode(result['file_content'])
    
    # Set response headers
    frappe.local.response.filename = result['filename']
    frappe.local.response.filecontent = file_content
    frappe.local.response.type = "download"


@frappe.whitelist()
def send_visitation_email(visitation_name, recipients, include_excel=1, additional_message=""):
    """
    Send visitation report via email with optional Excel attachment
    
    Args:
        visitation_name: Name of the Visitation document
        recipients: Comma-separated email addresses
        include_excel: Whether to attach Excel file
        additional_message: Additional custom message
    
    Returns:
        Dict with success status
    """
    try:
        # Get visitation document
        doc = frappe.get_doc("Visitation", visitation_name)
        
        # Parse recipients
        recipient_list = [email.strip() for email in recipients.split(',') if email.strip()]
        
        if not recipient_list:
            return {
                'success': False,
                'error': 'No valid email recipients provided'
            }
        
        # Get church settings
        settings = frappe.get_single("Church Settings")
        church_name = settings.church_name or "Our Church"
        
        # Generate email content
        email_content = get_visitation_report_email_template(doc, church_name, additional_message)
        
        # Prepare attachments
        attachments = []
        if include_excel:
            excel_result = export_visitation_to_excel(visitation_name)
            if excel_result.get('success'):
                file_content = base64.b64decode(excel_result['file_content'])
                attachments.append({
                    'fname': excel_result['filename'],
                    'fcontent': file_content
                })
        
        # Send email
        frappe.sendmail(
            recipients=recipient_list,
            subject=f"Visitation Report - {doc.visitee_full_name} ({doc.branch})",
            message=email_content,
            attachments=attachments,
            header=[f"Visitation Report - {doc.branch}", "blue"],
            delayed=False
        )
        
        # Log the email send
        frappe.db.set_value("Visitation", visitation_name, "email_sent", 1, update_modified=False)
        
        return {
            'success': True,
            'message': f'Email sent successfully to {len(recipient_list)} recipient(s)'
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), 'Send Visitation Email Error')
        return {
            'success': False,
            'error': str(e)
        }


def get_visitation_report_email_template(doc, church_name, additional_message=""):
    """Beautiful HTML email template for visitation report"""
    
    # Format date
    formatted_date = formatdate(doc.date_of_visitation, "dd MMMM yyyy")
    
    # Build additional message section
    additional_section = ""
    if additional_message:
        additional_section = f"""
            <div style="background: #E8F4F8; border-left: 5px solid #4FACFE; padding: 20px; margin: 25px 0; border-radius: 8px;">
                <p style="color: #2C3E50; margin: 0; font-size: 15px; line-height: 1.6;">
                    {additional_message}
                </p>
            </div>
        """
    
    # Build follow-up section
    follow_up_section = ""
    if doc.follow_up_required and doc.follow_up_notes:
        follow_up_section = f"""
            <div style="background: #FFF9C4; border-left: 5px solid #F39C12; padding: 20px; margin: 25px 0; border-radius: 8px;">
                <h4 style="color: #6D6D00; margin: 0 0 10px 0; font-size: 16px;">
                    🔄 Follow-up Required
                </h4>
                <p style="color: #6D6D00; margin: 0; font-size: 14px; line-height: 1.6;">
                    {doc.follow_up_notes}
                </p>
            </div>
        """
    
    # Clean HTML from report
    import re
    clean_report = doc.visitation_report or "No detailed report provided."
    clean_report = re.sub('<[^<]+?>', '', clean_report)
    clean_report = clean_report.replace('&nbsp;', ' ')
    
    email_html = f"""
    <div style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 700px; margin: 0 auto; background: #f8f9fa; padding: 0;">
        
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; text-align: center; border-radius: 10px 10px 0 0;">
            <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 700;">
                ⛪ {church_name}
            </h1>
            <p style="color: #e8f4f8; margin: 10px 0 0 0; font-size: 16px;">
                Visitation Report
            </p>
        </div>
        
        <!-- Main Content -->
        <div style="background: #ffffff; padding: 40px 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            
            <h2 style="color: #2c3e50; margin: 0 0 10px 0; font-size: 24px;">
                Visitation Report
            </h2>
            
            <p style="color: #7f8c8d; margin: 0 0 25px 0; font-size: 14px;">
                {formatted_date} • {doc.branch}
            </p>
            
            {additional_section}
            
            <!-- Visitation Summary Card -->
            <div style="background: linear-gradient(135deg, #e8f4f8 0%, #f0f8ff 100%); border-left: 5px solid #667eea; padding: 25px; margin: 25px 0; border-radius: 8px;">
                
                <h3 style="color: #667eea; margin: 0 0 20px 0; font-size: 20px; border-bottom: 2px solid #667eea; padding-bottom: 10px;">
                    📋 Visitation Details
                </h3>
                
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 10px 0; color: #7f8c8d; font-weight: 600; width: 40%;">Member Visited:</td>
                        <td style="padding: 10px 0; color: #2c3e50; font-weight: 700;">{doc.visitee_full_name}</td>
                    </tr>
                    <tr style="background: rgba(255,255,255,0.5);">
                        <td style="padding: 10px 0; color: #7f8c8d; font-weight: 600;">Type:</td>
                        <td style="padding: 10px 0; color: #2c3e50;">{doc.type}</td>
                    </tr>
                    <tr>
                        <td style="padding: 10px 0; color: #7f8c8d; font-weight: 600;">Team Leader:</td>
                        <td style="padding: 10px 0; color: #667eea; font-weight: 700;">{doc.full_name}</td>
                    </tr>
                    <tr style="background: rgba(255,255,255,0.5);">
                        <td style="padding: 10px 0; color: #7f8c8d; font-weight: 600;">Status:</td>
                        <td style="padding: 10px 0; color: #2ECC71; font-weight: 700;">{doc.status}</td>
                    </tr>
                </table>
                
            </div>
            
            <!-- Report Summary -->
            {f'''
            <div style="margin: 25px 0;">
                <h4 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 18px;">
                    📝 Summary
                </h4>
                <p style="color: #34495e; font-size: 15px; line-height: 1.7; margin: 0; padding: 20px; background: #f8f9fa; border-radius: 8px;">
                    {doc.report_summary}
                </p>
            </div>
            ''' if doc.report_summary else ''}
            
            <!-- Detailed Report -->
            <div style="margin: 25px 0;">
                <h4 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 18px;">
                    📄 Detailed Report
                </h4>
                <div style="color: #34495e; font-size: 14px; line-height: 1.7; padding: 20px; background: #f8f9fa; border-radius: 8px; white-space: pre-wrap;">
{clean_report}
                </div>
            </div>
            
            <!-- Prayer Requests -->
            {f'''
            <div style="background: #FFF3E0; border-left: 5px solid #FF9800; padding: 20px; margin: 25px 0; border-radius: 8px;">
                <h4 style="color: #E65100; margin: 0 0 10px 0; font-size: 16px;">
                    🙏 Prayer Requests
                </h4>
                <p style="color: #6D4C00; margin: 0; font-size: 14px; line-height: 1.6;">
                    {doc.prayer_requests}
                </p>
            </div>
            ''' if doc.prayer_requests else ''}
            
            <!-- Needs Identified -->
            {f'''
            <div style="background: #E3F2FD; border-left: 5px solid #2196F3; padding: 20px; margin: 25px 0; border-radius: 8px;">
                <h4 style="color: #0D47A1; margin: 0 0 10px 0; font-size: 16px;">
                    🎯 Needs Identified
                </h4>
                <p style="color: #1565C0; margin: 0; font-size: 14px; line-height: 1.6;">
                    {doc.needs_identified}
                </p>
            </div>
            ''' if doc.needs_identified else ''}
            
            {follow_up_section}
            
            <!-- Action Button -->
            <div style="text-align: center; margin: 35px 0;">
                <a href="{get_url()}/app/visitation/{doc.name}" 
                   style="display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: #ffffff; padding: 15px 40px; text-decoration: none; border-radius: 50px; font-weight: 700; font-size: 16px; box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);">
                    📋 View Full Report
                </a>
            </div>
            
        </div>
        
        <!-- Footer -->
        <div style="background: #2c3e50; padding: 25px 30px; text-align: center; border-radius: 0 0 10px 10px;">
            <p style="color: #95a5a6; margin: 0; font-size: 13px;">
                This report was generated by {church_name} Ecclesia Management System
            </p>
            <p style="color: #7f8c8d; margin: 10px 0 0 0; font-size: 12px;">
                📅 {formatdate(nowdate(), "dd MMMM yyyy")}
            </p>
        </div>
        
    </div>
    """
    
    return email_html