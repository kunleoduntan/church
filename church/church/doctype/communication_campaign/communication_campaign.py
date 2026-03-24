# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt

# -*- coding: utf-8 -*-
# Copyright (c) 2026, Church Management and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils import now_datetime, get_datetime, add_days, add_months, add_to_date
import os


class CommunicationCampaign(Document):
    def validate(self):
        """Validation before saving"""
        self.validate_recipients()
        self.validate_scheduling()
        self.validate_message_content()
        self.update_total_recipients()
    
    def validate_recipients(self):
        """Ensure recipients are present"""
        if not self.recipients:
            frappe.throw(_("Please add at least one recipient"))
    
    def validate_scheduling(self):
        """Validate scheduling logic"""
        if self.send_immediately and self.schedule:
            frappe.throw(_("Cannot select both 'Send Immediately' and 'Schedule for Later'"))
        
        if self.schedule and not self.scheduled_date:
            frappe.throw(_("Please specify Scheduled Date when scheduling campaign"))
        
        if self.is_recurring:
            if not self.recurrence_pattern:
                frappe.throw(_("Please select Recurrence Pattern for recurring campaign"))
            if not self.schedule:
                self.schedule = 1
                self.send_immediately = 0
    
    def validate_message_content(self):
        """Validate message content based on channel"""
        if not self.communication_channel:
            frappe.throw(_("Please select a Communication Channel"))
        
        # Email validation
        if "Email" in self.communication_channel:
            if not self.subject:
                frappe.throw(_("Subject is required for Email campaigns"))
            
            if self.message_format == "HTML" and not self.message_html:
                frappe.throw(_("HTML message is required when format is HTML"))
            elif self.message_format == "Text" and not self.message_body:
                frappe.throw(_("Message body is required when format is Text"))
        
        # SMS validation
        if self.communication_channel == "SMS" and not self.message_body:
            frappe.throw(_("Message body is required for SMS campaigns"))
        
        # WhatsApp validation
        if "WhatsApp" in self.communication_channel and not self.whatsapp_message_body:
            frappe.throw(_("WhatsApp message is required for WhatsApp campaigns"))
    
    def update_total_recipients(self):
        """Update total recipients count"""
        self.total_recipients = len(self.recipients)
    
    def on_submit(self):
        """Actions on submit"""
        if self.send_immediately and not self.is_test:
            # Send immediately
            self.status = "Sending"
            self.save()
            frappe.enqueue(
                send_campaign_async,
                queue='long',
                timeout=3000,
                docname=self.name
            )
        elif self.schedule:
            self.status = "Scheduled"
            self.save()


@frappe.whitelist()
def send_campaign(docname):
    """Main function to send campaign - called from button or scheduler"""
    doc = frappe.get_doc("Communication Campaign", docname)
    
    if doc.docstatus != 1:
        frappe.throw(_("Campaign must be submitted before sending"))
    
    if doc.sent and not doc.is_test:
        frappe.throw(_("Campaign has already been sent"))
    
    # Update status
    doc.status = "Sending"
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    
    # Process sending based on channel
    if "Email" in doc.communication_channel:
        send_email_campaign(doc)
    
    if doc.communication_channel == "SMS" or "SMS" in doc.communication_channel:
        send_sms_campaign(doc)
    
    if "WhatsApp" in doc.communication_channel:
        send_whatsapp_campaign(doc)
    
    # Update final status
    if doc.failed_count == 0:
        doc.status = "Completed"
        message = _("Campaign sent successfully to all recipients!")
    elif doc.sent_count > 0:
        doc.status = "Completed"
        message = _("Campaign sent with some failures. Check error logs.")
    else:
        doc.status = "Failed"
        message = _("Campaign failed to send. Check error logs.")
    
    doc.sent = 1
    doc.sent_at = now_datetime()
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    
    return message


@frappe.whitelist()
def send_test_campaign(docname, test_name, test_email=None, test_phone=None):
    """Send test campaign to specified test contacts"""
    doc = frappe.get_doc("Communication Campaign", docname)
    
    results = {
        'success': False,
        'email_sent': False,
        'sms_sent': False,
        'whatsapp_sent': False,
        'error': None
    }
    
    try:
        # Send test email
        if test_email and "Email" in doc.communication_channel:
            try:
                send_test_email(doc, test_name, test_email)
                results['email_sent'] = True
            except Exception as e:
                frappe.log_error(f"Test email error: {str(e)}", "Test Campaign Email Error")
                results['error'] = f"Email error: {str(e)}"
        
        # Send test SMS
        if test_phone and (doc.communication_channel == "SMS" or "SMS" in doc.communication_channel):
            try:
                send_sms(test_phone, doc.message_body)
                results['sms_sent'] = True
            except Exception as e:
                frappe.log_error(f"Test SMS error: {str(e)}", "Test Campaign SMS Error")
                if not results['error']:
                    results['error'] = f"SMS error: {str(e)}"
        
        # Send test WhatsApp
        if test_phone and "WhatsApp" in doc.communication_channel:
            try:
                send_whatsapp(test_phone, doc.whatsapp_message_body, doc.whatsapp_image_url)
                results['whatsapp_sent'] = True
            except Exception as e:
                frappe.log_error(f"Test WhatsApp error: {str(e)}", "Test Campaign WhatsApp Error")
                if not results['error']:
                    results['error'] = f"WhatsApp error: {str(e)}"
        
        # If at least one channel succeeded
        if results['email_sent'] or results['sms_sent'] or results['whatsapp_sent']:
            results['success'] = True
        else:
            if not results['error']:
                results['error'] = "No messages were sent. Please check your configuration."
        
        return results
        
    except Exception as e:
        frappe.log_error(f"Test campaign error: {str(e)}", "Test Campaign Error")
        results['error'] = str(e)
        return results


def send_test_email(doc, test_name, test_email):
    """Send test email"""
    # Get attachments
    files = frappe.get_all(
        "File",
        filters={"attached_to_doctype": "Communication Campaign", "attached_to_name": doc.name},
        fields=["file_url", "file_name", "is_private"]
    )
    
    image_urls = []
    attachment_files = []
    
    for f in files:
        url = f["file_url"].lower()
        if url.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
            image_urls.append(f["file_url"])
        else:
            attachment_files.append(f)
    
    # Build image tags
    image_tags = ""
    for url in image_urls:
        image_tags += f"<p style='text-align:center;'><img src='{url}' alt='Campaign Image' style='max-width:100%; height:auto; border-radius:6px;' /></p>"
    
    # Build message
    if doc.message_format == 'HTML':
        message_body = doc.message_html or ""
    else:
        message_body = doc.message_body or ""
    
    message_with_images = message_body + image_tags
    subject = doc.subject or "No Subject"
    
    # Prepare attachments
    attachments = []
    if doc.include_attachments:
        for file in attachment_files:
            file_doc = frappe.get_doc("File", {"file_url": file["file_url"]})
            file_path = frappe.get_site_path(
                "private" if file_doc.is_private else "public",
                file_doc.file_url.lstrip('/')
            )
            try:
                with open(file_path, "rb") as f:
                    attachments.append((file_doc.file_name, f.read()))
            except Exception as e:
                frappe.log_error(f"Error reading attachment: {str(e)}", "Test Campaign Attachment Error")
    
    # Render context
    context = {
        "full_name": test_name,
        "salutation": "Dear",
        "sender_name": doc.sender_name or "Church Management",
        "unsubscribe_url": "#"
    }
    rendered_message = frappe.render_template(message_with_images, context)
    
    # Send email
    frappe.sendmail(
        recipients=[test_email],
        subject=f"[TEST] {subject}",
        message=rendered_message,
        attachments=attachments,
        reference_doctype="Communication Campaign",
        reference_name=doc.name,
        template=False
    )


def send_campaign_async(docname):
    """Async wrapper for sending campaign"""
    return send_campaign(docname)


def send_email_campaign(doc):
    """Send email campaign to all recipients"""
    # Get attachments
    files = frappe.get_all(
        "File",
        filters={"attached_to_doctype": "Communication Campaign", "attached_to_name": doc.name},
        fields=["file_url", "file_name", "is_private"]
    )
    
    image_urls = []
    attachment_files = []
    
    for f in files:
        url = f["file_url"].lower()
        if url.endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
            image_urls.append(f["file_url"])
        else:
            attachment_files.append(f)
    
    # Build image tags
    image_tags = ""
    for url in image_urls:
        image_tags += f"<p style='text-align:center;'><img src='{url}' alt='Campaign Image' style='max-width:100%; height:auto; border-radius:6px;' /></p>"
    
    # Build message
    if doc.message_format == 'HTML':
        message_body = doc.message_html or ""
    else:
        message_body = doc.message_body or ""
    
    message_with_images = message_body + image_tags
    subject = doc.subject or "No Subject"
    
    # Prepare attachments
    attachments = []
    if doc.include_attachments:
        for file in attachment_files:
            file_doc = frappe.get_doc("File", {"file_url": file["file_url"]})
            file_path = frappe.get_site_path(
                "private" if file_doc.is_private else "public",
                file_doc.file_url.lstrip('/')
            )
            try:
                with open(file_path, "rb") as f:
                    attachments.append((file_doc.file_name, f.read()))
            except Exception as e:
                frappe.log_error(f"Error reading attachment: {str(e)}", "Campaign Attachment Error")
    
    # Send test email
    if doc.is_test and doc.test_email:
        context = {
            "full_name": "Test Recipient",
            "salutation": "Dear",
            "sender_name": doc.sender_name or "Church Management",
            "unsubscribe_url": "#"
        }
        rendered_message = frappe.render_template(message_with_images, context)
        
        try:
            frappe.sendmail(
                recipients=[doc.test_email],
                subject=f"{subject} (Test Email)",
                message=rendered_message,
                attachments=attachments,
                reference_doctype="Communication Campaign",
                reference_name=doc.name,
                template=False
            )
            frappe.msgprint(f"Test email sent to {doc.test_email}")
        except Exception as e:
            frappe.log_error(f"Test email error: {str(e)}", "Campaign Test Email Error")
        return
    
    # Send to all recipients
    failed = []
    sent_count = 0
    
    for recipient in doc.recipients:
        if not recipient.email:
            recipient.delivery_status = "Failed"
            recipient.error_message = "No email address"
            continue
        
        context = {
            "full_name": recipient.full_name,
            "salutation": recipient.salutation or "Dear",
            "sender_name": doc.sender_name or "Church Management",
            "unsubscribe_url": "#"
        }
        
        try:
            rendered_message = frappe.render_template(message_with_images, context)
            
            frappe.sendmail(
                recipients=[recipient.email],
                subject=subject,
                message=rendered_message,
                attachments=attachments,
                reference_doctype="Communication Campaign",
                reference_name=doc.name,
                template=False
            )
            
            recipient.delivery_status = "Sent"
            sent_count += 1
            
        except Exception as e:
            error_msg = str(e)
            failed.append(f"{recipient.email}: {error_msg}")
            recipient.delivery_status = "Failed"
            recipient.error_message = error_msg
            frappe.log_error(f"{recipient.email}: {error_msg}", "Campaign Email Error")
    
    # Update counts
    doc.sent_count = sent_count
    doc.failed_count = len(failed)
    
    if failed:
        frappe.log_error("\n".join(failed), "Campaign Partial Failure")


def send_sms_campaign(doc):
    """Send SMS campaign to all recipients"""
    # SMS sending logic - integrate with your SMS provider
    if doc.is_test and doc.test_phone_number:
        # Send test SMS
        send_sms(doc.test_phone_number, doc.message_body)
        frappe.msgprint(f"Test SMS sent to {doc.test_phone_number}")
        return
    
    sent_count = 0
    failed = []
    
    for recipient in doc.recipients:
        if not recipient.mobile_phone:
            recipient.delivery_status = "Failed"
            recipient.error_message = "No mobile phone number"
            continue
        
        try:
            # Replace with actual SMS sending logic
            send_sms(recipient.mobile_phone, doc.message_body)
            recipient.delivery_status = "Sent"
            sent_count += 1
        except Exception as e:
            error_msg = str(e)
            failed.append(f"{recipient.mobile_phone}: {error_msg}")
            recipient.delivery_status = "Failed"
            recipient.error_message = error_msg
            frappe.log_error(error_msg, "Campaign SMS Error")
    
    doc.sent_count = (doc.sent_count or 0) + sent_count
    doc.failed_count = (doc.failed_count or 0) + len(failed)


def send_whatsapp_campaign(doc):
    """Send WhatsApp campaign to all recipients"""
    # WhatsApp sending logic - integrate with your WhatsApp API
    if doc.is_test and doc.test_phone_number:
        # Send test WhatsApp
        send_whatsapp(doc.test_phone_number, doc.whatsapp_message_body, doc.whatsapp_image_url)
        frappe.msgprint(f"Test WhatsApp sent to {doc.test_phone_number}")
        return
    
    sent_count = 0
    failed = []
    
    for recipient in doc.recipients:
        if not recipient.mobile_phone:
            recipient.delivery_status = "Failed"
            recipient.error_message = "No mobile phone number"
            continue
        
        try:
            # Replace with actual WhatsApp sending logic
            send_whatsapp(recipient.mobile_phone, doc.whatsapp_message_body, doc.whatsapp_image_url)
            recipient.delivery_status = "Sent"
            sent_count += 1
        except Exception as e:
            error_msg = str(e)
            failed.append(f"{recipient.mobile_phone}: {error_msg}")
            recipient.delivery_status = "Failed"
            recipient.error_message = error_msg
            frappe.log_error(error_msg, "Campaign WhatsApp Error")
    
    doc.sent_count = (doc.sent_count or 0) + sent_count
    doc.failed_count = (doc.failed_count or 0) + len(failed)


def send_sms(phone_number, message):
    """Helper function to send SMS - uses Communication Settings"""
    settings = frappe.get_single("Communication Settings")
    
    if not settings.enable_sms or not settings.sms_enabled:
        frappe.throw(_("SMS is not enabled or properly configured in Communication Settings"))
    
    provider = settings.sms_provider
    
    try:
        if provider == "Twilio":
            return send_sms_twilio(phone_number, message, settings)
        elif provider == "Nexmo/Vonage":
            return send_sms_nexmo(phone_number, message, settings)
        elif provider == "Custom API":
            return send_sms_custom(phone_number, message, settings)
        else:
            frappe.throw(_("Invalid SMS provider selected"))
    except Exception as e:
        frappe.log_error(f"SMS Error: {str(e)}", "SMS Send Error")
        raise


def send_sms_twilio(phone_number, message, settings):
    """Send SMS via Twilio"""
    from twilio.rest import Client
    
    client = Client(settings.twilio_account_sid, settings.get_password('twilio_auth_token'))
    
    message_obj = client.messages.create(
        body=message,
        from_=settings.twilio_from_number,
        to=phone_number
    )
    
    return message_obj.sid


def send_sms_nexmo(phone_number, message, settings):
    """Send SMS via Nexmo/Vonage"""
    import vonage
    
    client = vonage.Client(
        key=settings.nexmo_api_key,
        secret=settings.get_password('nexmo_api_secret')
    )
    
    sms = vonage.Sms(client)
    
    response = sms.send_message({
        "from": settings.nexmo_from_number,
        "to": phone_number,
        "text": message
    })
    
    if response["messages"][0]["status"] == "0":
        return response["messages"][0]["message-id"]
    else:
        raise Exception(f"Nexmo Error: {response['messages'][0]['error-text']}")


def send_sms_custom(phone_number, message, settings):
    """Send SMS via Custom API"""
    import requests
    import json
    
    # Parse headers
    headers = {}
    if settings.custom_sms_headers:
        try:
            headers = json.loads(settings.custom_sms_headers)
            # Replace api_key placeholder
            headers = {k: v.replace('{api_key}', settings.get_password('custom_sms_api_key')) 
                      for k, v in headers.items()}
        except:
            frappe.log_error("Invalid SMS headers JSON", "Custom SMS Error")
    
    # Parse body template
    body = {}
    if settings.custom_sms_body_template:
        try:
            body_template = settings.custom_sms_body_template
            body_str = body_template.replace('{phone_number}', phone_number)
            body_str = body_str.replace('{message}', message)
            body_str = body_str.replace('{api_key}', settings.get_password('custom_sms_api_key'))
            body = json.loads(body_str)
        except Exception as e:
            frappe.log_error(f"Invalid SMS body template: {str(e)}", "Custom SMS Error")
            raise
    
    # Make API call
    if settings.custom_sms_method == "POST":
        response = requests.post(settings.custom_sms_api_endpoint, headers=headers, json=body)
    else:
        response = requests.get(settings.custom_sms_api_endpoint, headers=headers, params=body)
    
    if response.status_code in [200, 201]:
        return response.text
    else:
        raise Exception(f"Custom SMS API Error: {response.status_code} - {response.text}")


def send_whatsapp(phone_number, message, image_url=None):
    """Helper function to send WhatsApp - uses Communication Settings"""
    settings = frappe.get_single("Communication Settings")
    
    if not settings.enable_whatsapp or not settings.whatsapp_enabled:
        frappe.throw(_("WhatsApp is not enabled or properly configured in Communication Settings"))
    
    provider = settings.whatsapp_provider
    
    try:
        if provider == "Twilio":
            return send_whatsapp_twilio(phone_number, message, image_url, settings)
        elif provider == "WATI":
            return send_whatsapp_wati(phone_number, message, image_url, settings)
        elif provider == "Custom API":
            return send_whatsapp_custom(phone_number, message, image_url, settings)
        else:
            frappe.throw(_("Invalid WhatsApp provider selected"))
    except Exception as e:
        frappe.log_error(f"WhatsApp Error: {str(e)}", "WhatsApp Send Error")
        raise


def send_whatsapp_twilio(phone_number, message, image_url, settings):
    """Send WhatsApp via Twilio"""
    from twilio.rest import Client
    
    client = Client(settings.twilio_whatsapp_sid, settings.get_password('twilio_whatsapp_token'))
    
    msg_params = {
        'body': message,
        'from_': f'whatsapp:{settings.twilio_whatsapp_number}',
        'to': f'whatsapp:{phone_number}'
    }
    
    if image_url:
        msg_params['media_url'] = [image_url]
    
    message_obj = client.messages.create(**msg_params)
    
    return message_obj.sid


def send_whatsapp_wati(phone_number, message, image_url, settings):
    """Send WhatsApp via WATI"""
    import requests
    
    url = f"{settings.wati_api_endpoint}/api/v1/sendSessionMessage/{phone_number}"
    
    headers = {
        'Authorization': f'Bearer {settings.get_password("wati_access_token")}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'messageText': message
    }
    
    if image_url:
        payload['media'] = {
            'url': image_url
        }
    
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code == 200:
        return response.json().get('messageId', 'success')
    else:
        raise Exception(f"WATI Error: {response.status_code} - {response.text}")


def send_whatsapp_custom(phone_number, message, image_url, settings):
    """Send WhatsApp via Custom API"""
    import requests
    import json
    
    # Parse headers
    headers = {}
    if settings.custom_whatsapp_headers:
        try:
            headers = json.loads(settings.custom_whatsapp_headers)
            headers = {k: v.replace('{api_key}', settings.get_password('custom_whatsapp_api_key')) 
                      for k, v in headers.items()}
        except:
            frappe.log_error("Invalid WhatsApp headers JSON", "Custom WhatsApp Error")
    
    # Parse body template
    body = {}
    if settings.custom_whatsapp_body_template:
        try:
            body_template = settings.custom_whatsapp_body_template
            body_str = body_template.replace('{phone_number}', phone_number)
            body_str = body_str.replace('{message}', message)
            body_str = body_str.replace('{image_url}', image_url or '')
            body_str = body_str.replace('{api_key}', settings.get_password('custom_whatsapp_api_key'))
            body = json.loads(body_str)
        except Exception as e:
            frappe.log_error(f"Invalid WhatsApp body template: {str(e)}", "Custom WhatsApp Error")
            raise
    
    # Make API call
    if settings.custom_whatsapp_method == "POST":
        response = requests.post(settings.custom_whatsapp_api_endpoint, headers=headers, json=body)
    else:
        response = requests.get(settings.custom_whatsapp_api_endpoint, headers=headers, params=body)
    
    if response.status_code in [200, 201]:
        return response.text
    else:
        raise Exception(f"Custom WhatsApp API Error: {response.status_code} - {response.text}")


# ========================================
# HELPER FUNCTIONS FOR GETTING RECIPIENTS
# ========================================

@frappe.whitelist()
def get_students_from_programs_courses(programs, courses):
    """Get students from selected programs and courses"""
    import json
    
    if isinstance(programs, str):
        programs = json.loads(programs)
    if isinstance(courses, str):
        courses = json.loads(courses)
    
    filters = []
    if programs:
        filters.append(["program", "in", programs])
    if courses:
        filters.append(["program", "in", courses])  # Assuming courses are linked to programs
    
    students = frappe.get_all(
        "Student",
        fields=["name", "student_name", "student_email_id", "student_mobile_number", "salutation"],
        filters=filters,
        limit_page_length=0
    )
    
    return students


@frappe.whitelist()
def get_students_from_groups(groups):
    """Get students from selected student groups"""
    import json
    
    if isinstance(groups, str):
        groups = json.loads(groups)
    
    # Get student group members
    students_list = []
    for group in groups:
        group_members = frappe.get_all(
            "Student Group Student",
            filters={"parent": group},
            fields=["student"]
        )
        students_list.extend([s.student for s in group_members])
    
    # Remove duplicates
    students_list = list(set(students_list))
    
    # Get student details
    students = frappe.get_all(
        "Student",
        filters={"name": ["in", students_list]},
        fields=["name", "student_name", "student_email_id", "student_mobile_number", "salutation"],
        limit_page_length=0
    )
    
    return students


@frappe.whitelist()
def get_members_from_departments(departments):
    """Get members from selected church departments"""
    import json
    
    if isinstance(departments, str):
        departments = json.loads(departments)
    
    # Get department members
    members_list = []
    for dept in departments:
        dept_members = frappe.get_all(
            "Department Member",  # Assuming you have this child table
            filters={"parent": dept},
            fields=["member"]
        )
        members_list.extend([m.member for m in dept_members])
    
    # Remove duplicates
    members_list = list(set(members_list))
    
    # Get member details
    members = frappe.get_all(
        "Member",
        filters={"name": ["in", members_list]},
        fields=["name", "full_name", "email", "mobile_phone", "salutation"],
        limit_page_length=0
    )
    
    return members


# ========================================
# SCHEDULED JOB TO PROCESS SCHEDULED CAMPAIGNS
# ========================================

def process_scheduled_campaigns():
    """Process all scheduled campaigns - run every 15 minutes"""
    now = now_datetime()
    
    # Get all scheduled campaigns that are due
    campaigns = frappe.get_all(
        "Communication Campaign",
        filters={
            "schedule": 1,
            "sent": 0,
            "docstatus": 1,
            "status": "Scheduled",
            "scheduled_date": ["<=", now]
        },
        fields=["name"]
    )
    
    for campaign in campaigns:
        try:
            send_campaign(campaign.name)
        except Exception as e:
            frappe.log_error(f"Error processing campaign {campaign.name}: {str(e)}", "Scheduled Campaign Error")


def process_recurring_campaigns():
    """Process recurring campaigns - run daily"""
    today = frappe.utils.today()
    
    # Get all active recurring campaigns
    campaigns = frappe.get_all(
        "Communication Campaign",
        filters={
            "is_recurring": 1,
            "docstatus": 1,
            "status": ["in", ["Completed", "Scheduled"]]
        },
        fields=["name", "recurrence_pattern", "recurrence_interval", "recurrence_end_date", "sent_at"]
    )
    
    for campaign_info in campaigns:
        try:
            campaign = frappe.get_doc("Communication Campaign", campaign_info.name)
            
            # Check if recurrence has ended
            if campaign.recurrence_end_date and get_datetime(campaign.recurrence_end_date) < get_datetime(today):
                continue
            
            # Calculate next send date
            if campaign.sent_at:
                next_send_date = calculate_next_send_date(
                    campaign.sent_at,
                    campaign.recurrence_pattern,
                    campaign.recurrence_interval
                )
                
                if get_datetime(next_send_date).date() == get_datetime(today).date():
                    # Create a copy and send
                    new_campaign = create_recurring_campaign_copy(campaign)
                    send_campaign(new_campaign.name)
        
        except Exception as e:
            frappe.log_error(f"Error processing recurring campaign {campaign_info.name}: {str(e)}", "Recurring Campaign Error")


def calculate_next_send_date(last_sent, pattern, interval):
    """Calculate the next send date based on recurrence pattern"""
    last_sent = get_datetime(last_sent)
    
    if pattern == "Daily":
        return add_days(last_sent, interval)
    elif pattern == "Weekly":
        return add_days(last_sent, interval * 7)
    elif pattern == "Monthly":
        return add_months(last_sent, interval)
    elif pattern == "Quarterly":
        return add_months(last_sent, interval * 3)
    elif pattern == "Yearly":
        return add_to_date(last_sent, years=interval)
    
    return last_sent


def create_recurring_campaign_copy(original_campaign):
    """Create a copy of the campaign for recurring send"""
    new_campaign = frappe.copy_doc(original_campaign)
    new_campaign.sent = 0
    new_campaign.sent_at = None
    new_campaign.sent_count = 0
    new_campaign.delivered_count = 0
    new_campaign.opened_count = 0
    new_campaign.clicked_count = 0
    new_campaign.failed_count = 0
    new_campaign.status = "Sending"
    
    # Reset recipient statuses
    for recipient in new_campaign.recipients:
        recipient.delivery_status = "Pending"
        recipient.error_message = ""
    
    new_campaign.insert(ignore_permissions=True)
    new_campaign.submit()
    
    return new_campaign


# ========================================
# EXCEL IMPORT/EXPORT FUNCTIONS
# ========================================

@frappe.whitelist()
def import_recipients_from_excel(file_url, campaign_name=None):
    """Import recipients from uploaded Excel file"""
    import openpyxl
    from openpyxl import load_workbook
    import re
    
    try:
        # Get file path
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = frappe.get_site_path(
            "private" if file_doc.is_private else "public",
            file_doc.file_url.lstrip('/')
        )
        
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        recipients = []
        errors = []
        row_num = 1
        
        # Expected columns
        expected_headers = ['Full Name', 'Email', 'Mobile Phone', 'Salutation', 'Reference DocType', 'Reference Name']
        
        # Read header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Validate headers
        if not all(h in headers for h in ['Full Name']):
            return {
                'success': False,
                'error': 'Excel file must have "Full Name" column'
            }
        
        # Get column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            if header in expected_headers:
                col_indices[header] = idx
        
        # Process data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            
            if not row or all(cell is None for cell in row):
                continue
            
            try:
                # Extract values
                full_name = row[col_indices.get('Full Name', 0)] if 'Full Name' in col_indices else None
                email = row[col_indices.get('Email', 1)] if 'Email' in col_indices else None
                mobile_phone = row[col_indices.get('Mobile Phone', 2)] if 'Mobile Phone' in col_indices else None
                salutation = row[col_indices.get('Salutation', 3)] if 'Salutation' in col_indices else 'Dear'
                reference_doctype = row[col_indices.get('Reference DocType', 4)] if 'Reference DocType' in col_indices else None
                reference_name = row[col_indices.get('Reference Name', 5)] if 'Reference Name' in col_indices else None
                
                # Validate required fields
                if not full_name:
                    errors.append({
                        'row': row_num,
                        'message': 'Full Name is required'
                    })
                    continue
                
                # Clean and validate email
                if email:
                    email = str(email).strip()
                    if not validate_email(email):
                        errors.append({
                            'row': row_num,
                            'message': f'Invalid email format: {email}'
                        })
                
                # Clean phone number
                if mobile_phone:
                    mobile_phone = str(mobile_phone).strip()
                
                # Add to recipients
                recipient = {
                    'full_name': str(full_name).strip(),
                    'email': email or '',
                    'mobile_phone': mobile_phone or '',
                    'salutation': str(salutation).strip() if salutation else 'Dear',
                    'reference_doctype': str(reference_doctype).strip() if reference_doctype else '',
                    'reference_name': str(reference_name).strip() if reference_name else ''
                }
                
                recipients.append(recipient)
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'message': str(e)
                })
        
        return {
            'success': True,
            'recipients': recipients,
            'count': len(recipients),
            'errors': errors
        }
        
    except Exception as e:
        frappe.log_error(f"Excel import error: {str(e)}", "Excel Import Error")
        return {
            'success': False,
            'error': str(e)
        }


@frappe.whitelist()
def generate_excel_template(data=None):
    """Generate Excel template for recipient import"""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import json
    
    if isinstance(data, str):
        data = json.loads(data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Recipients Template"
    
    # Default template data if none provided
    if not data:
        data = [
            ['Full Name', 'Email', 'Mobile Phone', 'Salutation', 'Reference DocType', 'Reference Name'],
            ['John Doe', 'john.doe@example.com', '+1234567890', 'Mr.', 'Member', 'MEM-001'],
            ['Jane Smith', 'jane.smith@example.com', '+1234567891', 'Mrs.', 'Member', 'MEM-002'],
            ['Example User', 'example@example.com', '+1234567892', 'Dr.', '', '']
        ]
    
    # Write data
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style header row
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0369a1", end_color="0369a1", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    column_widths = {
        'A': 25,  # Full Name
        'B': 35,  # Email
        'C': 20,  # Mobile Phone
        'D': 15,  # Salutation
        'E': 20,  # Reference DocType
        'F': 20   # Reference Name
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save file
    file_name = "recipients_import_template.xlsx"
    file_path = frappe.get_site_path("private", "files", file_name)
    
    # Ensure directory exists
    import os
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    wb.save(file_path)
    
    # Create File document
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": file_name,
        "is_private": 0,
        "file_url": f"/files/{file_name}"
    })
    
    # Check if file already exists
    existing_file = frappe.db.exists("File", {"file_name": file_name})
    if existing_file:
        file_doc = frappe.get_doc("File", existing_file)
    else:
        file_doc.insert(ignore_permissions=True)
    
    return {
        'file_url': file_doc.file_url,
        'file_name': file_name
    }


def validate_email(email):
    """Validate email format"""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'  # Added closing quote and $
    return re.match(pattern, email) is not None