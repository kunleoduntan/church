# -*- coding: utf-8 -*-
# Copyright (c) 2026, Church Management and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import (
    nowdate, date_diff, getdate, formatdate, now, get_datetime,
    add_days, add_months, flt, cint, fmt_money
)
from frappe.utils.data import add_to_date
import base64
import io
import secrets
import json
import qrcode
from collections import defaultdict
from frappe.utils.file_manager import save_file
from frappe.utils import now_datetime

# Optional face recognition import
try:
    import face_recognition
    import numpy as np
    FACE_AVAILABLE = True
except Exception:
    FACE_AVAILABLE = False


class Member(Document):
    def validate(self):
        """Validate before saving - OPTIMIZED"""
        self.set_full_name()
        self.calculate_age()
        self.assign_age_category()
        self.assign_demographic_group()
        self.calculate_department_count()
        self.validate_departments()
        self.validate_contact_info()
        self.set_default_branch()
        self.track_status_changes()

    def before_save(self):
        """Before save operations"""
        self.validate_unique_primary_department()
        # NOTE: QR generation is NO LONGER called here.
        # It is handled on-demand via the whitelisted `generate_personal_qr` method
        # to avoid triggering required-field validation errors prematurely.

    def on_update(self):
        """After save operations"""
        self.update_department_members()
        self.log_significant_changes()

    def set_full_name(self):
        """Set full name from first, middle, last names"""
        parts = [self.first_name, self.middle_name, self.last_name]
        self.full_name = " ".join([p for p in parts if p])

    def calculate_age(self):
        """Calculate age from date of birth - OPTIMIZED"""
        if not self.date_of_birth:
            self.age = None
            return

        dob = getdate(self.date_of_birth)
        today = getdate()

        age = today.year - dob.year

        # Adjust if birthday hasn't occurred this year
        if (today.month, today.day) < (dob.month, dob.day):
            age -= 1

        self.age = max(0, age)

    def assign_age_category(self):
        """Assign age category based on age - SMART"""
        if not self.age:
            self.category = None
            return

        if self.age < 13:
            self.category = "Child"
        elif 13 <= self.age < 18:
            self.category = "Teenager"
        elif 18 <= self.age < 36:
            self.category = "Youth"
        elif 36 <= self.age < 60:
            self.category = "Adult"
        else:
            self.category = "Elder"

    def assign_demographic_group(self):
        """Assign demographic group based on Church Settings rules - OPTIMIZED"""
        if not self.age or not self.gender:
            return

        if not hasattr(self, '_demographic_rules'):
            self._demographic_rules = frappe.cache().get_value(
                'demographic_rules',
                generator=lambda: frappe.get_all(
                    "Demographic Group Rule",
                    filters={
                        "parenttype": "Church Settings",
                        "parent": "Church Settings"
                    },
                    fields=["member_group", "gender", "min_age", "max_age", "priority"],
                    order_by="priority desc"
                )
            )

        for rule in self._demographic_rules:
            min_age = cint(rule.min_age) if rule.min_age else 0
            max_age = cint(rule.max_age) if rule.max_age else 999

            if min_age <= self.age <= max_age:
                if rule.gender == "Both" or rule.gender == self.gender:
                    self.demographic_group = rule.member_group
                    return

    def calculate_department_count(self):
        """Calculate total active departments - SMART"""
        if not hasattr(self, 'departments') or not self.departments:
            self.department_count = 0
            return

        active_count = sum(1 for dept in self.departments if dept.is_active)
        self.department_count = active_count

    def validate_departments(self):
        """Validate department assignments - SMART"""
        if not hasattr(self, 'departments') or not self.departments:
            return

        dept_names = [d.department for d in self.departments if d.department]
        if len(dept_names) != len(set(dept_names)):
            frappe.throw(_("Cannot assign the same department multiple times"))

        for dept in self.departments:
            if dept.from_date and dept.to_date:
                if getdate(dept.from_date) > getdate(dept.to_date):
                    frappe.throw(_(f"Invalid date range for {dept.department}: From Date cannot be after To Date"))

    def validate_unique_primary_department(self):
        """Ensure only one primary department - CRITICAL"""
        if not hasattr(self, 'departments') or not self.departments:
            return

        primary_count = sum(1 for dept in self.departments if dept.is_primary and dept.is_active)

        if primary_count > 1:
            frappe.throw(_("Only one department can be marked as Primary"))

        if primary_count == 0 and any(dept.is_active for dept in self.departments):
            for dept in self.departments:
                if dept.is_active:
                    dept.is_primary = 1
                    frappe.msgprint(
                        _(f"{dept.department} has been automatically set as Primary Department"),
                        indicator='blue'
                    )
                    break

    def validate_contact_info(self):
        """Validate contact information - SMART"""
        if not self.mobile_phone and not self.email:
            frappe.msgprint(
                _("Warning: Member has no contact information (phone or email)"),
                indicator='orange',
                alert=True
            )

        if self.email:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, self.email):
                frappe.throw(_("Invalid email format"))

        if self.mobile_phone:
            self.mobile_phone = format_phone_number(self.mobile_phone)
        if self.alternative_phone:
            self.alternative_phone = format_phone_number(self.alternative_phone)

    def set_default_branch(self):
        """Set default branch from Church Settings if not set"""
        if not self.branch:
            church_settings = frappe.get_single('Church Settings')
            if church_settings.default_branch:
                self.branch = church_settings.default_branch

    def track_status_changes(self):
        """Track status changes for reporting - POWERFUL"""
        if self.is_new():
            return

        old_doc = self.get_doc_before_save()
        if not old_doc:
            return

        if old_doc.member_status != self.member_status:
            self.append('status_changes', {
                'from_status': old_doc.member_status,
                'to_status': self.member_status,
                'change_date': nowdate(),
                'changed_by': frappe.session.user,
                'remarks': f'Status changed from {old_doc.member_status} to {self.member_status}'
            })

    def update_department_members(self):
        """Update department member lists when departments change - SMART"""
        if not hasattr(self, 'departments'):
            return

        frappe.enqueue(
            'church.church.doctype.member.member.sync_member_to_departments',
            member_name=self.name,
            queue='short',
            timeout=300
        )

    def log_significant_changes(self):
        """Log significant changes for audit trail"""
        if self.is_new():
            return

        old_doc = self.get_doc_before_save()
        if not old_doc:
            return

        significant_fields = [
            'member_status', 'branch', 'demographic_group',
            'mobile_phone', 'email', 'address'
        ]

        changes = []
        for field in significant_fields:
            old_val = old_doc.get(field)
            new_val = self.get(field)
            if old_val != new_val:
                changes.append(f"{field}: {old_val} → {new_val}")

        if changes:
            frappe.logger().info(
                f"Member {self.name} updated: {', '.join(changes)}"
            )


def format_phone_number(phone):
    """Format phone number to international format"""
    if not phone:
        return None

    phone = ''.join(c for c in phone if c.isdigit() or c == '+')

    if not phone.startswith('+'):
        if len(phone) == 10 or len(phone) == 11:
            phone = '+234' + phone.lstrip('0')
        else:
            phone = '+' + phone

    return phone


def sync_member_to_departments(member_name):
    """Background job to sync member to department member lists"""
    try:
        member = frappe.get_doc('Member', member_name)

        if not hasattr(member, 'departments'):
            return

        active_depts = [
            dept.department for dept in member.departments
            if dept.is_active and dept.department
        ]

        for dept_name in active_depts:
            try:
                dept_doc = frappe.get_doc('Church Department', dept_name)
                existing = any(m.member_id == member_name for m in dept_doc.members or [])

                if not existing:
                    frappe.enqueue(
                        'church.church.doctype.church_department.church_department.fetch_department_members',
                        department_name=dept_name,
                        fetch_mode='active_only',
                        queue='short'
                    )
            except Exception as e:
                frappe.log_error(f"Failed to sync member to {dept_name}: {str(e)}")

    except Exception as e:
        frappe.log_error(f"Sync member to departments failed: {str(e)}")


# ============================================================================
# QR CODE GENERATION - FIXED
# ============================================================================

@frappe.whitelist()
def generate_personal_qr(member_name):
    """
    Generate a secure personal QR code for a member ON-DEMAND.

    KEY FIX: Uses frappe.db.set_value() instead of doc.save() so that
    required-field validation (Member Status, Gender, Parish, etc.) is
    completely bypassed. The QR is written directly to the database without
    triggering the full validation chain.

    Called from the client-side button in member.js — NOT hooked to
    before_save / after_insert to prevent premature validation errors.
    """

    # Confirm the member exists
    if not frappe.db.exists("Member", member_name):
        return {"status": "error", "message": _("Member not found")}

    # Fetch only the fields we need — no full doc.save() later
    existing = frappe.db.get_value(
        "Member",
        member_name,
        ["member_qr_code", "personal_qr_code", "qr_token"],
        as_dict=True
    )

    # Already generated — return the existing URL
    if existing.member_qr_code or existing.personal_qr_code:
        return {
            "status": "exists",
            "message": _("QR code already generated"),
            "file_url": existing.member_qr_code or existing.personal_qr_code
        }

    try:
        # Generate or reuse token
        qr_token = existing.qr_token or secrets.token_hex(12)

        qr_data = f"/api/method/church.church.doctype.member.member.qr_checkin?token={qr_token}"

        # Build QR image
        qr = qrcode.make(qr_data)
        buffer = io.BytesIO()
        qr.save(buffer, format="PNG")
        buffer.seek(0)

        # Save file — attached to the Member doctype record
        file_doc = save_file(
            fname=f"{member_name}_qr.png",
            content=buffer.getvalue(),
            dt="Member",
            dn=member_name,
            is_private=0
        )

        file_url = file_doc.file_url

        # Build update dict dynamically based on which fields exist
        field_map = frappe.get_meta("Member").get_fieldnames_with_value()
        updates = {"qr_token": qr_token}

        if "member_qr_code" in field_map:
            updates["member_qr_code"] = file_url
        if "personal_qr_code" in field_map:
            updates["personal_qr_code"] = file_url

        # Write directly to DB — skips all validation hooks entirely
        frappe.db.set_value("Member", member_name, updates, update_modified=False)
        frappe.db.commit()

        return {
            "status": "success",
            "message": _("QR code generated successfully"),
            "file_url": file_url
        }

    except Exception:
        frappe.log_error(frappe.get_traceback(), "QR Generation Failed")
        return {
            "status": "error",
            "message": _("QR code generation failed. Please check error logs.")
        }


# ============================================================================
# QR CHECKIN API
# ============================================================================

@frappe.whitelist(allow_guest=True)
def qr_checkin(token, service_type, service_date, branch):
    """Check in a member via QR code scan"""

    member = frappe.db.get_value(
        "Member",
        {"qr_token": token},
        ["name", "full_name"],
        as_dict=True
    )

    if not member:
        return {"status": "error", "message": "Invalid QR"}

    exists = frappe.db.exists(
        "Church Attendance",
        {
            "member_id": member.name,
            "service_date": service_date,
            "service_type": service_type
        }
    )

    if exists:
        return {"status": "exists", "message": "Attendance already recorded"}

    try:
        doc = frappe.get_doc({
            "doctype": "Church Attendance",
            "member_id": member.name,
            "member_name": member.full_name,
            "service_date": service_date,
            "service_type": service_type,
            "branch": branch,
            "present": 1,
            "checkin_method": "QR",
            "marked_by": frappe.session.user,
            "marked_at": now_datetime()
        })

        doc.insert(ignore_permissions=True)
        doc.submit()

        update_service_capacity(doc.service_instance)

        return {"status": "success", "member": member.full_name}

    except Exception:
        frappe.log_error(frappe.get_traceback(), "QR Attendance Failure")
        return {"status": "error", "message": "Attendance failed"}


# ============================================================================
# FACE RECOGNITION
# ============================================================================

def register_face(member, image_bytes):
    """Register a member's face encoding"""
    if not FACE_AVAILABLE:
        frappe.throw("Face recognition library not installed")

    image = face_recognition.load_image_file(image_bytes)
    encodings = face_recognition.face_encodings(image)

    if not encodings:
        frappe.throw("No face detected")

    member.face_encoding = json.dumps(encodings[0].tolist())
    member.save()


def recognize_face(image_bytes):
    """Identify a member from a face image"""
    if not FACE_AVAILABLE:
        return None

    unknown = face_recognition.load_image_file(image_bytes)
    unknown_encodings = face_recognition.face_encodings(unknown)

    if not unknown_encodings:
        return None

    unknown_encoding = unknown_encodings[0]

    members = frappe.get_all(
        "Member",
        filters={"face_encoding": ["!=", ""]},
        fields=["name", "full_name", "face_encoding"]
    )

    for m in members:
        stored = np.array(json.loads(m.face_encoding))
        match = face_recognition.compare_faces([stored], unknown_encoding)[0]
        if match:
            return m.name

    return None


# ============================================================================
# VOLUNTEER CHECKIN
# ============================================================================

def volunteer_checkin(volunteer, service_date, service_type, team=None):
    """Record volunteer attendance"""
    doc = frappe.get_doc({
        "doctype": "Volunteer Attendance",
        "volunteer": volunteer,
        "team": team,
        "service_date": service_date,
        "service_type": service_type,
        "checkin_time": now_datetime()
    })

    doc.insert(ignore_permissions=True)
    return doc.name


# ============================================================================
# SERVICE CAPACITY TRACKING
# ============================================================================

def update_service_capacity(service_name):
    """Update current attendance count and capacity percentage"""
    if not service_name:
        return

    total = frappe.db.count("Church Attendance", {"service_instance": service_name})

    capacity = frappe.db.get_value("Service", service_name, "capacity") or 0
    pct = (total / capacity * 100) if capacity else 0

    frappe.db.set_value(
        "Service",
        service_name,
        {"current_attendance": total, "capacity_pct": pct}
    )


# ============================================================================
# ANALYTICS
# ============================================================================

def get_weekly_attendance_trend(branch=None):
    """Return weekly attendance totals"""
    conditions = ""
    if branch:
        conditions = f"WHERE branch='{branch}'"

    return frappe.db.sql(
        f"""
        SELECT WEEK(service_date) week_no,
               COUNT(name) total
        FROM `tabChurch Attendance`
        {conditions}
        GROUP BY WEEK(service_date)
        ORDER BY week_no DESC
        """,
        as_dict=True
    )


def get_demographic_breakdown(service_date=None, branch=None):
    """Return attendance breakdown by demographic group"""
    conditions = ["1=1"]
    if service_date:
        conditions.append(f"service_date='{service_date}'")
    if branch:
        conditions.append(f"branch='{branch}'")

    return frappe.db.sql(
        f"""
        SELECT demographic_group,
               COUNT(name) total
        FROM `tabChurch Attendance`
        WHERE {" AND ".join(conditions)}
        GROUP BY demographic_group
        """,
        as_dict=True
    )


# ============================================================================
# WHITELISTED API FUNCTIONS
# ============================================================================

@frappe.whitelist()
def get_demographic_group(age, gender):
    """
    Returns the correct Demographic Group based on Church Settings rules.
    OPTIMIZED with caching.
    """
    try:
        age = int(age)
    except (ValueError, TypeError):
        return None

    rules = frappe.cache().get_value(
        'demographic_rules',
        generator=lambda: frappe.get_all(
            "Demographic Group Rule",
            filters={
                "parenttype": "Church Settings",
                "parent": "Church Settings"
            },
            fields=["member_group", "gender", "min_age", "max_age", "priority"],
            order_by="priority desc"
        )
    )

    for rule in rules:
        min_age = int(rule.min_age) if rule.min_age else 0
        max_age = int(rule.max_age) if rule.max_age else 999

        if min_age <= age <= max_age:
            if rule.gender == "Both" or rule.gender == gender:
                return rule.member_group

    return None


@frappe.whitelist()
def reclassify_members():
    """
    Bulk reclassify all members - OPTIMIZED.
    Runs in background for performance.
    """
    frappe.enqueue(
        'church.church.doctype.member.member.process_reclassification',
        queue='long',
        timeout=3600,
        job_name='member_reclassification'
    )

    return {
        'success': True,
        'message': _('Member reclassification started in background. You will be notified when complete.')
    }


def process_reclassification():
    """Background job for member reclassification"""
    members = frappe.get_all(
        "Member",
        fields=["name", "date_of_birth", "gender"],
        filters={'member_status': 'Active'}
    )

    updated_count = 0
    error_count = 0

    for member in members:
        if not member.date_of_birth or not member.gender:
            continue

        try:
            age = date_diff(nowdate(), member.date_of_birth) // 365
            group = get_demographic_group(age, member.gender)

            if group:
                frappe.db.set_value(
                    "Member",
                    member.name,
                    {"age": age, "demographic_group": group},
                    update_modified=False
                )
                updated_count += 1
        except Exception as e:
            error_count += 1
            frappe.log_error(
                f"Failed to reclassify {member.name}: {str(e)}",
                "Member Reclassification Error"
            )

    frappe.db.commit()

    frappe.publish_realtime(
        'member_reclassification_complete',
        {
            'updated': updated_count,
            'errors': error_count,
            'message': f'Reclassified {updated_count} members ({error_count} errors)'
        },
        user=frappe.session.user
    )

    frappe.logger().info(f"Member reclassification: {updated_count} updated, {error_count} errors")


@frappe.whitelist()
def bulk_update_members(members, field, value):
    """
    Bulk update a field for multiple members - POWERFUL.
    """
    try:
        if isinstance(members, str):
            members = json.loads(members)

        if not members:
            frappe.throw(_("No members selected"))

        allowed_fields = [
            'branch', 'member_status', 'type',
            'demographic_group', 'parish', 'area', 'zone'
        ]

        if field not in allowed_fields:
            frappe.throw(_(f"Field '{field}' is not allowed for bulk update"))

        updated = 0
        failed = 0

        for member_name in members:
            try:
                frappe.db.set_value('Member', member_name, field, value)
                updated += 1
            except Exception as e:
                failed += 1
                frappe.log_error(f"Bulk update failed for {member_name}: {str(e)}")

        frappe.db.commit()

        return {
            'success': True,
            'updated': updated,
            'failed': failed,
            'message': _(f'Updated {updated} members ({failed} failed)')
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Bulk Update Error")
        frappe.throw(_("Bulk update failed: {0}").format(str(e)))


@frappe.whitelist()
def get_member_statistics(filters=None):
    """
    Get comprehensive member statistics - POWERFUL ANALYTICS.
    """
    try:
        if filters and isinstance(filters, str):
            filters = json.loads(filters)

        if not filters:
            filters = {'member_status': 'Active'}

        members = frappe.get_all(
            'Member',
            filters=filters,
            fields=[
                'name', 'gender', 'age', 'demographic_group',
                'branch', 'member_status', 'type', 'department_count',
                'date_of_joining'
            ]
        )

        stats = {
            'total': len(members),
            'by_gender': defaultdict(int),
            'by_demographic': defaultdict(int),
            'by_branch': defaultdict(int),
            'by_status': defaultdict(int),
            'by_type': defaultdict(int),
            'age_distribution': {
                '0-12': 0, '13-17': 0, '18-35': 0, '36-59': 0, '60+': 0
            },
            'department_distribution': defaultdict(int),
            'recent_joiners': 0,
            'no_contact': 0,
            'incomplete_profiles': 0
        }

        thirty_days_ago = add_days(nowdate(), -30)

        for member in members:
            if member.gender:
                stats['by_gender'][member.gender] += 1
            if member.demographic_group:
                stats['by_demographic'][member.demographic_group] += 1
            if member.branch:
                stats['by_branch'][member.branch] += 1
            if member.member_status:
                stats['by_status'][member.member_status] += 1
            if member.type:
                stats['by_type'][member.type] += 1

            if member.age:
                if member.age <= 12:
                    stats['age_distribution']['0-12'] += 1
                elif member.age <= 17:
                    stats['age_distribution']['13-17'] += 1
                elif member.age <= 35:
                    stats['age_distribution']['18-35'] += 1
                elif member.age <= 59:
                    stats['age_distribution']['36-59'] += 1
                else:
                    stats['age_distribution']['60+'] += 1

            dept_count = member.department_count or 0
            if dept_count == 0:
                stats['department_distribution']['None'] += 1
            elif dept_count == 1:
                stats['department_distribution']['1 Department'] += 1
            elif dept_count <= 3:
                stats['department_distribution']['2-3 Departments'] += 1
            else:
                stats['department_distribution']['4+ Departments'] += 1

            if member.date_of_joining and getdate(member.date_of_joining) >= getdate(thirty_days_ago):
                stats['recent_joiners'] += 1

            member_contact = frappe.db.get_value(
                'Member', member.name, ['mobile_phone', 'email'], as_dict=1
            )
            if not member_contact.mobile_phone and not member_contact.email:
                stats['no_contact'] += 1

            if not member.age or not member.demographic_group or not member.branch:
                stats['incomplete_profiles'] += 1

        stats['by_gender'] = dict(stats['by_gender'])
        stats['by_demographic'] = dict(stats['by_demographic'])
        stats['by_branch'] = dict(stats['by_branch'])
        stats['by_status'] = dict(stats['by_status'])
        stats['by_type'] = dict(stats['by_type'])
        stats['department_distribution'] = dict(stats['department_distribution'])

        return {'success': True, 'statistics': stats}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Member Statistics Error")
        return {'success': False, 'error': str(e)}


@frappe.whitelist()
def get_member_profile_html(member_name):
    """Generate beautiful HTML profile - ENHANCED"""
    member = frappe.get_doc("Member", member_name)

    # Photo or initials avatar
    if member.photo:
        photo_html = f'<img src="{member.photo}" style="width: 150px; height: 150px; border-radius: 50%; object-fit: cover; border: 4px solid #667eea;">'
    else:
        initials = (
            (member.first_name[0] if member.first_name else "") +
            (member.last_name[0] if member.last_name else "")
        )
        photo_html = f'''
        <div style="width: 150px; height: 150px; border-radius: 50%;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    display: flex; align-items: center; justify-content: center;
                    color: white; font-size: 48px; font-weight: bold;">
            {initials}
        </div>
        '''

    # Departments table
    if hasattr(member, 'departments') and member.departments:
        dept_rows = []
        for dept in member.departments:
            status_badge = '✓ Active' if dept.is_active else '✗ Inactive'
            status_color = '#2ecc71' if dept.is_active else '#95a5a6'
            primary_badge = ' 🌟 Primary' if dept.is_primary else ''
            dept_rows.append(f"""
                <tr style="border-bottom: 1px solid #ecf0f1;">
                    <td style="padding: 10px;">{dept.department}</td>
                    <td style="padding: 10px;">
                        <span style="color: {status_color}; font-weight: bold;">{status_badge}</span>
                        {primary_badge}
                    </td>
                    <td style="padding: 10px;">{formatdate(dept.from_date) if dept.from_date else 'N/A'}</td>
                </tr>
            """)
        departments_html = f"""
            <table style="width: 100%; border-collapse: collapse;">
                <thead style="background: #f8f9fa;">
                    <tr>
                        <th style="padding: 10px; text-align: left;">Department</th>
                        <th style="padding: 10px; text-align: left;">Status</th>
                        <th style="padding: 10px; text-align: left;">Since</th>
                    </tr>
                </thead>
                <tbody>{''.join(dept_rows)}</tbody>
            </table>
        """
    else:
        departments_html = "<p style='color: #7f8c8d; font-style: italic;'>No departments assigned</p>"

    # Membership duration
    membership_duration = ""
    if member.date_of_joining:
        days = date_diff(nowdate(), member.date_of_joining)
        years = days // 365
        months = (days % 365) // 30
        membership_duration = f"{years} years, {months} months"

    # QR code section
    qr_url = getattr(member, 'member_qr_code', None) or getattr(member, 'personal_qr_code', None)
    qr_section = f'<img src="{qr_url}" style="width:120px;height:120px;">' if qr_url else \
        "<p style='color:#7f8c8d;font-style:italic;'>No QR code generated yet</p>"

    html = f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 900px; margin: 20px auto;
                background: #ffffff; border-radius: 15px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); overflow: hidden;">

        <!-- Header -->
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 40px 30px; color: white;">
            <div style="display: flex; align-items: center; gap: 30px;">
                <div>{photo_html}</div>
                <div style="flex-grow: 1;">
                    <h1 style="margin: 0 0 10px 0; font-size: 32px;">
                        {member.salutation or ''} {member.full_name}
                    </h1>
                    <p style="margin: 5px 0; opacity: 0.9; font-size: 16px;">
                        <strong>ID:</strong> {member.name} &nbsp;|&nbsp;
                        <span style="background: rgba(255,255,255,0.2); padding: 4px 12px; border-radius: 12px;">
                            {member.member_status}
                        </span>
                    </p>
                    <p style="margin: 5px 0; opacity: 0.9;">
                        {member.demographic_group or 'N/A'} &nbsp;•&nbsp; {member.age or 'N/A'} years
                    </p>
                </div>
            </div>
        </div>

        <div style="padding: 30px;">
            <!-- Quick Stats -->
            <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 30px;">
                <div style="background: linear-gradient(135deg, #3498db, #2980b9); padding: 20px; border-radius: 10px; color: white; text-align: center;">
                    <div style="font-size: 32px; font-weight: bold;">{member.department_count or 0}</div>
                    <div style="opacity: 0.9;">Departments</div>
                </div>
                <div style="background: linear-gradient(135deg, #2ecc71, #27ae60); padding: 20px; border-radius: 10px; color: white; text-align: center;">
                    <div style="font-size: 20px; font-weight: bold;">{membership_duration or 'N/A'}</div>
                    <div style="opacity: 0.9;">Membership</div>
                </div>
            </div>

            <!-- Personal Info -->
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #3498db;">
                <h3 style="margin-top: 0; color: #2980b9;">📋 Personal Information</h3>
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px;">
                    <div><strong>Gender:</strong> {member.gender or 'N/A'}</div>
                    <div><strong>Date of Birth:</strong> {formatdate(member.date_of_birth) if member.date_of_birth else 'N/A'}</div>
                    <div><strong>Marital Status:</strong> {member.marital_status or 'N/A'}</div>
                    <div><strong>Occupation:</strong> {member.occupation or 'N/A'}</div>
                </div>
            </div>

            <!-- Contact Info -->
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #e74c3c;">
                <h3 style="margin-top: 0; color: #c0392b;">📞 Contact Information</h3>
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px;">
                    <div><strong>Mobile:</strong> {member.mobile_phone or 'N/A'}</div>
                    <div><strong>Alt Phone:</strong> {member.alternative_phone or 'N/A'}</div>
                    <div style="grid-column: 1 / -1;"><strong>Email:</strong> {member.email or 'N/A'}</div>
                    <div style="grid-column: 1 / -1;"><strong>Address:</strong> {member.address or 'N/A'}</div>
                    <div><strong>City:</strong> {member.city or 'N/A'}</div>
                    <div><strong>State:</strong> {member.state or 'N/A'}</div>
                </div>
            </div>

            <!-- Church Info -->
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #9b59b6;">
                <h3 style="margin-top: 0; color: #8e44ad;">⛪ Church Information</h3>
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px;">
                    <div><strong>Branch:</strong> {member.branch or 'N/A'}</div>
                    <div><strong>Parish:</strong> {member.parish or 'N/A'}</div>
                    <div><strong>Date Joined:</strong> {formatdate(member.date_of_joining) if member.date_of_joining else 'N/A'}</div>
                    <div><strong>Member Type:</strong> {member.type or 'N/A'}</div>
                </div>
            </div>

            <!-- Departments -->
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #f39c12;">
                <h3 style="margin-top: 0; color: #d68910;">🏢 Departments & Ministries ({member.department_count or 0})</h3>
                {departments_html}
            </div>

            <!-- QR Code -->
            <div style="background: #f8f9fa; padding: 25px; border-radius: 10px; border-left: 4px solid #1abc9c; text-align: center;">
                <h3 style="margin-top: 0; color: #16a085;">🔲 Member QR Code</h3>
                {qr_section}
            </div>

            <!-- Footer -->
            <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 2px solid #ecf0f1; color: #7f8c8d; font-size: 12px;">
                <p>Generated on {formatdate(nowdate())} at {now()}</p>
                <p>⛪ Church Management System</p>
            </div>
        </div>
    </div>
    """

    return html


@frappe.whitelist()
def export_members_to_excel(filters=None):
    """Export members to Excel - ENHANCED with filters"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.throw(_("Excel export requires 'openpyxl' library"))

    if filters and isinstance(filters, str):
        filters = json.loads(filters)

    if not filters:
        filters = {}

    members = frappe.get_all(
        "Member",
        filters=filters,
        fields=[
            "name", "salutation", "first_name", "middle_name", "last_name", "full_name",
            "gender", "date_of_birth", "age", "marital_status", "occupation",
            "mobile_phone", "alternative_phone", "email", "address", "city", "state",
            "branch", "parish", "area", "zone", "date_of_joining", "member_status",
            "type", "demographic_group", "department_count"
        ],
        order_by="full_name asc"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members List"

    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    headers = [
        "Member ID", "Full Name", "Gender", "Age", "DOB", "Marital Status",
        "Mobile", "Email", "Branch", "Parish", "Status", "Type",
        "Demographic", "Departments", "Date Joined"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    for row_idx, member in enumerate(members, 2):
        fill = alt_row_fill if row_idx % 2 == 0 else PatternFill()
        data = [
            member.name,
            member.full_name or "",
            member.gender or "",
            member.age or "",
            formatdate(member.date_of_birth) if member.date_of_birth else "",
            member.marital_status or "",
            member.mobile_phone or "",
            member.email or "",
            member.branch or "",
            member.parish or "",
            member.member_status or "",
            member.type or "",
            member.demographic_group or "",
            member.department_count or 0,
            formatdate(member.date_of_joining) if member.date_of_joining else ""
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = fill

    ws.freeze_panes = 'A2'

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    file_content = base64.b64encode(excel_file.read()).decode('utf-8')
    filename = f"Members_Export_{formatdate(nowdate(), 'yyyy-MM-dd')}.xlsx"

    return {'success': True, 'file_content': file_content, 'filename': filename}


@frappe.whitelist()
def send_birthday_wishes():
    """Send birthday wishes - runs daily via scheduler"""
    today = getdate(nowdate())
    church_settings = frappe.get_single('Church Settings')

    members = frappe.db.sql("""
        SELECT name, first_name, full_name, salutation, email, mobile_phone,
               gender, date_of_birth, branch, age
        FROM `tabMember`
        WHERE date_of_birth IS NOT NULL
        AND DAY(date_of_birth) = %(day)s
        AND MONTH(date_of_birth) = %(month)s
        AND member_status = 'Active'
    """, {"day": today.day, "month": today.month}, as_dict=True)

    if not members:
        return {'success': True, 'sent': 0}

    sent = 0
    failed = 0

    for member in members:
        try:
            age = member.age if member.age else calculate_age_from_dob(member.date_of_birth)
            message_html = generate_birthday_html(member, church_settings, age)

            if member.email:
                frappe.sendmail(
                    recipients=[member.email],
                    subject=f"🎉 Happy Birthday {member.first_name}!",
                    message=message_html,
                    reference_doctype="Member",
                    reference_name=member.name
                )
                sent += 1

        except Exception as e:
            failed += 1
            frappe.log_error(f"Birthday wish failed for {member.full_name}: {str(e)}")

    frappe.db.commit()
    return {'success': True, 'sent': sent, 'failed': failed}


@frappe.whitelist()
def send_birthday_wishes_manual(member_name):
    """Manually send a birthday wish to a specific member"""
    try:
        member = frappe.get_doc("Member", member_name)
        church_settings = frappe.get_single('Church Settings')

        if not member.email:
            return {'success': False, 'message': _('No email address found for this member')}

        age = member.age or calculate_age_from_dob(member.date_of_birth)
        message_html = generate_birthday_html(member, church_settings, age)

        frappe.sendmail(
            recipients=[member.email],
            subject=f"🎉 Happy Birthday {member.first_name}!",
            message=message_html,
            reference_doctype="Member",
            reference_name=member.name
        )

        return {'success': True, 'message': _('Birthday wish sent successfully')}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Manual Birthday Wish Failed")
        return {'success': False, 'message': str(e)}


def calculate_age_from_dob(dob):
    """Helper to calculate age from date of birth"""
    if not dob:
        return None
    return date_diff(nowdate(), dob) // 365


def generate_birthday_html(member, settings, age):
    """Generate beautiful birthday email HTML"""
    church_name = settings.church_name if hasattr(settings, 'church_name') and settings.church_name else "Church"

    return f"""
    <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px; margin: 0 auto;
                background: linear-gradient(135deg, #FF6B9D, #C06C84); padding: 40px 20px; border-radius: 15px;">
        <div style="text-align: center; color: white;">
            <h1 style="font-size: 48px; margin: 0;">🎉🎂🎈</h1>
            <h2 style="margin: 15px 0; font-size: 32px;">Happy {age}th Birthday!</h2>
            <h3 style="margin: 10px 0; font-size: 24px;">{member.first_name}</h3>
        </div>
        <div style="background: white; padding: 30px; border-radius: 12px; margin: 25px 0; color: #2c3e50;">
            <p style="font-size: 16px; line-height: 1.8; margin: 0;">
                May this special day bring you endless joy and God's abundant blessings! 🌟
                <br><br>
                Wishing you another amazing year filled with love, laughter, and divine favor.
                <br><br>
                <strong>With love,<br>{church_name}</strong>
            </p>
        </div>
        <div style="text-align: center;">
            <p style="color: white; opacity: 0.9; font-size: 14px;">🎁 Enjoy your special day!</p>
        </div>
    </div>
    """