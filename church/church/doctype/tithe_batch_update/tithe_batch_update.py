# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt


from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils import flt, getdate, nowdate, now_datetime, get_datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
import io
from datetime import datetime

class TitheBatchUpdate(Document):
    def validate(self):
        """Validate the document before saving"""
        self.validate_dates()
        self.validate_transactions()
        self.calculate_totals()
    
    def before_submit(self):
        """Final validation before submission"""
        if not self.tithe_transaction_entry:
            frappe.throw(_("Cannot submit batch without any transactions"))
        
        self.validate_no_duplicate_members()
        self.set_receipt_status("Pending")
    
    def on_submit(self):
        """Actions to perform on submit"""
        # Generate receipts automatically
        if frappe.db.get_single_value("Church Settings", "auto_generate_receipts_on_batch_submit"):
            frappe.enqueue(
                generate_receipts,
                queue='long',
                timeout=3000,
                docname=self.name,
                is_async=True
            )
    
    def on_cancel(self):
        """Cancel all linked receipts"""
        self.cancel_linked_receipts()
        self.set_receipt_status("")
    
    def validate_dates(self):
        """Validate date ranges"""
        if self.reporting_date and self.start_date:
            if getdate(self.reporting_date) < getdate(self.start_date):
                frappe.throw(_("Reporting date cannot be earlier than start date"))
        
        if self.start_date and self.end_date:
            if getdate(self.start_date) > getdate(self.end_date):
                frappe.throw(_("Start date cannot be later than end date"))
    
    def validate_transactions(self):
        """Validate all transaction entries"""
        if not self.tithe_transaction_entry:
            return
        
        for idx, row in enumerate(self.tithe_transaction_entry, start=1):
            if not row.member_id:
                frappe.throw(_("Row {0}: Member ID is required").format(idx))
            
            if not row.date:
                frappe.throw(_("Row {0}: Date is required").format(idx))
            
            if not row.type:
                frappe.throw(_("Row {0}: Source type is required").format(idx))
            
            if flt(row.amount_paid) <= 0:
                frappe.throw(_("Row {0}: Amount must be greater than zero").format(idx))
            
            # Validate member exists
            if not frappe.db.exists("Member", row.member_id):
                frappe.throw(_("Row {0}: Member {1} does not exist").format(idx, row.member_id))
    
    def validate_no_duplicate_members(self):
        """Check for duplicate member entries in the same batch"""
        member_list = []
        for row in self.tithe_transaction_entry:
            if row.member_id in member_list:
                frappe.msgprint(
                    _("Warning: Member {0} appears multiple times in this batch").format(row.member_id),
                    indicator='orange',
                    alert=True
                )
            member_list.append(row.member_id)
    
    def calculate_totals(self):
        """Calculate total amounts"""
        total_amount = 0
        total_workers = 0
        total_members = 0
        
        for row in self.tithe_transaction_entry:
            # Calculate base currency amount
            amount_in_base = flt(row.amount_paid) * flt(row.exchange_rate, 1)
            row.total_base_currency = amount_in_base
            
            # Calculate worker/member tithe
            if row.type == "Worker":
                row.worker_tithe = amount_in_base
                row.member_tithe = 0
                total_workers += amount_in_base
            elif row.type == "Member":
                row.member_tithe = amount_in_base
                row.worker_tithe = 0
                total_members += amount_in_base
            
            total_amount += amount_in_base
        
        self.total_transactions = len(self.tithe_transaction_entry)
        self.amount_paid = total_amount
        self.total_for_workers = total_workers
        self.total_for_members = total_members
    
    def set_receipt_status(self, status):
        """Update receipt generation status"""
        self.db_set('receipt_status', status, update_modified=False)
    
    def cancel_linked_receipts(self):
        """Cancel all receipts created from this batch"""
        receipts = frappe.get_all(
            "Receipts",
            filters={"tithe_batch_reference": self.name, "docstatus": 1},
            pluck="name"
        )
        
        for receipt_name in receipts:
            try:
                receipt = frappe.get_doc("Receipts", receipt_name)
                receipt.cancel()
            except Exception as e:
                frappe.log_error(
                    message=frappe.get_traceback(),
                    title=f"Error cancelling receipt {receipt_name}"
                )


@frappe.whitelist()
def generate_receipts(docname):
    """
    Generate receipts for all transactions in the batch
    Returns a summary of success and failures
    """
    doc = frappe.get_doc("Tithe Batch Update", docname)
    
    if doc.docstatus != 1:
        frappe.throw(_("Batch must be submitted before generating receipts"))
    
    # Update status
    doc.set_receipt_status("In Progress")
    doc.db_set('last_receipt_generation', now_datetime(), update_modified=False)
    
    success_count = 0
    failure_count = 0
    error_log = []
    
    # Get company default currency
    company = frappe.defaults.get_user_default("Company")
    local_currency = frappe.db.get_value('Company', company, 'default_currency') or 'NGN'
    
    for idx, row in enumerate(doc.tithe_transaction_entry, start=1):
        try:
            # Check if receipt already exists
            existing_receipt = frappe.db.get_value(
                "Receipts",
                {"referenced_document_no": row.name, "docstatus": ["!=", 2]},
                "name"
            )
            
            if existing_receipt:
                row.db_set('receipt_reference', existing_receipt, update_modified=False)
                error_log.append(f"Row {idx}: Receipt already exists ({existing_receipt})")
                continue
            
            # Create receipt
            receipt = create_single_receipt(doc, row, local_currency)
            
            if receipt:
                row.db_set('receipt_reference', receipt.name, update_modified=False)
                success_count += 1
            else:
                failure_count += 1
                error_log.append(f"Row {idx}: Failed to create receipt")
        
        except Exception as e:
            failure_count += 1
            error_msg = f"Row {idx} ({row.member_id}): {str(e)}"
            error_log.append(error_msg)
            frappe.log_error(
                message=frappe.get_traceback(),
                title=f"Receipt Generation Error - {doc.name} Row {idx}"
            )
    
    # Update batch status
    doc.db_set('receipts_created', success_count, update_modified=False)
    doc.db_set('receipts_failed', failure_count, update_modified=False)
    
    if failure_count == 0:
        doc.set_receipt_status("Completed")
    elif success_count == 0:
        doc.set_receipt_status("Failed")
    else:
        doc.set_receipt_status("Partially Failed")
    
    # Save error log
    if error_log:
        log_text = "\n".join(error_log)
        doc.db_set('generation_log', log_text, update_modified=False)
    
    frappe.db.commit()
    
    return {
        "success": failure_count == 0,
        "message": _("Generated {0} receipts successfully. {1} failed.").format(success_count, failure_count),
        "success_count": success_count,
        "failure_count": failure_count,
        "errors": error_log if error_log else None
    }


def create_single_receipt(batch_doc, transaction_row, local_currency):
    """Create a single receipt from a transaction row"""
    
    receipt = frappe.new_doc("Receipts")
    
    # Basic information
    receipt.naming_series = 'REC-.YYYY.-.MM.-.#####'
    receipt.transaction_date = transaction_row.date
    receipt.transaction_type = 'Member Tithe'
    receipt.member_id = transaction_row.member_id
    receipt.branch = transaction_row.branch or batch_doc.branch
    receipt.source = transaction_row.type
    receipt.transaction_purposes = f"Member's Tithe - {transaction_row.type} ({batch_doc.name})"
    
    # Financial information
    receipt.receipt_currency = transaction_row.currency
    receipt.exchange_rate = flt(transaction_row.exchange_rate, 1)
    
    # Amount handling based on currency
    if transaction_row.currency == local_currency:
        receipt.amount_paid = flt(transaction_row.amount_paid)
        receipt.amount_paid_in_fc = 0
    else:
        receipt.amount_paid = 0
        receipt.amount_paid_in_fc = flt(transaction_row.amount_paid)
    
    # Payment details
    receipt.mode_of_payment = "Wire Transfer"
    receipt.reference_no = transaction_row.name.upper()
    receipt.referenced_document_no = transaction_row.name.upper()
    receipt.remittance_bank = transaction_row.designated_bank_acct or batch_doc.designated_bank_acct
    
    # Link back to batch
    receipt.tithe_batch_reference = batch_doc.name
    
    # Accounting
    receipt.create_accounting_entries = 1
    
    # Additional details
    if transaction_row.other_details:
        receipt.remarks = transaction_row.other_details
    
    # Save with flags to bypass some validations if needed
    receipt.flags.ignore_permissions = True
    receipt.insert()
    
    # Auto-submit if configured
    if frappe.db.get_single_value("Church Settings", "auto_submit_receipts_from_batch"):
        receipt.submit()
    
    receipt.flags.ignore_permissions = False
    
    return receipt


@frappe.whitelist()
def download_import_template(branch=None, currency=None):
    """Generate and download an Excel template for importing tithe transactions"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tithe Transactions"
    
    # Define headers
    headers = [
        "Member ID*",
        "Full Name",
        "Date*",
        "Source*",
        "Amount Paid*",
        "Exchange Rate",
        "Currency*",
        "Other Details"
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add instructions row
    instructions = [
        "Enter member ID",
        "Auto-filled from member",
        "Format: YYYY-MM-DD",
        "Worker or Member",
        "Numeric value",
        "Default: 1.0",
        currency or "NGN",
        "Optional notes"
    ]
    
    for col_num, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = instruction
        cell.font = Font(italic=True, color="666666", size=9)
        cell.alignment = Alignment(horizontal='left')
    
    # Add sample data rows
    sample_data = [
        ["MBR-0001", "John Doe", "2024-01-28", "Member", 50000, 1.0, currency or "NGN", "January tithe"],
        ["WKR-0001", "Jane Smith", "2024-01-28", "Worker", 75000, 1.0, currency or "NGN", ""]
    ]
    
    for row_num, row_data in enumerate(sample_data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Format numeric columns
            if col_num in [5, 6]:
                cell.number_format = '#,##0.00'
    
    # Set column widths
    column_widths = [15, 25, 12, 12, 15, 15, 12, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions_text = [
        ["Tithe Import Template Instructions", ""],
        ["", ""],
        ["Required Fields (marked with *):", ""],
        ["Member ID", "Must be a valid member ID from your system"],
        ["Date", "Format: YYYY-MM-DD (e.g., 2024-01-28)"],
        ["Source", "Must be either 'Worker' or 'Member'"],
        ["Amount Paid", "Numeric value greater than zero"],
        ["Currency", "Three-letter currency code (e.g., NGN, USD)"],
        ["", ""],
        ["Optional Fields:", ""],
        ["Exchange Rate", "Defaults to 1.0 if not provided"],
        ["Other Details", "Any additional notes or remarks"],
        ["", ""],
        ["Tips:", ""],
        ["1. Full Name will be auto-filled when you import", ""],
        ["2. Remove the sample data rows before importing", ""],
        ["3. Ensure all member IDs exist in your system", ""],
        ["4. Date should be within the batch date range", ""],
        ["5. Use consistent currency across all transactions", ""]
    ]
    
    for row_num, row_data in enumerate(instructions_text, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_num, column=col_num)
            cell.value = value
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="4472C4")
            elif col_num == 1 and value:
                cell.font = Font(bold=True)
    
    ws_instructions.column_dimensions['A'].width = 40
    ws_instructions.column_dimensions['B'].width = 50
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Encode to base64
    file_content = base64.b64encode(output.read()).decode('utf-8')
    
    filename = f"Tithe_Import_Template_{nowdate()}.xlsx"
    
    return {
        "file_content": file_content,
        "filename": filename
    }


@frappe.whitelist()
def import_from_excel(docname, file_url):
    """Import tithe transactions from an uploaded Excel file"""
    
    doc = frappe.get_doc("Tithe Batch Update", docname)
    
    if doc.docstatus != 0:
        frappe.throw(_("Cannot import to a submitted batch"))
    
    # Get file content
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    imported_count = 0
    error_list = []
    
    # Skip header and instruction rows
    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip empty rows
        if not any(row):
            continue
        
        try:
            member_id = row[0]
            date = row[2]
            source = row[3]
            amount_paid = row[4]
            exchange_rate = row[5] or 1.0
            currency = row[6]
            other_details = row[7]
            
            # Validate required fields
            if not member_id or not date or not source or not amount_paid:
                error_list.append(f"Row {row_num}: Missing required fields")
                continue
            
            # Validate member exists
            if not frappe.db.exists("Member", member_id):
                error_list.append(f"Row {row_num}: Member {member_id} does not exist")
                continue
            
            # Parse date
            if isinstance(date, str):
                try:
                    date = datetime.strptime(date, "%Y-%m-%d").date()
                except:
                    error_list.append(f"Row {row_num}: Invalid date format")
                    continue
            
            # Get member details
            member = frappe.get_doc("Member", member_id)
            
            # Add transaction row
            doc.append("tithe_transaction_entry", {
                "member_id": member_id,
                "full_name": member.full_name,
                "date": date,
                "type": source,
                "amount_paid": flt(amount_paid),
                "exchange_rate": flt(exchange_rate, 1),
                "currency": currency or doc.currency,
                "branch": doc.branch,
                "designated_bank_acct": doc.designated_bank_acct,
                "other_details": other_details
            })
            
            imported_count += 1
        
        except Exception as e:
            error_list.append(f"Row {row_num}: {str(e)}")
            continue
    
    # Save the document
    if imported_count > 0:
        doc.save()
    
    return {
        "imported": imported_count,
        "errors": error_list if error_list else None
    }


@frappe.whitelist()
def export_to_excel(docname):
    """Export tithe batch transactions to Excel"""
    
    doc = frappe.get_doc("Tithe Batch Update", docname)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tithe Transactions"
    
    # Style definitions
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="4472C4")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Tithe Collection Report - {doc.name}"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add batch info
    info_row = 2
    ws.cell(row=info_row, column=1).value = "Branch:"
    ws.cell(row=info_row, column=2).value = doc.branch
    ws.cell(row=info_row, column=3).value = "Period:"
    ws.cell(row=info_row, column=4).value = f"{doc.start_date} to {doc.end_date}"
    ws.cell(row=info_row, column=5).value = "Currency:"
    ws.cell(row=info_row, column=6).value = doc.currency
    
    # Headers
    headers = [
        "Member ID",
        "Full Name",
        "Date",
        "Source",
        "Amount",
        "Exchange Rate",
        "Worker Tithe",
        "Member Tithe",
        "Total (Base)",
        "Receipt Ref"
    ]
    
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    data_start_row = 5
    for row_num, transaction in enumerate(doc.tithe_transaction_entry, start=data_start_row):
        ws.cell(row=row_num, column=1).value = transaction.member_id
        ws.cell(row=row_num, column=2).value = transaction.full_name
        ws.cell(row=row_num, column=3).value = transaction.date
        ws.cell(row=row_num, column=4).value = transaction.type
        ws.cell(row=row_num, column=5).value = transaction.amount_paid
        ws.cell(row=row_num, column=6).value = transaction.exchange_rate
        ws.cell(row=row_num, column=7).value = transaction.worker_tithe
        ws.cell(row=row_num, column=8).value = transaction.member_tithe
        ws.cell(row=row_num, column=9).value = transaction.total_base_currency
        ws.cell(row=row_num, column=10).value = transaction.receipt_reference or ""
        
        # Apply border and format numbers
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num in [5, 6, 7, 8, 9]:
                cell.number_format = '#,##0.00'
    
    # Add totals row
    total_row = data_start_row + len(doc.tithe_transaction_entry)
    ws.cell(row=total_row, column=4).value = "TOTALS:"
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=7).value = doc.total_for_workers
    ws.cell(row=total_row, column=8).value = doc.total_for_members
    ws.cell(row=total_row, column=9).value = doc.amount_paid
    
    for col_num in [7, 8, 9]:
        cell = ws.cell(row=total_row, column=col_num)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.border = border
    
    # Set column widths
    column_widths = [15, 25, 12, 12, 15, 12, 15, 15, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Encode to base64
    file_content = base64.b64encode(output.read()).decode('utf-8')
    
    filename = f"{doc.name}_Tithe_Collection_{nowdate()}.xlsx"
    
    return {
        "file_content": file_content,
        "filename": filename
    }