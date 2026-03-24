# -*- coding: utf-8 -*-
# Copyright (c) 2025, Value Impacts Consulting and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import flt, getdate, get_first_day, get_last_day, add_days, nowdate
from frappe.model.document import Document

class OfferingSheet(Document):
    def validate(self):
        """Validate offering sheet before saving"""
        self.validate_dates()
        self.set_month_from_date()
        self.calculate_totals()
        self.validate_branch_consistency()
        
    def validate_dates(self):
        """Ensure date range is valid"""
        if self.date_from and self.date_to:
            if getdate(self.date_from) > getdate(self.date_to):
                frappe.throw(_("Date From cannot be after Date To"))
                
    def set_month_from_date(self):
        """Auto-set month from reporting date"""
        if self.reporting_date:
            reporting_date = getdate(self.reporting_date)
            self.month = reporting_date.strftime('%B %Y')
            
    def validate_branch_consistency(self):
        """Ensure all offering entries have the same branch"""
        if self.branch and self.offering:
            for row in self.offering:
                row.branch = self.branch
                
    def calculate_totals(self):
        """Calculate total amounts from offering entries"""
        total = 0
        for row in self.offering:
            if row.amount_in_lc:
                total += flt(row.amount_in_lc)
        self.total = total
        
    def on_submit(self):
        """Create receipt vouchers on submit"""
        self.create_receipt_vouchers()
        
    def create_receipt_vouchers(self):
        """Create individual receipt vouchers for each offering entry"""
        if not self.offering:
            frappe.throw(_("No offering entries to process"))
            
        created_receipts = []
        
        for row in self.offering:
            # Skip if receipt already exists
            if frappe.db.exists("Receipts", {"referenced_document_no": row.name}):
                frappe.msgprint(_("Receipt for {0} already exists").format(row.name))
                continue
                
            # Validate offering type
            if row.offering_type in ['Pledges', 'Tithe offering']:
                frappe.throw(_("Cannot process Pledges and Tithe offerings through Offering Sheet"))
                
            try:
                receipt = self._create_receipt(row)
                created_receipts.append(receipt.name)
            except Exception as e:
                frappe.log_error(frappe.get_traceback(), f"Offering Sheet Receipt Creation Error - {row.name}")
                frappe.throw(_("Error creating receipt for {0}: {1}").format(row.name, str(e)))
                
        if created_receipts:
            frappe.msgprint(_("Successfully created {0} receipt(s): {1}").format(
                len(created_receipts), 
                ", ".join(created_receipts)
            ), alert=True)
            
    def _create_receipt(self, row):
        """Create a single receipt voucher"""
        purpose = f"Member's Offering from {row.programme} for {row.offering_type} - {row.day}"
        
        receipt = frappe.get_doc({
            "doctype": "Receipts",
            "naming_series": "REC-",
            "transaction_date": row.date,
            "transaction_type": "Revenue",
            "received_from": f"Member's Offering from {row.programme}",
            "branch": row.branch or self.branch,
            "create_accounting_entries": 1,
            "source": row.offering_type,
            "transaction_purposes": purpose,
            "receipt_currency": row.currency or self.currency,
            "exchange_rate": 1 if row.currency == "NGN" else flt(row.exchange_rate, 2),
            "amount_paid": row.amount_paid if row.currency == "NGN" else 0,
            "amount_paid_in_fc": row.amount_paid if row.currency != "NGN" else 0,
            "account_to_credit": row.income_account,
            "mode_of_payment": "Cash",
            "reference_no": row.name.upper(),
            "referenced_document_no": row.name.upper(),
            "remittance_bank": self.bank_account
        })
        
        receipt.flags.ignore_permissions = True
        receipt.insert()
        
        return receipt


@frappe.whitelist()
def auto_create_offering_sheets():
    """
    Scheduled function to create offering sheets every Sunday for all branches.
    Add to Scheduled Job Type: Weekly on Sunday
    """
    from datetime import datetime
    
    today = getdate(nowdate())
    
    # Check if today is Sunday (weekday 6)
    if today.weekday() != 6:
        return
        
    branches = frappe.get_all("Branch", filters={"disabled": 0}, pluck="name")
    
    created_sheets = []
    
    for branch in branches:
        # Check if sheet already exists for this week
        existing = frappe.db.exists("Offering Sheet", {
            "branch": branch,
            "reporting_date": today
        })
        
        if existing:
            continue
            
        try:
            sheet = create_weekly_offering_sheet(branch, today)
            created_sheets.append(sheet.name)
        except Exception as e:
            frappe.log_error(frappe.get_traceback(), f"Auto Create Offering Sheet - {branch}")
            
    if created_sheets:
        frappe.publish_realtime(
            "msgprint",
            _("Created {0} offering sheets: {1}").format(len(created_sheets), ", ".join(created_sheets))
        )
        
    return created_sheets


def create_weekly_offering_sheet(branch, reporting_date):
    """Create an offering sheet for a specific branch and week"""
    date_from = add_days(reporting_date, -6)  # Previous Sunday
    date_to = reporting_date
    
    # Get default bank account for branch
    bank_account = frappe.db.get_value("Branch", branch, "default_bank_account")
    
    sheet = frappe.get_doc({
        "doctype": "Offering Sheet",
        "naming_series": "OFF-SHEET-.YYYY.-",
        "reporting_date": reporting_date,
        "date_from": date_from,
        "date_to": date_to,
        "branch": branch,
        "bank_account": bank_account,
        "prepared_by": frappe.session.user
    })
    
    sheet.insert(ignore_permissions=True)
    
    return sheet


@frappe.whitelist()
def get_offering_template():
    """Generate Excel template for bulk offering import"""
    from frappe.utils.xlsxutils import make_xlsx
    
    data = [
        ["Date", "Programme", "Offering Type", "Amount Paid", "Currency", "Exchange Rate", "Note"],
        ["2025-02-02", "Sunday Service", "First Fruit", "50000", "NGN", "1.00", "Sample entry"],
        ["2025-02-02", "Sunday Service", "Thanksgiving", "25000", "NGN", "1.00", ""],
    ]
    
    xlsx_file = make_xlsx(data, "Offering Import Template")
    
    frappe.response['filename'] = 'offering_import_template.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def export_offering_sheet(offering_sheet_name):
    """Export offering sheet to colorful Excel format"""
    from frappe.utils.xlsxutils import make_xlsx
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    doc = frappe.get_doc("Offering Sheet", offering_sheet_name)
    
    # Prepare data
    data = [
        [f"OFFERING SHEET - {doc.branch}"],
        [f"Reporting Period: {doc.date_from} to {doc.date_to}"],
        [f"Month: {doc.month}"],
        [""],
        ["Date", "Day", "Programme", "Offering Type", "Currency", "Amount", "Exchange Rate", "Amount (LC)", "Note"]
    ]
    
    for row in doc.offering:
        data.append([
            row.date,
            row.day,
            row.programme,
            row.offering_type,
            row.currency or "",
            flt(row.amount_paid, 2),
            flt(row.exchange_rate, 2),
            flt(row.amount_in_lc, 2),
            row.note or ""
        ])
        
    data.append(["", "", "", "", "", "", "TOTAL:", flt(doc.total, 2), ""])
    
    # Create Excel with styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offering Sheet"
    
    # Add data
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Header styling
            if row_idx <= 3:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Column headers
            elif row_idx == 5:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Data rows - alternating colors
            elif row_idx > 5 and row_idx < len(data):
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
            # Total row
            elif row_idx == len(data):
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                
            # Borders for all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    frappe.response['filename'] = f'offering_sheet_{doc.name}.xlsx'
    frappe.response['filecontent'] = excel_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def email_offering_sheet(offering_sheet_name, recipients=None, cc=None, subject=None, message=None, attach_excel=True):
    """
    Email offering sheet with beautiful HTML template and optional Excel attachment
    
    Args:
        offering_sheet_name: Name of the Offering Sheet document
        recipients: Comma-separated email addresses or list
        cc: Comma-separated CC email addresses or list
        subject: Email subject (optional, will use default if not provided)
        message: Additional message to include in email body
        attach_excel: Whether to attach Excel file (default: True)
    """
    
    doc = frappe.get_doc("Offering Sheet", offering_sheet_name)
    
    # Validate document
    if doc.docstatus != 1:
        frappe.throw(_("Only submitted Offering Sheets can be emailed"))
    
    # Parse recipients
    if not recipients:
        # Get default recipients from Branch or Church Settings
        recipients = get_default_email_recipients(doc.branch)
    
    if isinstance(recipients, str):
        recipients = [email.strip() for email in recipients.split(",")]
    
    if not recipients:
        frappe.throw(_("Please provide at least one recipient email address"))
    
    # Parse CC
    if cc and isinstance(cc, str):
        cc = [email.strip() for email in cc.split(",")]
    
    # Generate email subject
    if not subject:
        subject = f"Offering Sheet - {doc.branch} ({doc.date_from} to {doc.date_to})"
    
    # Generate HTML email body
    email_body = generate_email_template(doc, message)
    
    # Prepare attachments
    attachments = []
    if attach_excel:
        excel_file = generate_excel_attachment(doc)
        attachments.append(excel_file)
    
    # Send email
    try:
        frappe.sendmail(
            recipients=recipients,
            cc=cc,
            subject=subject,
            message=email_body,
            attachments=attachments,
            reference_doctype="Offering Sheet",
            reference_name=doc.name,
            delayed=False
        )
        
        # Log email sent
        log_email_activity(doc, recipients, cc)
        
        frappe.msgprint(
            _("Email sent successfully to {0}").format(", ".join(recipients)),
            alert=True,
            indicator="green"
        )
        
        return {"status": "success", "recipients": recipients}
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), f"Offering Sheet Email Error - {doc.name}")
        frappe.throw(_("Failed to send email: {0}").format(str(e)))


def generate_email_template(doc, additional_message=None):
    """Generate beautiful HTML email template"""
    
    # Calculate statistics
    total_entries = len(doc.offering)
    unique_programmes = len(set([row.programme for row in doc.offering]))
    unique_offering_types = len(set([row.offering_type for row in doc.offering]))
    
    # Group offerings by type for summary
    offering_summary = {}
    for row in doc.offering:
        if row.offering_type not in offering_summary:
            offering_summary[row.offering_type] = 0
        offering_summary[row.offering_type] += flt(row.amount_in_lc)
    
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                margin: 0;
                padding: 0;
                background-color: #f4f7f9;
            }}
            .email-container {{
                max-width: 700px;
                margin: 20px auto;
                background: #ffffff;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 40px 30px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0 0 10px 0;
                font-size: 28px;
                font-weight: 700;
            }}
            .header p {{
                margin: 0;
                font-size: 16px;
                opacity: 0.95;
            }}
            .content {{
                padding: 30px;
            }}
            .info-box {{
                background: linear-gradient(135deg, #f6f8fa 0%, #e9ecef 100%);
                border-left: 4px solid #667eea;
                padding: 20px;
                margin-bottom: 25px;
                border-radius: 6px;
            }}
            .info-row {{
                display: flex;
                justify-content: space-between;
                padding: 8px 0;
                border-bottom: 1px solid #dee2e6;
            }}
            .info-row:last-child {{
                border-bottom: none;
            }}
            .info-label {{
                font-weight: 600;
                color: #495057;
            }}
            .info-value {{
                color: #212529;
            }}
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                gap: 15px;
                margin: 25px 0;
            }}
            .stat-card {{
                background: #f8f9fa;
                padding: 20px;
                border-radius: 8px;
                text-align: center;
                border: 2px solid #e9ecef;
            }}
            .stat-number {{
                font-size: 32px;
                font-weight: 700;
                color: #667eea;
                margin-bottom: 5px;
            }}
            .stat-label {{
                font-size: 12px;
                color: #6c757d;
                text-transform: uppercase;
                letter-spacing: 1px;
            }}
            .summary-table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
                background: white;
                border-radius: 8px;
                overflow: hidden;
            }}
            .summary-table thead {{
                background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                color: white;
            }}
            .summary-table th {{
                padding: 15px;
                text-align: left;
                font-weight: 600;
                text-transform: uppercase;
                font-size: 12px;
                letter-spacing: 1px;
            }}
            .summary-table td {{
                padding: 12px 15px;
                border-bottom: 1px solid #e9ecef;
            }}
            .summary-table tr:last-child td {{
                border-bottom: none;
            }}
            .summary-table tbody tr:nth-child(even) {{
                background: #f8f9fa;
            }}
            .amount {{
                font-weight: 600;
                color: #2c5282;
                text-align: right;
            }}
            .total-box {{
                background: linear-gradient(135deg, #48bb78 0%, #38a169 100%);
                color: white;
                padding: 25px;
                border-radius: 8px;
                text-align: center;
                margin: 25px 0;
            }}
            .total-label {{
                font-size: 14px;
                text-transform: uppercase;
                letter-spacing: 2px;
                margin-bottom: 10px;
                opacity: 0.95;
            }}
            .total-amount {{
                font-size: 42px;
                font-weight: 700;
            }}
            .message-box {{
                background: #fff3cd;
                border-left: 4px solid #ffc107;
                padding: 15px;
                margin: 20px 0;
                border-radius: 6px;
                color: #856404;
            }}
            .cta-button {{
                display: inline-block;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 15px 35px;
                text-decoration: none;
                border-radius: 8px;
                font-weight: 600;
                margin: 20px 0;
                text-align: center;
            }}
            .footer {{
                background: #2d3748;
                color: #cbd5e0;
                padding: 25px 30px;
                text-align: center;
                font-size: 13px;
            }}
            .footer p {{
                margin: 5px 0;
            }}
            .footer a {{
                color: #90cdf4;
                text-decoration: none;
            }}
            @media only screen and (max-width: 600px) {{
                .stats-grid {{
                    grid-template-columns: 1fr;
                }}
                .header h1 {{
                    font-size: 22px;
                }}
                .total-amount {{
                    font-size: 32px;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="email-container">
            <!-- Header -->
            <div class="header">
                <h1>⛪ Offering Sheet Report</h1>
                <p>{doc.branch} - {doc.month}</p>
            </div>
            
            <!-- Content -->
            <div class="content">
                <!-- Document Info -->
                <div class="info-box">
                    <div class="info-row">
                        <span class="info-label">Document No:</span>
                        <span class="info-value">{doc.name}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Reporting Date:</span>
                        <span class="info-value">{frappe.format(doc.reporting_date, {'fieldtype': 'Date'})}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Period:</span>
                        <span class="info-value">{frappe.format(doc.date_from, {'fieldtype': 'Date'})} - {frappe.format(doc.date_to, {'fieldtype': 'Date'})}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Prepared By:</span>
                        <span class="info-value">{doc.prepared_by or 'N/A'}</span>
                    </div>
                </div>
                
                <!-- Statistics -->
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-number">{total_entries}</div>
                        <div class="stat-label">Total Entries</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{unique_programmes}</div>
                        <div class="stat-label">Programmes</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{unique_offering_types}</div>
                        <div class="stat-label">Offering Types</div>
                    </div>
                </div>
                
                <!-- Summary by Offering Type -->
                <h3 style="color: #2d3748; margin-top: 30px;">Summary by Offering Type</h3>
                <table class="summary-table">
                    <thead>
                        <tr>
                            <th>Offering Type</th>
                            <th style="text-align: right;">Amount ({doc.currency})</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # Add offering type summaries
    for offering_type, amount in sorted(offering_summary.items(), key=lambda x: x[1], reverse=True):
        html_template += f"""
                        <tr>
                            <td>{offering_type}</td>
                            <td class="amount">{frappe.format(amount, {'fieldtype': 'Currency', 'options': doc.currency})}</td>
                        </tr>
        """
    
    html_template += """
                    </tbody>
                </table>
                
                <!-- Total Amount -->
                <div class="total-box">
                    <div class="total-label">Total Amount Received</div>
                    <div class="total-amount">{}</div>
                </div>
    """.format(frappe.format(doc.total, {'fieldtype': 'Currency', 'options': doc.currency}))
    
    # Add custom message if provided
    if additional_message:
        html_template += f"""
                <div class="message-box">
                    <strong>Note:</strong> {additional_message}
                </div>
        """
    
    # Add view document button
    doc_url = frappe.utils.get_url_to_form("Offering Sheet", doc.name)
    html_template += f"""
                <div style="text-align: center;">
                    <a href="{doc_url}" class="cta-button">View Full Report</a>
                </div>
                
                <p style="color: #6c757d; font-size: 13px; margin-top: 30px;">
                    <strong>Note:</strong> The complete offering sheet with detailed entries is attached as an Excel file for your records.
                </p>
            </div>
            
            <!-- Footer -->
            <div class="footer">
                <p><strong>Church Management System</strong></p>
                <p>Generated automatically on {frappe.utils.now()}</p>
                <p>This is an automated email. Please do not reply.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_template


def generate_excel_attachment(doc):
    """Generate Excel file as attachment"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # Prepare compact data
    data = [
        [f"OFFERING SHEET - {doc.branch}"],
        [f"Period: {doc.date_from} to {doc.date_to} | Month: {doc.month}"],
        [""],
        ["Date", "Day", "Programme", "Offering Type", "Currency", "Amount (LC)", "Note"]
    ]
    
    for row in doc.offering:
        data.append([
            str(row.date),
            row.day,
            row.programme,
            row.offering_type,
            row.currency or "",
            flt(row.amount_in_lc, 2),
            row.note or ""
        ])
        
    data.append(["", "", "", "", "TOTAL:", flt(doc.total, 2), ""])
    
    # Create Excel with styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offering Sheet"
    
    # Add data with styling
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx == 2:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx > 4 and row_idx < len(data):
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                if col_idx == 6:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
            elif row_idx == len(data):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                if col_idx == 6:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Merge header cells
    ws.merge_cells('A1:G1')
    ws.merge_cells('A2:G2')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return {
        'fname': f'offering_sheet_{doc.name}.xlsx',
        'fcontent': excel_file.getvalue()
    }


def get_default_email_recipients(branch):
    """Get default email recipients for a branch"""
    recipients = []
    
    # Get branch email
    branch_email = frappe.db.get_value("Branch", branch, "email")
    if branch_email:
        recipients.append(branch_email)
    
    # Get Church Settings default recipients
    church_settings = frappe.get_single("Church Settings")
    if hasattr(church_settings, "offering_report_recipients"):
        recipients.extend([r.email for r in church_settings.offering_report_recipients if r.email])
    
    return recipients


def log_email_activity(doc, recipients, cc=None):
    """Log email activity in document timeline"""
    comment = f"Offering Sheet emailed to: {', '.join(recipients)}"
    if cc:
        comment += f"\nCC: {', '.join(cc)}"
    
    doc.add_comment("Info", comment)