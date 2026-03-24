# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt



from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import flt, getdate, nowdate
from frappe.model.document import Document

class AnonymousTithesBatchUpdate(Document):
    def validate(self):
        """Validate anonymous tithes before saving"""
        self.validate_dates()
        self.set_month_from_date()
        self.calculate_totals()
        self.validate_branch_consistency()
        self.propagate_parent_fields()
        
    def validate_dates(self):
        """Ensure date range is valid"""
        if self.start_date and self.end_date:
            if getdate(self.start_date) > getdate(self.end_date):
                frappe.throw(_("Start Date cannot be after End Date"))
                
        if self.reporting_date:
            if getdate(self.reporting_date) < getdate(self.start_date):
                frappe.throw(_("Reporting Date cannot be earlier than Start Date"))
            if getdate(self.reporting_date) < getdate(self.end_date):
                frappe.throw(_("Reporting Date cannot be earlier than End Date"))
                
    def set_month_from_date(self):
        """Auto-set month from reporting date"""
        if self.reporting_date:
            reporting_date = getdate(self.reporting_date)
            self.month = reporting_date.strftime('%B %Y')
            
        if not self.description:
            self.description = 'Anonymous tithes received from various members'
            
    def validate_branch_consistency(self):
        """Ensure all tithe entries have the same branch"""
        if self.branch and self.tithe_transaction_entry_for_anonymous:
            for row in self.tithe_transaction_entry_for_anonymous:
                row.branch = self.branch
                
    def propagate_parent_fields(self):
        """Propagate parent fields to child entries"""
        if self.tithe_transaction_entry_for_anonymous:
            for row in self.tithe_transaction_entry_for_anonymous:
                row.branch = self.branch
                row.designated_bank_acct = self.designated_bank_acct
                row.currency = row.currency or self.currency
                row.other_details = row.other_details or self.description
                
    def calculate_totals(self):
        """Calculate total amounts from tithe entries"""
        total = 0
        for row in self.tithe_transaction_entry_for_anonymous:
            if row.amount_in_lc:
                total += flt(row.amount_in_lc)
        self.amount_paid = total
        
    def on_submit(self):
        """Create receipt vouchers on submit"""
        self.create_receipt_vouchers()
        
    def create_receipt_vouchers(self):
        """Create individual receipt vouchers for each tithe entry"""
        if not self.tithe_transaction_entry_for_anonymous:
            frappe.throw(_("No tithe entries to process"))
            
        created_receipts = []
        
        for row in self.tithe_transaction_entry_for_anonymous:
            # Skip if receipt already exists
            if frappe.db.exists("Receipts", {"referenced_document_no": row.name}):
                frappe.msgprint(_("Receipt for {0} already exists").format(row.name))
                continue
                
            try:
                receipt = self._create_receipt(row)
                created_receipts.append(receipt.name)
            except Exception as e:
                frappe.log_error(frappe.get_traceback(), f"Anonymous Tithes Receipt Creation Error - {row.name}")
                frappe.throw(_("Error creating receipt for {0}: {1}").format(row.name, str(e)))
                
        if created_receipts:
            frappe.msgprint(_("Successfully created {0} receipt(s): {1}").format(
                len(created_receipts), 
                ", ".join(created_receipts)
            ), alert=True)
            
    def _create_receipt(self, row):
        """Create a single receipt voucher for anonymous tithe"""
        
        # Get tithe income account from Church Settings or default
        tithe_account = self.get_tithe_income_account()
        
        purpose = f"Anonymous Tithe - {row.full_name} - {self.month}"
        
        receipt = frappe.get_doc({
            "doctype": "Receipts",
            "naming_series": "REC-",
            "transaction_date": row.date,
            "transaction_type": "Revenue",
            "received_from": f"Anonymous Tithe - {row.full_name}",
            "branch": row.branch or self.branch,
            "create_accounting_entries": 1,
            "source": "Tithe offering",
            "transaction_purposes": purpose,
            "receipt_currency": row.currency or self.currency,
            "exchange_rate": 1 if row.currency == "NGN" else flt(row.exchange_rate, 2),
            "amount_paid": row.amount_paid if row.currency == "NGN" else 0,
            "amount_paid_in_fc": row.amount_paid if row.currency != "NGN" else 0,
            "account_to_credit": tithe_account,
            "mode_of_payment": "Cash",
            "reference_no": row.name.upper(),
            "referenced_document_no": row.name.upper(),
            "remittance_bank": self.designated_bank_acct
        })
        
        receipt.flags.ignore_permissions = True
        receipt.insert()
        
        return receipt
        
    def get_tithe_income_account(self):
        """Get tithe income account from settings or offering type"""
        
        # Try to get from Offering Type
        tithe_offering = frappe.db.get_value("Offering Type", "Tithe offering", "income_account")
        
        if tithe_offering:
            return tithe_offering
            
        # Fallback to Church Settings
        church_settings = frappe.get_single("Church Settings")
        if hasattr(church_settings, "default_tithe_income_account"):
            return church_settings.default_tithe_income_account
            
        # Final fallback - search for any tithe-related account
        account = frappe.db.get_value(
            "Account",
            {"account_name": ["like", "%Tithe%"], "is_group": 0},
            "name"
        )
        
        if not account:
            frappe.throw(_("Please configure Tithe Income Account in Church Settings or Offering Type"))
            
        return account


@frappe.whitelist()
def export_anonymous_tithes(anonymous_tithes_name):
    """Export anonymous tithes to colorful Excel format"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    doc = frappe.get_doc("Anonymous Tithes Batch Update", anonymous_tithes_name)
    
    # Prepare compact data
    data = [
        [f"ANONYMOUS TITHES BATCH - {doc.branch}"],
        [f"Period: {doc.start_date} to {doc.end_date} | Month: {doc.month}"],
        [""],
        ["Date", "Name/Description", "Currency", "Amount (LC)", "Note"]
    ]
    
    for row in doc.tithe_transaction_entry_for_anonymous:
        data.append([
            str(row.date),
            row.full_name,
            row.currency or "",
            flt(row.amount_in_lc, 2),
            row.other_details or ""
        ])
        
    data.append(["", "", "TOTAL:", flt(doc.amount_paid, 2), ""])
    
    # Create Excel with styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anonymous Tithes"
    
    # Add data
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Header styling - Row 1
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Subheader - Row 2
            elif row_idx == 2:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Column headers - Row 4
            elif row_idx == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Data rows - alternating colors
            elif row_idx > 4 and row_idx < len(data):
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                
                # Right-align amount column
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                    
            # Total row
            elif row_idx == len(data):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                
            # Borders for all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    # Merge header cells
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12   # Date
    ws.column_dimensions['B'].width = 30   # Name
    ws.column_dimensions['C'].width = 10   # Currency
    ws.column_dimensions['D'].width = 18   # Amount LC
    ws.column_dimensions['E'].width = 40   # Note
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    frappe.response['filename'] = f'anonymous_tithes_{doc.name}.xlsx'
    frappe.response['filecontent'] = excel_file.getvalue()
    frappe.response['type'] = 'binary'
    
    return True


@frappe.whitelist()
def email_anonymous_tithes(anonymous_tithes_name, recipients=None, cc=None, subject=None, message=None, attach_excel=True):
    """Email anonymous tithes with beautiful HTML template"""
    
    doc = frappe.get_doc("Anonymous Tithes Batch Update", anonymous_tithes_name)
    
    if doc.docstatus != 1:
        frappe.throw(_("Only submitted Anonymous Tithes can be emailed"))
    
    if not recipients:
        recipients = get_default_email_recipients(doc.branch)
    
    if isinstance(recipients, str):
        recipients = [email.strip() for email in recipients.split(",")]
    
    if not recipients:
        frappe.throw(_("Please provide at least one recipient email address"))
    
    if cc and isinstance(cc, str):
        cc = [email.strip() for email in cc.split(",")]
    
    if not subject:
        subject = f"Anonymous Tithes Batch - {doc.branch} ({doc.start_date} to {doc.end_date})"
    
    email_body = generate_anonymous_tithes_email(doc, message)
    
    attachments = []
    if attach_excel:
        excel_file = generate_anonymous_tithes_excel(doc)
        attachments.append(excel_file)
    
    try:
        frappe.sendmail(
            recipients=recipients,
            cc=cc,
            subject=subject,
            message=email_body,
            attachments=attachments,
            reference_doctype="Anonymous Tithes Batch Update",
            reference_name=doc.name,
            delayed=False
        )
        
        log_email_activity(doc, recipients, cc)
        
        frappe.msgprint(
            _("Email sent successfully to {0}").format(", ".join(recipients)),
            alert=True,
            indicator="green"
        )
        
        return {"status": "success", "recipients": recipients}
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), f"Anonymous Tithes Email Error - {doc.name}")
        frappe.throw(_("Failed to send email: {0}").format(str(e)))


def generate_anonymous_tithes_email(doc, additional_message=None):
    """Generate beautiful HTML email template for anonymous tithes"""
    
    total_entries = len(doc.tithe_transaction_entry_for_anonymous)
    
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 0; background-color: #f4f7f9; }}
            .container {{ max-width: 700px; margin: 20px auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #5B9BD5 0%, #70AD47 100%); color: white; padding: 40px 30px; text-align: center; }}
            .header h1 {{ margin: 0 0 10px 0; font-size: 28px; font-weight: 700; }}
            .content {{ padding: 30px; }}
            .info-box {{ background: linear-gradient(135deg, #f6f8fa 0%, #e9ecef 100%); border-left: 4px solid #5B9BD5; padding: 20px; margin-bottom: 25px; border-radius: 6px; }}
            .info-row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #dee2e6; }}
            .info-row:last-child {{ border-bottom: none; }}
            .info-label {{ font-weight: 600; color: #495057; }}
            .info-value {{ color: #212529; }}
            .stats-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; margin: 25px 0; }}
            .stat-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; border: 2px solid #e9ecef; }}
            .stat-number {{ font-size: 32px; font-weight: 700; color: #5B9BD5; margin-bottom: 5px; }}
            .stat-label {{ font-size: 12px; color: #6c757d; text-transform: uppercase; letter-spacing: 1px; }}
            .total-box {{ background: linear-gradient(135deg, #48bb78 0%, #38a169 100%); color: white; padding: 25px; border-radius: 8px; text-align: center; margin: 25px 0; }}
            .total-label {{ font-size: 14px; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 10px; opacity: 0.95; }}
            .total-amount {{ font-size: 42px; font-weight: 700; }}
            .footer {{ background: #2d3748; color: #cbd5e0; padding: 25px 30px; text-align: center; font-size: 13px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>💰 Anonymous Tithes Report</h1>
                <p>{doc.branch} - {doc.month}</p>
            </div>
            
            <div class="content">
                <div class="info-box">
                    <div class="info-row">
                        <span class="info-label">Document No:</span>
                        <span class="info-value">{doc.name}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Reporting Date:</span>
                        <span class="info-value">{frappe.format(doc.reporting_date, {'fieldtype': 'Date'})}</span>
                    </div>
                    <div class="info-row">
                        <span class="info-label">Period:</span>
                        <span class="info-value">{frappe.format(doc.start_date, {'fieldtype': 'Date'})} - {frappe.format(doc.end_date, {'fieldtype': 'Date'})}</span>
                    </div>
                </div>
                
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-number">{total_entries}</div>
                        <div class="stat-label">Total Entries</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">{doc.currency}</div>
                        <div class="stat-label">Currency</div>
                    </div>
                </div>
                
                <div class="total-box">
                    <div class="total-label">Total Anonymous Tithes</div>
                    <div class="total-amount">{frappe.format(doc.amount_paid, {'fieldtype': 'Currency', 'options': doc.currency})}</div>
                </div>
    """
    
    if additional_message:
        html_template += f"""
                <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 6px; color: #856404;">
                    <strong>Note:</strong> {additional_message}
                </div>
        """
    
    doc_url = frappe.utils.get_url_to_form("Anonymous Tithes Batch Update", doc.name)
    html_template += f"""
                <div style="text-align: center; margin-top: 30px;">
                    <a href="{doc_url}" style="display: inline-block; background: linear-gradient(135deg, #5B9BD5 0%, #70AD47 100%); color: white; padding: 15px 35px; text-decoration: none; border-radius: 8px; font-weight: 600;">View Full Report</a>
                </div>
            </div>
            
            <div class="footer">
                <p><strong>Church Management System</strong></p>
                <p>Generated on {frappe.utils.now()}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html_template


def generate_anonymous_tithes_excel(doc):
    """Generate Excel file as attachment"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    data = [
        [f"ANONYMOUS TITHES BATCH - {doc.branch}"],
        [f"Period: {doc.start_date} to {doc.end_date} | Month: {doc.month}"],
        [""],
        ["Date", "Name/Description", "Currency", "Amount (LC)", "Note"]
    ]
    
    for row in doc.tithe_transaction_entry_for_anonymous:
        data.append([
            str(row.date),
            row.full_name,
            row.currency or "",
            flt(row.amount_in_lc, 2),
            row.other_details or ""
        ])
        
    data.append(["", "", "TOTAL:", flt(doc.amount_paid, 2), ""])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anonymous Tithes"
    
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx == 2:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif row_idx > 4 and row_idx < len(data):
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
            elif row_idx == len(data):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
    
    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return {
        'fname': f'anonymous_tithes_{doc.name}.xlsx',
        'fcontent': excel_file.getvalue()
    }


def get_default_email_recipients(branch):
    """Get default email recipients"""
    recipients = []
    
    branch_email = frappe.db.get_value("Branch", branch, "email")
    if branch_email:
        recipients.append(branch_email)
    
    church_settings = frappe.get_single("Church Settings")
    if hasattr(church_settings, "offering_report_recipients"):
        recipients.extend([r.email for r in church_settings.offering_report_recipients if r.email])
    
    return recipients


def log_email_activity(doc, recipients, cc=None):
    """Log email activity"""
    comment = f"Anonymous Tithes emailed to: {', '.join(recipients)}"
    if cc:
        comment += f"\nCC: {', '.join(cc)}"
    
    doc.add_comment("Info", comment)