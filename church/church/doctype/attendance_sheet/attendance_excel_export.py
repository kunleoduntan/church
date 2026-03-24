# -*- coding: utf-8 -*-
"""
Beautiful Excel Export for Attendance Sheet
Professional formatting with colors, charts, and styling
"""

from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import getdate, formatdate, now_datetime, nowdate, format_date, format_datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import io
import base64


@frappe.whitelist()
def export_attendance_to_excel(attendance_sheet_name):
    """
    Export Attendance Sheet to beautiful Excel file
    
    Args:
        attendance_sheet_name: Name of the Attendance Sheet
    
    Returns:
        Dict with file content and filename
    """
    
    # Check if openpyxl is installed
    if not HAS_OPENPYXL:
        frappe.throw(_("""
            Excel export requires 'openpyxl' library. 
            Please install it using: bench pip install openpyxl
        """))
    
    try:
        # Get attendance sheet document
        doc = frappe.get_doc("Attendance Sheet", attendance_sheet_name)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        create_summary_sheet(wb, doc)
        create_detailed_sheet(wb, doc)
        create_charts_sheet(wb, doc)
        create_programme_breakdown_sheet(wb, doc)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Encode to base64
        file_content = base64.b64encode(excel_file.read()).decode('utf-8')
        
        # Generate filename
        filename = f"Attendance_Report_{doc.branch}_{formatdate(doc.reporting_date, 'yyyy-MM-dd')}.xlsx"
        
        return {
            'success': True,
            'file_content': file_content,
            'filename': filename,
            'message': _('Excel file generated successfully')
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), 'Excel Export Error')
        return {
            'success': False,
            'error': str(e)
        }


def create_summary_sheet(wb, doc):
    """Create beautifully formatted summary overview sheet"""
    from frappe.utils import format_datetime, format_date
    
    ws = wb.create_sheet("Summary", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    
    # Define beautiful color palette
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    subheader_fill = PatternFill(start_color="4FACFE", end_color="00F2FE", fill_type="solid")
    accent_fill = PatternFill(start_color="FA709A", end_color="FEE140", fill_type="solid")
    total_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    highlight_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=24, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    subheader_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    label_font = Font(name='Calibri', size=11, bold=True, color="2C3E50")
    value_font = Font(name='Calibri', size=11, color="34495E")
    big_number_font = Font(name='Calibri', size=16, bold=True, color="667EEA")
    percentage_font = Font(name='Calibri', size=10, italic=True, color="7F8C8D")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # TITLE SECTION (Rows 1-3)
    ws.merge_cells('A1:D1')
    ws['A1'] = "📊 ATTENDANCE REPORT"
    ws['A1'].font = title_font
    ws['A1'].fill = header_gradient
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 45
    
    # BRANCH & DATE (Row 2)
    ws.merge_cells('A2:D2')
    formatted_date = format_date(doc.reporting_date, "dd MMMM yyyy") if doc.reporting_date else ""
    ws['A2'] = f"{doc.branch} • {formatted_date}"
    ws['A2'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    ws['A2'].fill = subheader_fill
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].border = thick_border
    ws.row_dimensions[2].height = 30
    
    # Empty row for spacing
    ws.row_dimensions[3].height = 10
    
    current_row = 4
    
    # KEY METRICS SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "📈 KEY METRICS"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = total_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 35
    current_row += 1
    
    # Column Headers
    headers = ['Metric', 'Count', 'Percentage', 'Category']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = subheader_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    # TOTAL ATTENDANCE (Big Number)
    ws[f'A{current_row}'] = "Total Attendance"
    ws[f'A{current_row}'].font = label_font
    ws[f'B{current_row}'] = doc.total_first or 0
    ws[f'B{current_row}'].font = big_number_font
    ws[f'C{current_row}'] = "100%"
    ws[f'C{current_row}'].font = percentage_font
    ws[f'D{current_row}'] = "Overall"
    ws[f'D{current_row}'].font = value_font
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{current_row}'].fill = highlight_fill
        ws[f'{col}{current_row}'].border = thin_border
        ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # MEMBERS BREAKDOWN
    member_data = [
        ("Men", doc.total_men or 0, calculate_percentage(doc.total_men, doc.total_first), "Members"),
        ("Women", doc.total_women or 0, calculate_percentage(doc.total_women, doc.total_first), "Members"),
        ("Children", doc.total_children or 0, calculate_percentage(doc.total_children, doc.total_first), "Members"),
    ]
    
    for label, count, pct, category in member_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = label_font
        ws[f'B{current_row}'] = count
        ws[f'B{current_row}'].font = value_font
        ws[f'C{current_row}'] = f"{pct}%"
        ws[f'C{current_row}'].font = percentage_font
        ws[f'D{current_row}'] = category
        ws[f'D{current_row}'].font = value_font
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
        current_row += 1
    
    # Spacing
    current_row += 1
    
    # VISITORS SECTION
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "🌟 VISITORS"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = total_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].border = thick_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Column Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = subheader_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    current_row += 1
    
    # Total Visitors
    ws[f'A{current_row}'] = "Total Visitors"
    ws[f'A{current_row}'].font = label_font
    ws[f'B{current_row}'] = doc.total_second or 0
    ws[f'B{current_row}'].font = big_number_font
    ws[f'C{current_row}'] = f"{calculate_percentage(doc.total_second, doc.total_first)}%"
    ws[f'C{current_row}'].font = percentage_font
    ws[f'D{current_row}'] = "New"
    ws[f'D{current_row}'].font = value_font
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{current_row}'].fill = highlight_fill
        ws[f'{col}{current_row}'].border = thin_border
        ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1
    
    # Visitor Breakdown
    visitor_data = [
        ("New Men", doc.total_new_men or 0, calculate_percentage(doc.total_new_men, doc.total_second), "Visitors"),
        ("New Women", doc.total_new_women or 0, calculate_percentage(doc.total_new_women, doc.total_second), "Visitors"),
        ("New Children", doc.total_new_children or 0, calculate_percentage(doc.total_new_children, doc.total_second), "Visitors"),
    ]
    
    for label, count, pct, category in visitor_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = label_font
        ws[f'B{current_row}'] = count
        ws[f'B{current_row}'].font = value_font
        ws[f'C{current_row}'] = f"{pct}%"
        ws[f'C{current_row}'].font = percentage_font
        ws[f'D{current_row}'] = category
        ws[f'D{current_row}'].font = value_font
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
        current_row += 1
    
    # Spacing
    current_row += 2
    
    # FOOTER - Generation timestamp
    ws.merge_cells(f'A{current_row}:D{current_row}')
    generation_time = format_datetime(now_datetime(), "dd MMM yyyy HH:mm")
    ws[f'A{current_row}'] = f"📅 Generated on: {generation_time}"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=9, italic=True, color="95A5A6")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add church branding footer
    current_row += 1
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "⛪ Ecclesia Church Management System"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True, color="667EEA")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')


def create_detailed_sheet(wb, doc):
    """Create beautifully formatted detailed breakdown sheet"""
    from frappe.utils import format_date
    
    ws = wb.create_sheet("Detailed Breakdown")
    
    # Define beautiful color palette
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    subheader_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    total_fill = PatternFill(start_color="2ECC71", end_color="27AE60", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=10, color="2C3E50")
    bold_data_font = Font(name='Calibri', size=10, bold=True, color="2C3E50")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # TITLE ROW
    ws.merge_cells('A1:K1')
    ws['A1'] = "📋 DETAILED ATTENDANCE BREAKDOWN"
    ws['A1'].font = title_font
    ws['A1'].fill = header_gradient
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 40
    
    # Headers with categories
    headers = [
        ("Date", 12, "General"),
        ("Day", 10, "General"),
        ("Programme", 25, "General"),
        ("Men", 10, "Members"),
        ("Women", 10, "Members"),
        ("Children", 12, "Members"),
        ("Total", 10, "Members"),
        ("New Men", 10, "Visitors"),
        ("New Women", 12, "Visitors"),
        ("New Children", 12, "Visitors"),
        ("Visitors", 10, "Visitors")
    ]
    
    # Set column widths and write headers
    for col_idx, (header, width, category) in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        
        # Color code by category
        if category == "Members":
            cell.fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
        elif category == "Visitors":
            cell.fill = PatternFill(start_color="FA709A", end_color="FA709A", fill_type="solid")
        else:
            cell.fill = subheader_fill
        
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[2].height = 30
    
    # Write data rows with zebra striping
    row_idx = 3
    for row_data in doc.church_attendance_analysis:
        # Alternate row colors for readability
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        
        # Format date properly
        formatted_date = format_date(row_data.date, "dd MMM") if row_data.date else ""
        
        data = [
            formatted_date,
            row_data.day or "",
            row_data.programme,
            row_data.men or 0,
            row_data.women or 0,
            row_data.children or 0,
            row_data.total or 0,
            row_data.new_men or 0,
            row_data.new_women or 0,
            row_data.new_children or 0,
            row_data.new_total or 0
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.fill = fill
            
            # Bold programme name
            if col_idx == 3:
                cell.font = bold_data_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
            # Right align numbers, highlight totals
            elif col_idx > 3:
                cell.font = data_font
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Highlight total columns
                if col_idx in [7, 11]:
                    cell.font = Font(name='Calibri', size=10, bold=True, color="667EEA")
            else:
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
    
    # Add GRAND TOTAL row with beautiful styling
    totals = [
        "", "", "GRAND TOTAL",
        doc.total_men or 0,
        doc.total_women or 0,
        doc.total_children or 0,
        doc.total_first or 0,
        doc.total_new_men or 0,
        doc.total_new_women or 0,
        doc.total_new_children or 0,
        doc.total_second or 0
    ]
    
    for col_idx, value in enumerate(totals, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thick_border
        
        if col_idx <= 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    ws.row_dimensions[row_idx].height = 35
    
    # Freeze panes for scrolling
    ws.freeze_panes = 'A3'
    
    # Add summary footer
    row_idx += 2
    ws.merge_cells(f'A{row_idx}:K{row_idx}')
    ws[f'A{row_idx}'] = f"Total Programmes: {len(doc.church_attendance_analysis)} | Total Attendance: {doc.total_first or 0} | Visitors: {doc.total_second or 0}"
    ws[f'A{row_idx}'].font = Font(name='Calibri', size=10, italic=True, color="7F8C8D")
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')


def create_charts_sheet(wb, doc):
    """Create beautifully formatted charts and visualizations sheet"""
    ws = wb.create_sheet("Charts & Analytics")
    
    # Define beautiful colors
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    chart_header_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    data_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=20, bold=True, color="FFFFFF")
    section_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=10, color="2C3E50")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    # TITLE
    ws.merge_cells('A1:F1')
    ws['A1'] = "📊 VISUAL ANALYTICS & INSIGHTS"
    ws['A1'].font = title_font
    ws['A1'].fill = header_gradient
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 45
    
    # SECTION 1: ATTENDANCE DISTRIBUTION
    ws.merge_cells('A3:D3')
    ws['A3'] = "👥 Attendance Distribution"
    ws['A3'].font = section_font
    ws['A3'].fill = chart_header_fill
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].border = thick_border
    ws.row_dimensions[3].height = 30
    
    # Data headers
    headers = ['Category', 'Count', 'Percentage', 'Visual']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="FA709A", end_color="FA709A", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[4].height = 25
    
    # Prepare data for charts
    categories = ['Men', 'Women', 'Children']
    values = [
        doc.total_men or 0,
        doc.total_women or 0,
        doc.total_children or 0
    ]
    
    total = sum(values)
    
    # Write distribution data with visual bars
    for idx, (cat, val) in enumerate(zip(categories, values), 5):
        pct = calculate_percentage(val, total)
        
        ws[f'A{idx}'] = cat
        ws[f'A{idx}'].font = data_font
        ws[f'A{idx}'].border = thin_border
        ws[f'A{idx}'].fill = data_fill
        
        ws[f'B{idx}'] = val
        ws[f'B{idx}'].font = Font(name='Calibri', size=11, bold=True, color="667EEA")
        ws[f'B{idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'B{idx}'].border = thin_border
        ws[f'B{idx}'].fill = data_fill
        
        ws[f'C{idx}'] = f"{pct}%"
        ws[f'C{idx}'].font = data_font
        ws[f'C{idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{idx}'].border = thin_border
        ws[f'C{idx}'].fill = data_fill
        
        # Visual bar representation
        bar_length = int(pct / 5)  # Scale for display
        ws[f'D{idx}'] = "█" * bar_length
        ws[f'D{idx}'].font = Font(name='Calibri', size=11, color="667EEA")
        ws[f'D{idx}'].border = thin_border
        ws[f'D{idx}'].fill = data_fill
        
        ws.row_dimensions[idx].height = 22
    
    # Create Pie Chart
    try:
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=5, max_row=7)
        data = Reference(ws, min_col=2, min_row=4, max_row=7)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Attendance by Category"
        pie.height = 12
        pie.width = 20
        ws.add_chart(pie, "F4")
    except:
        pass  # Chart creation might fail in some environments
    
    # SECTION 2: VISITOR STATISTICS
    ws.merge_cells('A10:D10')
    ws['A10'] = "🌟 Visitor Statistics"
    ws['A10'].font = section_font
    ws['A10'].fill = chart_header_fill
    ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A10'].border = thick_border
    ws.row_dimensions[10].height = 30
    
    # Visitor headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="F093FB", end_color="F093FB", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[11].height = 25
    
    visitor_data = [
        ("New Men", doc.total_new_men or 0),
        ("New Women", doc.total_new_women or 0),
        ("New Children", doc.total_new_children or 0),
        ("Total Visitors", doc.total_second or 0)
    ]
    
    visitor_total = doc.total_second or 1  # Avoid division by zero
    
    for idx, (cat, val) in enumerate(visitor_data, 12):
        pct = calculate_percentage(val, visitor_total)
        
        ws[f'A{idx}'] = cat
        ws[f'A{idx}'].font = data_font
        ws[f'A{idx}'].border = thin_border
        ws[f'A{idx}'].fill = data_fill
        
        ws[f'B{idx}'] = val
        ws[f'B{idx}'].font = Font(name='Calibri', size=11, bold=True, color="E91E63")
        ws[f'B{idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'B{idx}'].border = thin_border
        ws[f'B{idx}'].fill = data_fill
        
        ws[f'C{idx}'] = f"{pct}%"
        ws[f'C{idx}'].font = data_font
        ws[f'C{idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{idx}'].border = thin_border
        ws[f'C{idx}'].fill = data_fill
        
        # Visual bar
        bar_length = int(pct / 5)
        ws[f'D{idx}'] = "█" * bar_length
        ws[f'D{idx}'].font = Font(name='Calibri', size=11, color="E91E63")
        ws[f'D{idx}'].border = thin_border
        ws[f'D{idx}'].fill = data_fill
        
        ws.row_dimensions[idx].height = 22
    
    # Create Bar Chart for visitors
    try:
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Visitor Breakdown"
        bar.y_axis.title = 'Count'
        bar.x_axis.title = 'Category'
        
        labels = Reference(ws, min_col=1, min_row=12, max_row=15)
        data = Reference(ws, min_col=2, min_row=11, max_row=15)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        bar.height = 12
        bar.width = 20
        ws.add_chart(bar, "F17")
    except:
        pass
    
    # KEY INSIGHTS Section
    ws.merge_cells('A18:D18')
    ws['A18'] = "💡 Key Insights"
    ws['A18'].font = section_font
    ws['A18'].fill = PatternFill(start_color="2ECC71", end_color="27AE60", fill_type="solid")
    ws['A18'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A18'].border = thick_border
    ws.row_dimensions[18].height = 30
    
    # Calculate insights
    visitor_rate = calculate_percentage(doc.total_second, doc.total_first)
    member_rate = 100 - visitor_rate
    
    insights = [
        f"• {visitor_rate}% of total attendance were visitors",
        f"• {member_rate}% were existing members",
        f"• Largest demographic: {'Men' if doc.total_men > doc.total_women and doc.total_men > doc.total_children else 'Women' if doc.total_women > doc.total_children else 'Children'}",
        f"• Total programmes covered: {len(doc.church_attendance_analysis)}"
    ]
    
    for idx, insight in enumerate(insights, 19):
        ws.merge_cells(f'A{idx}:D{idx}')
        ws[f'A{idx}'] = insight
        ws[f'A{idx}'].font = Font(name='Calibri', size=10, color="2C3E50")
        ws[f'A{idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[idx].height = 20


def create_programme_breakdown_sheet(wb, doc):
    """Create beautifully formatted programme-wise breakdown sheet"""
    ws = wb.create_sheet("Programme Analysis")
    
    # Define beautiful colors
    header_gradient = PatternFill(start_color="667EEA", end_color="764BA2", fill_type="solid")
    column_header_fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="2ECC71", end_color="27AE60", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    
    # Define fonts
    title_font = Font(name='Calibri', size=20, bold=True, color="FFFFFF")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=10, color="2C3E50")
    bold_data_font = Font(name='Calibri', size=10, bold=True, color="2C3E50")
    percentage_font = Font(name='Calibri', size=10, italic=True, color="7F8C8D")
    
    # Define borders
    thick_border = Border(
        left=Side(style='medium', color="667EEA"),
        right=Side(style='medium', color="667EEA"),
        top=Side(style='medium', color="667EEA"),
        bottom=Side(style='medium', color="667EEA")
    )
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )
    
    # TITLE
    ws.merge_cells('A1:G1')
    ws['A1'] = "📈 PROGRAMME-WISE ATTENDANCE ANALYSIS"
    ws['A1'].font = title_font
    ws['A1'].fill = header_gradient
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thick_border
    ws.row_dimensions[1].height = 45
    
    # Set column widths
    column_widths = [30, 18, 15, 15, 12, 12, 15]
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col, width in zip(columns, column_widths):
        ws.column_dimensions[col].width = width
    
    # Headers with beautiful styling
    headers = [
        "Programme", 
        "Total Attendance", 
        "Members", 
        "Visitors", 
        "Visitor %",
        "Rank",
        "Performance"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 35
    
    # Group data by programme
    programme_data = {}
    for row in doc.church_attendance_analysis:
        prog = row.programme
        if prog not in programme_data:
            programme_data[prog] = {
                'total': 0,
                'members': 0,
                'visitors': 0
            }
        
        programme_data[prog]['total'] += (row.total or 0)
        programme_data[prog]['members'] += (row.existing_total or 0)
        programme_data[prog]['visitors'] += (row.new_total or 0)
    
    # Sort programmes by total attendance (descending)
    sorted_programmes = sorted(programme_data.items(), key=lambda x: x[1]['total'], reverse=True)
    
    # Calculate overall max for performance indicator
    max_attendance = max([data['total'] for _, data in sorted_programmes]) if sorted_programmes else 1
    
    # Write data with ranking and performance indicators
    row_idx = 3
    for rank, (programme, data) in enumerate(sorted_programmes, 1):
        total = data['total']
        members = data['members']
        visitors = data['visitors']
        visitor_pct = calculate_percentage(visitors, total)
        performance_pct = (total / max_attendance) * 100 if max_attendance > 0 else 0
        
        # Alternate row colors
        fill = alt_row_fill if row_idx % 2 == 0 else white_fill
        
        # Highlight top performer
        if rank == 1:
            fill = highlight_fill
        
        # Programme name
        ws[f'A{row_idx}'] = programme
        ws[f'A{row_idx}'].font = bold_data_font if rank <= 3 else data_font
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{row_idx}'].border = thin_border
        ws[f'A{row_idx}'].fill = fill
        
        # Total attendance
        ws[f'B{row_idx}'] = total
        ws[f'B{row_idx}'].font = Font(name='Calibri', size=11, bold=True, color="667EEA") if rank <= 3 else data_font
        ws[f'B{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'B{row_idx}'].border = thin_border
        ws[f'B{row_idx}'].fill = fill
        
        # Members
        ws[f'C{row_idx}'] = members
        ws[f'C{row_idx}'].font = data_font
        ws[f'C{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{row_idx}'].border = thin_border
        ws[f'C{row_idx}'].fill = fill
        
        # Visitors
        ws[f'D{row_idx}'] = visitors
        ws[f'D{row_idx}'].font = data_font
        ws[f'D{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{row_idx}'].border = thin_border
        ws[f'D{row_idx}'].fill = fill
        
        # Visitor percentage
        ws[f'E{row_idx}'] = f"{visitor_pct}%"
        ws[f'E{row_idx}'].font = percentage_font
        ws[f'E{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'E{row_idx}'].border = thin_border
        ws[f'E{row_idx}'].fill = fill
        
        # Rank with medal for top 3
        rank_display = f"🥇 {rank}" if rank == 1 else f"🥈 {rank}" if rank == 2 else f"🥉 {rank}" if rank == 3 else str(rank)
        ws[f'F{row_idx}'] = rank_display
        ws[f'F{row_idx}'].font = bold_data_font if rank <= 3 else data_font
        ws[f'F{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'F{row_idx}'].border = thin_border
        ws[f'F{row_idx}'].fill = fill
        
        # Performance bar
        bar_length = int(performance_pct / 10)
        ws[f'G{row_idx}'] = "█" * bar_length
        ws[f'G{row_idx}'].font = Font(name='Calibri', size=11, color="2ECC71" if rank == 1 else "667EEA")
        ws[f'G{row_idx}'].border = thin_border
        ws[f'G{row_idx}'].fill = fill
        
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1
    
    # Add SUMMARY row
    row_idx += 1
    summary_data = [
        "OVERALL SUMMARY",
        doc.total_first or 0,
        doc.total_third or 0,
        doc.total_second or 0,
        f"{calculate_percentage(doc.total_second, doc.total_first)}%",
        "—",
        "100%"
    ]
    
    for col_idx, value in enumerate(summary_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center' if col_idx in [1, 6, 7] else 'right', vertical='center')
    
    ws.row_dimensions[row_idx].height = 35
    
    # Add insights footer
    row_idx += 2
    ws.merge_cells(f'A{row_idx}:G{row_idx}')
    top_programme = sorted_programmes[0][0] if sorted_programmes else "N/A"
    ws[f'A{row_idx}'] = f"💡 Top Performing Programme: {top_programme} | Total Programmes: {len(programme_data)}"
    ws[f'A{row_idx}'].font = Font(name='Calibri', size=11, italic=True, color="7F8C8D")
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')


def calculate_percentage(value, total):
    """Calculate percentage"""
    if not total or total == 0:
        return 0
    return round((value / total) * 100, 1)


@frappe.whitelist()
def download_excel(attendance_sheet_name):
    """
    Download Excel file directly
    Used by client-side to trigger download
    """
    result = export_attendance_to_excel(attendance_sheet_name)
    
    if result.get('success'):
        return {
            'success': True,
            'file_url': f'/api/method/church.church.doctype.attendance_sheet.attendance_excel_export.get_excel_file?attendance_sheet={attendance_sheet_name}'
        }
    else:
        frappe.throw(result.get('error', 'Failed to generate Excel file'))


@frappe.whitelist()
def get_excel_file(attendance_sheet):
    """
    Return Excel file for download
    """
    result = export_attendance_to_excel(attendance_sheet)
    
    if not result.get('success'):
        frappe.throw(result.get('error', 'Failed to generate Excel file'))
    
    # Decode base64
    file_content = base64.b64decode(result['file_content'])
    
    # Set response headers
    frappe.local.response.filename = result['filename']
    frappe.local.response.filecontent = file_content
    frappe.local.response.type = "download"