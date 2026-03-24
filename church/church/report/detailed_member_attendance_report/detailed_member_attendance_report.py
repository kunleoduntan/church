# Copyright (c) 2026, Church Management and contributors
# For license information, please see license.txt

"""
Detailed Member Attendance Report - Individual Level

Features:
- Shows individual member names
- Present vs Absent status
- Service-by-service breakdown
- Color-coded Excel (Green=Present, Red=Absent)
- Beautiful HTML with attendance badges
- Filter by demographic, branch, date range
- Attendance percentage calculation
"""

import frappe
from frappe import _
from frappe.utils import getdate, format_date, flt, cint, nowdate
from collections import defaultdict
import json


def execute(filters=None):
    """
    Main report execution function
    
    Returns:
        columns: List of column definitions
        data: List of data rows with individual members
        message: Optional message
        chart: Chart configuration
        report_summary: Summary statistics
    """
    
    if not filters:
        filters = {}
    
    # Validate filters
    validate_filters(filters)
    
    # Get columns
    columns = get_columns(filters)
    
    # Get data
    data = get_data(filters)
    
    # Get chart
    chart = get_chart_data(data, filters)
    
    # Get report summary
    report_summary = get_report_summary(data, filters)
    
    return columns, data, None, chart, report_summary


def validate_filters(filters):
    """Validate and process report filters with flexible date options"""
    
    # Process date filters based on what user selected
    from_date = None
    to_date = None
    
    # Priority 1: Custom Date Range
    if filters.get('from_date') and filters.get('to_date'):
        from_date = getdate(filters['from_date'])
        to_date = getdate(filters['to_date'])
    
    # Priority 2: Period (This Month, Last Month, This Quarter, etc.)
    elif filters.get('period'):
        from_date, to_date = get_period_dates(filters['period'])
    
    # Priority 3: Service Instance
    elif filters.get('service_instance'):
        from_date, to_date = get_service_instance_dates(filters['service_instance'])
    
    # Priority 4: Specific Month
    elif filters.get('month') and filters.get('year'):
        from_date, to_date = get_month_dates(filters['month'], filters['year'])
    
    # Priority 5: Specific Week
    elif filters.get('week') and filters.get('year'):
        from_date, to_date = get_week_dates(filters['week'], filters['year'])
    
    # Priority 6: Year only
    elif filters.get('year'):
        from_date, to_date = get_year_dates(filters['year'])
    
    # Default: Last 30 days
    else:
        to_date = getdate(nowdate())
        from_date = frappe.utils.add_days(to_date, -30)
    
    # Validate date range
    if from_date > to_date:
        frappe.throw(_("From Date cannot be after To Date"))
    
    # Set processed dates back to filters
    filters['from_date'] = from_date
    filters['to_date'] = to_date
    
    return filters


def get_period_dates(period):
    """Get date range for predefined periods"""
    
    today = getdate(nowdate())
    
    if period == 'Today':
        return today, today
    
    elif period == 'Yesterday':
        yesterday = frappe.utils.add_days(today, -1)
        return yesterday, yesterday
    
    elif period == 'This Week':
        # Week starts on Sunday
        week_start = frappe.utils.get_first_day_of_week(today)
        return week_start, today
    
    elif period == 'Last Week':
        last_week_end = frappe.utils.add_days(frappe.utils.get_first_day_of_week(today), -1)
        last_week_start = frappe.utils.get_first_day_of_week(last_week_end)
        return last_week_start, last_week_end
    
    elif period == 'This Month':
        month_start = frappe.utils.get_first_day(today)
        return month_start, today
    
    elif period == 'Last Month':
        last_month_end = frappe.utils.add_days(frappe.utils.get_first_day(today), -1)
        last_month_start = frappe.utils.get_first_day(last_month_end)
        return last_month_start, last_month_end
    
    elif period == 'This Quarter':
        quarter_start = frappe.utils.get_quarter_start(today)
        return quarter_start, today
    
    elif period == 'Last Quarter':
        last_quarter_end = frappe.utils.add_days(frappe.utils.get_quarter_start(today), -1)
        last_quarter_start = frappe.utils.get_quarter_start(last_quarter_end)
        return last_quarter_start, last_quarter_end
    
    elif period == 'This Year':
        year_start = frappe.utils.get_year_start(today)
        return year_start, today
    
    elif period == 'Last Year':
        last_year_end = frappe.utils.add_days(frappe.utils.get_year_start(today), -1)
        last_year_start = frappe.utils.get_year_start(last_year_end)
        return last_year_start, last_year_end
    
    elif period == 'Last 7 Days':
        return frappe.utils.add_days(today, -7), today
    
    elif period == 'Last 30 Days':
        return frappe.utils.add_days(today, -30), today
    
    elif period == 'Last 90 Days':
        return frappe.utils.add_days(today, -90), today
    
    elif period == 'Year to Date':
        year_start = frappe.utils.get_year_start(today)
        return year_start, today
    
    else:
        # Default to last 30 days
        return frappe.utils.add_days(today, -30), today


def get_service_instance_dates(service_instance_name):
    """Get date range from Service Instance"""
    
    try:
        service_instance = frappe.get_doc('Service Instance', service_instance_name)
        
        # Service Instance has service_date field
        service_date = getdate(service_instance.service_date)
        
        # Return single day range (service date only)
        return service_date, service_date
        
    except Exception as e:
        frappe.log_error(f"Error getting service instance dates: {str(e)}")
        # Fallback to today
        today = getdate(nowdate())
        return today, today


def get_month_dates(month, year):
    """Get first and last day of a specific month"""
    
    from calendar import monthrange
    
    try:
        # Get number of days in month
        _, last_day = monthrange(int(year), int(month))
        
        # Create date strings
        from_date = getdate(f"{year}-{month:02d}-01")
        to_date = getdate(f"{year}-{month:02d}-{last_day}")
        
        return from_date, to_date
        
    except Exception as e:
        frappe.log_error(f"Error getting month dates: {str(e)}")
        # Fallback to current month
        today = getdate(nowdate())
        return frappe.utils.get_first_day(today), frappe.utils.get_last_day(today)


def get_week_dates(week_number, year):
    """Get first and last day of a specific week"""
    
    import datetime
    
    try:
        # Get the first day of the year
        jan_1 = datetime.date(int(year), 1, 1)
        
        # Calculate the date of Monday of the given week
        # ISO week starts on Monday
        week_start = jan_1 + datetime.timedelta(weeks=int(week_number)-1)
        
        # Adjust to the Monday of that week
        week_start = week_start - datetime.timedelta(days=week_start.weekday())
        
        # Week ends on Sunday
        week_end = week_start + datetime.timedelta(days=6)
        
        return getdate(week_start), getdate(week_end)
        
    except Exception as e:
        frappe.log_error(f"Error getting week dates: {str(e)}")
        # Fallback to current week
        today = getdate(nowdate())
        return frappe.utils.get_first_day_of_week(today), frappe.utils.add_days(today, 6)


def get_year_dates(year):
    """Get first and last day of a specific year"""
    
    try:
        from_date = getdate(f"{year}-01-01")
        to_date = getdate(f"{year}-12-31")
        
        return from_date, to_date
        
    except Exception as e:
        frappe.log_error(f"Error getting year dates: {str(e)}")
        # Fallback to current year
        today = getdate(nowdate())
        return frappe.utils.get_year_start(today), frappe.utils.get_year_ending(today)


def get_columns(filters):
    """
    Define report columns
    """
    
    columns = [
        {
            'fieldname': 'member_id',
            'label': _('Member ID'),
            'fieldtype': 'Link',
            'options': 'Member',
            'width': 120
        },
        {
            'fieldname': 'full_name',
            'label': _('Full Name'),
            'fieldtype': 'Data',
            'width': 200
        },
        {
            'fieldname': 'demographic_group',
            'label': _('Demographic'),
            'fieldtype': 'Data',
            'width': 120
        },
        {
            'fieldname': 'branch',
            'label': _('Branch'),
            'fieldtype': 'Link',
            'options': 'Branch',
            'width': 150
        },
        {
            'fieldname': 'total_services',
            'label': _('Total Services'),
            'fieldtype': 'Int',
            'width': 120
        },
        {
            'fieldname': 'services_attended',
            'label': _('Present'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'services_absent',
            'label': _('Absent'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'attendance_percentage',
            'label': _('Attendance %'),
            'fieldtype': 'Percent',
            'width': 120
        },
        {
            'fieldname': 'status',
            'label': _('Status'),
            'fieldtype': 'Data',
            'width': 120
        },
        {
            'fieldname': 'last_attended',
            'label': _('Last Attended'),
            'fieldtype': 'Date',
            'width': 120
        }
    ]
    
    # Add service columns if show_services is enabled
    if filters.get('show_services'):
        service_dates = get_service_dates(filters)
        for service_date in service_dates[:10]:  # Limit to 10 services for display
            columns.append({
                'fieldname': f'service_{service_date}',
                'label': format_date(service_date),
                'fieldtype': 'Data',
                'width': 100
            })
    
    return columns


def get_data(filters):
    """
    Fetch detailed attendance data for each member
    """
    
    from_date = getdate(filters.get('from_date'))
    to_date = getdate(filters.get('to_date'))
    branch = filters.get('branch')
    demographic = filters.get('demographic_group')
    status_filter = filters.get('status')
    
    # Get all active members
    member_filters = {
        'member_status': 'Active'
    }
    
    if branch:
        member_filters['branch'] = branch
    
    if demographic:
        member_filters['demographic_group'] = demographic
    
    members = frappe.get_all('Member',
        filters=member_filters,
        fields=['name', 'full_name', 'demographic_group', 'branch'],
        order_by='full_name'
    )
    
    # Get all service dates in the period
    service_dates = get_service_dates(filters)
    total_services = len(service_dates)
    
    if total_services == 0:
        frappe.msgprint(_("No services found in the selected period"))
        return []
    
    data = []
    
    for member in members:
        # Get attendance records for this member
        attendance_data = get_member_attendance(
            member['name'], 
            from_date, 
            to_date, 
            branch
        )
        
        services_attended = len(attendance_data['present_dates'])
        services_absent = total_services - services_attended
        
        # Calculate attendance percentage
        attendance_pct = (services_attended / total_services * 100) if total_services > 0 else 0
        
        # Determine status
        if attendance_pct >= 80:
            status = '✅ Excellent'
        elif attendance_pct >= 60:
            status = '🟢 Good'
        elif attendance_pct >= 40:
            status = '🟡 Fair'
        elif attendance_pct >= 20:
            status = '🟠 Poor'
        else:
            status = '🔴 Very Poor'
        
        # Get last attended date
        last_attended = max(attendance_data['present_dates']) if attendance_data['present_dates'] else None
        
        row = {
            'member_id': member['name'],
            'full_name': member['full_name'],
            'demographic_group': member['demographic_group'],
            'branch': member['branch'],
            'total_services': total_services,
            'services_attended': services_attended,
            'services_absent': services_absent,
            'attendance_percentage': attendance_pct,
            'status': status,
            'last_attended': last_attended
        }
        
        # Add service-specific attendance if requested
        if filters.get('show_services'):
            for service_date in service_dates[:10]:
                if service_date in attendance_data['present_dates']:
                    row[f'service_{service_date}'] = '✅ Present'
                else:
                    row[f'service_{service_date}'] = '❌ Absent'
        
        # Filter by status if specified
        if status_filter:
            if status_filter == 'Excellent' and attendance_pct < 80:
                continue
            elif status_filter == 'Good' and (attendance_pct < 60 or attendance_pct >= 80):
                continue
            elif status_filter == 'Fair' and (attendance_pct < 40 or attendance_pct >= 60):
                continue
            elif status_filter == 'Poor' and (attendance_pct < 20 or attendance_pct >= 40):
                continue
            elif status_filter == 'Very Poor' and attendance_pct >= 20:
                continue
        
        data.append(row)
    
    return data


def get_service_dates(filters):
    """Get all unique service dates in the period"""
    
    from_date = getdate(filters.get('from_date'))
    to_date = getdate(filters.get('to_date'))
    branch = filters.get('branch')
    
    try:
        filters_dict = {
            'service_date': ['between', [from_date, to_date]]
        }
        
        if branch:
            filters_dict['branch'] = branch
        
        dates = frappe.get_all('Church Attendance',
            filters=filters_dict,
            fields=['service_date'],
            distinct=True,
            order_by='service_date'
        )
        
        return [d['service_date'] for d in dates]
        
    except Exception as e:
        frappe.log_error(f"Error getting service dates: {str(e)}")
        return []


def get_member_attendance(member_id, from_date, to_date, branch=None):
    """
    Get attendance records for a specific member
    
    Returns:
        dict with present_dates and absent_dates
    """
    
    try:
        filters_dict = {
            'member_id': member_id,
            'service_date': ['between', [from_date, to_date]]
        }
        
        if branch:
            filters_dict['branch'] = branch
        
        # Get all attendance records
        attendance = frappe.get_all('Church Attendance',
            filters=filters_dict,
            fields=['service_date', 'present'],
            order_by='service_date'
        )
        
        present_dates = []
        
        for record in attendance:
            if record['present']:
                present_dates.append(record['service_date'])
        
        return {
            'present_dates': present_dates
        }
        
    except Exception as e:
        frappe.log_error(f"Error getting member attendance: {str(e)}")
        return {'present_dates': []}


def get_chart_data(data, filters):
    """Generate attendance distribution chart"""
    
    if not data:
        return None
    
    # Count members by status
    status_counts = defaultdict(int)
    for row in data:
        status = row['status'].split(' ')[1] if ' ' in row['status'] else row['status']
        status_counts[status] += 1
    
    labels = list(status_counts.keys())
    values = list(status_counts.values())
    
    chart = {
        'data': {
            'labels': labels,
            'datasets': [{
                'name': 'Members',
                'values': values
            }]
        },
        'type': 'pie',
        'height': 300,
        'colors': ['#2ecc71', '#4CAF50', '#FFC107', '#FF9800', '#f44336']
    }
    
    return chart


def get_report_summary(data, filters):
    """Generate report summary"""
    
    if not data:
        return []
    
    total_members = len(data)
    
    # Count by status
    excellent = sum(1 for r in data if r['attendance_percentage'] >= 80)
    good = sum(1 for r in data if 60 <= r['attendance_percentage'] < 80)
    fair = sum(1 for r in data if 40 <= r['attendance_percentage'] < 60)
    poor = sum(1 for r in data if 20 <= r['attendance_percentage'] < 40)
    very_poor = sum(1 for r in data if r['attendance_percentage'] < 20)
    
    # Average attendance
    avg_attendance = sum(r['attendance_percentage'] for r in data) / total_members if total_members > 0 else 0
    
    summary = [
        {
            'value': total_members,
            'label': _('Total Members'),
            'indicator': 'Blue',
            'datatype': 'Int'
        },
        {
            'value': avg_attendance,
            'label': _('Average Attendance'),
            'indicator': 'Green',
            'datatype': 'Percent'
        },
        {
            'value': excellent,
            'label': _('Excellent (80%+)'),
            'indicator': 'Green',
            'datatype': 'Int'
        },
        {
            'value': poor + very_poor,
            'label': _('Poor (<40%)'),
            'indicator': 'Red',
            'datatype': 'Int'
        }
    ]
    
    return summary


# ============================================================================
# HTML REPORT GENERATION
# ============================================================================

def get_html(filters):
    """Generate beautiful HTML report with color-coded attendance"""
    
    columns, data, _, chart, summary = execute(filters)
    
    church_settings = frappe.get_single('Church Settings')
    church_name = church_settings.church_name if church_settings else 'Church'
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Detailed Attendance Report</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 20px;
            }}
            .container {{
                max-width: 1600px;
                margin: 0 auto;
                background: white;
                border-radius: 15px;
                overflow: hidden;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px;
                color: white;
                text-align: center;
            }}
            .header h1 {{
                font-size: 36px;
                margin-bottom: 10px;
                text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
            }}
            .summary-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                padding: 30px;
                background: #f8f9fa;
            }}
            .summary-card {{
                background: white;
                padding: 20px;
                border-radius: 12px;
                text-align: center;
                box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            }}
            .summary-card .value {{
                font-size: 32px;
                font-weight: bold;
                margin: 10px 0;
            }}
            .summary-card .label {{
                color: #7f8c8d;
                font-size: 14px;
                text-transform: uppercase;
            }}
            .data-table {{
                padding: 30px;
                overflow-x: auto;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                background: white;
            }}
            th {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 15px 10px;
                text-align: left;
                font-weight: 600;
                font-size: 12px;
                text-transform: uppercase;
                position: sticky;
                top: 0;
                z-index: 10;
            }}
            td {{
                padding: 12px 10px;
                border-bottom: 1px solid #ecf0f1;
                font-size: 13px;
            }}
            tr:hover {{
                background: #f8f9fa;
            }}
            .status-excellent {{ color: #2ecc71; font-weight: bold; }}
            .status-good {{ color: #4CAF50; font-weight: bold; }}
            .status-fair {{ color: #FFC107; font-weight: bold; }}
            .status-poor {{ color: #FF9800; font-weight: bold; }}
            .status-very-poor {{ color: #f44336; font-weight: bold; }}
            .present {{ color: #2ecc71; }}
            .absent {{ color: #f44336; }}
            @media print {{
                body {{ background: white; padding: 0; }}
                .container {{ box-shadow: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📋 Detailed Member Attendance Report</h1>
                <p>{church_name}</p>
                <p>{format_date(filters.get('from_date'))} - {format_date(filters.get('to_date'))}</p>
            </div>
            
            <div class="summary-grid">
    """
    
    # Summary cards
    for card in summary:
        value_display = card['value']
        if card['datatype'] == 'Percent':
            value_display = f"{card['value']:.1f}%"
        elif card['datatype'] == 'Int':
            value_display = f"{int(card['value']):,}"
        
        indicator_colors = {
            'Blue': '#1976D2',
            'Green': '#4CAF50',
            'Red': '#f44336'
        }
        color = indicator_colors.get(card['indicator'], '#667eea')
        
        html += f"""
                <div class="summary-card">
                    <div class="label">{card['label']}</div>
                    <div class="value" style="color: {color};">{value_display}</div>
                </div>
        """
    
    html += """
            </div>
            
            <div class="data-table">
                <table>
                    <thead>
                        <tr>
    """
    
    # Headers
    for col in columns:
        if not col['fieldname'].startswith('service_'):
            html += f"<th>{col['label']}</th>"
    
    html += """
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # Data rows
    for row in data:
        html += "<tr>"
        
        for col in columns:
            if col['fieldname'].startswith('service_'):
                continue
                
            value = row.get(col['fieldname'], '')
            
            # Format value
            if col['fieldtype'] == 'Percent':
                value = f"{float(value):.1f}%"
            elif col['fieldtype'] == 'Int':
                value = f"{int(value):,}"
            elif col['fieldtype'] == 'Date' and value:
                value = format_date(value)
            
            # Special styling
            cell_class = ''
            if col['fieldname'] == 'status':
                if '✅' in str(value):
                    cell_class = 'status-excellent'
                elif '🟢' in str(value):
                    cell_class = 'status-good'
                elif '🟡' in str(value):
                    cell_class = 'status-fair'
                elif '🟠' in str(value):
                    cell_class = 'status-poor'
                elif '🔴' in str(value):
                    cell_class = 'status-very-poor'
            
            html += f'<td class="{cell_class}">{value}</td>'
        
        html += "</tr>"
    
    html += f"""
                    </tbody>
                </table>
            </div>
            
            <div style="background: #2c3e50; color: white; padding: 20px; text-align: center;">
                <p>Generated on {format_date(nowdate())} | {church_name}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html


# ============================================================================
# EXCEL REPORT GENERATION  
# ============================================================================

def get_excel(filters):
    """Generate color-coded Excel report"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except ImportError:
        frappe.throw(_("Excel export requires openpyxl"))
    
    columns, data, _, chart, summary = execute(filters)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Attendance Report")
    
    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = '📋 Detailed Member Attendance Report'
    ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Report details
    ws.merge_cells('A2:J2')
    church_settings = frappe.get_single('Church Settings')
    church_name = church_settings.church_name if church_settings else 'Church'
    ws['A2'] = f"{church_name} | {format_date(filters.get('from_date'))} - {format_date(filters.get('to_date'))}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Column headers
    for col_idx, col in enumerate(columns, 1):
        if col['fieldname'].startswith('service_'):
            continue
            
        cell = ws.cell(row=4, column=col_idx)
        cell.value = col['label']
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='764BA2', end_color='764BA2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get('width', 100) / 7
    
    # Data rows with color coding
    for row_idx, row_data in enumerate(data, 5):
        attendance_pct = row_data.get('attendance_percentage', 0)
        
        # Determine row color based on attendance
        if attendance_pct >= 80:
            row_color = 'D4EDDA'  # Light green
        elif attendance_pct >= 60:
            row_color = 'D1ECF1'  # Light blue
        elif attendance_pct >= 40:
            row_color = 'FFF3CD'  # Light yellow
        elif attendance_pct >= 20:
            row_color = 'FFE5CC'  # Light orange
        else:
            row_color = 'F8D7DA'  # Light red
        
        col_idx = 1
        for col in columns:
            if col['fieldname'].startswith('service_'):
                continue
                
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col['fieldname'], '')
            
            # Format value
            if col['fieldtype'] == 'Percent':
                cell.value = float(value) / 100 if value else 0
                cell.number_format = '0.0%'
            elif col['fieldtype'] == 'Int':
                cell.value = int(value) if value else 0
                cell.number_format = '#,##0'
            else:
                cell.value = str(value)
            
            # Styling
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col_idx > 2 else 'left', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            col_idx += 1
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Save
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()