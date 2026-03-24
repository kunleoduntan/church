# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt

# Copyright (c) 2026, Church Management and contributors
# For license information, please see license.txt

"""
Demographic Attendance Analysis - Powerful Script Report

Features:
- Interactive filters (Date range, Branch, Demographic)
- Real-time data aggregation
- Beautiful HTML export with charts
- Colorful Excel export with formatting
- Visitor tracking and conversion analysis
- Trend indicators and statistics
"""

import frappe
from frappe import _
from frappe.utils import getdate, format_date, flt, cint, nowdate
from collections import defaultdict
import json


def execute(filters=None):
    """
    Main report execution function
    
    Returns:
        columns: List of column definitions
        data: List of data rows
        message: Optional message
        chart: Chart configuration
        report_summary: Summary statistics
    """
    
    if not filters:
        filters = {}
    
    # Validate filters
    validate_filters(filters)
    
    # Get columns
    columns = get_columns(filters)
    
    # Get data
    data = get_data(filters)
    
    # Get chart
    chart = get_chart_data(data, filters)
    
    # Get report summary
    report_summary = get_report_summary(data, filters)
    
    return columns, data, None, chart, report_summary


def validate_filters(filters):
    """Validate report filters"""
    
    if not filters.get('from_date'):
        filters['from_date'] = frappe.utils.add_months(nowdate(), -1)
    
    if not filters.get('to_date'):
        filters['to_date'] = nowdate()
    
    if getdate(filters['from_date']) > getdate(filters['to_date']):
        frappe.throw(_("From Date cannot be after To Date"))


def get_columns(filters):
    """
    Define report columns with beautiful formatting
    """
    
    columns = [
        {
            'fieldname': 'demographic_group',
            'label': _('Demographic'),
            'fieldtype': 'Data',
            'width': 120
        },
        {
            'fieldname': 'total_attendance',
            'label': _('Total Attendance'),
            'fieldtype': 'Int',
            'width': 130
        },
        {
            'fieldname': 'church_attendance',
            'label': _('Church'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'sunday_school',
            'label': _('Sunday School'),
            'fieldtype': 'Int',
            'width': 120
        },
        {
            'fieldname': 'unique_members',
            'label': _('Unique Members'),
            'fieldtype': 'Int',
            'width': 130
        },
        {
            'fieldname': 'total_visitors',
            'label': _('Visitors'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'first_time_visitors',
            'label': _('First Time'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'converted_visitors',
            'label': _('Converted'),
            'fieldtype': 'Int',
            'width': 100
        },
        {
            'fieldname': 'conversion_rate',
            'label': _('Conversion %'),
            'fieldtype': 'Percent',
            'width': 110
        },
        {
            'fieldname': 'average_per_service',
            'label': _('Avg/Service'),
            'fieldtype': 'Float',
            'width': 110,
            'precision': 1
        },
        {
            'fieldname': 'trend',
            'label': _('Trend'),
            'fieldtype': 'Data',
            'width': 100
        },
        {
            'fieldname': 'percentage',
            'label': _('% of Total'),
            'fieldtype': 'Percent',
            'width': 110
        }
    ]
    
    return columns


def get_data(filters):
    """
    Fetch and process data for the report
    """
    
    from_date = getdate(filters.get('from_date'))
    to_date = getdate(filters.get('to_date'))
    branch = filters.get('branch')
    demographic = filters.get('demographic_group')
    
    # Define demographic groups
    demographic_groups = ['Men', 'Women', 'Youth', 'Teens', 'Children']
    
    # Filter by specific demographic if selected
    if demographic:
        demographic_groups = [demographic]
    
    data = []
    total_attendance_all = 0
    
    # Collect data for each demographic
    for demo_group in demographic_groups:
        row_data = get_demographic_data(from_date, to_date, demo_group, branch)
        data.append(row_data)
        total_attendance_all += row_data.get('total_attendance', 0)
    
    # Calculate percentages
    for row in data:
        if total_attendance_all > 0:
            row['percentage'] = (row['total_attendance'] / total_attendance_all) * 100
        else:
            row['percentage'] = 0
    
    return data


def get_demographic_data(from_date, to_date, demographic_group, branch=None):
    """
    Get comprehensive data for a demographic group
    """
    
    # Church Attendance
    church_attendance = get_church_attendance_count(from_date, to_date, demographic_group, branch)
    
    # Sunday School
    sunday_school = get_sunday_school_count(from_date, to_date, demographic_group, branch)
    
    # Unique Members
    unique_members = get_unique_members_count(from_date, to_date, demographic_group, branch)
    
    # Visitors
    visitor_data = get_visitor_data(from_date, to_date, demographic_group, branch)
    
    # Service count for average
    service_count = get_service_count(from_date, to_date, branch)
    
    # Calculate totals
    total_attendance = church_attendance + sunday_school
    
    # Calculate average per service
    avg_per_service = total_attendance / service_count if service_count > 0 else 0
    
    # Calculate conversion rate
    conversion_rate = 0
    if visitor_data['total'] > 0:
        conversion_rate = (visitor_data['converted'] / visitor_data['total']) * 100
    
    # Determine trend
    trend = calculate_trend(from_date, to_date, demographic_group, branch)
    
    return {
        'demographic_group': demographic_group,
        'total_attendance': total_attendance,
        'church_attendance': church_attendance,
        'sunday_school': sunday_school,
        'unique_members': unique_members,
        'total_visitors': visitor_data['total'],
        'first_time_visitors': visitor_data['first_time'],
        'converted_visitors': visitor_data['converted'],
        'conversion_rate': conversion_rate,
        'average_per_service': avg_per_service,
        'trend': trend,
        'percentage': 0  # Will be calculated later
    }


def get_church_attendance_count(from_date, to_date, demographic_group, branch=None):
    """Get church attendance count"""
    
    try:
        conditions = []
        values = {
            'from_date': from_date,
            'to_date': to_date,
            'demographic_group': demographic_group
        }
        
        if branch:
            conditions.append("ca.branch = %(branch)s")
            values['branch'] = branch
        
        where_clause = " AND " + " AND ".join(conditions) if conditions else ""
        
        query = f"""
            SELECT COUNT(DISTINCT ca.name) as count
            FROM `tabChurch Attendance` ca
            INNER JOIN `tabMember` m ON ca.member_id = m.name
            WHERE ca.service_date BETWEEN %(from_date)s AND %(to_date)s
                AND ca.present = 1
                AND m.demographic_group = %(demographic_group)s
                AND m.member_status = 'Active'
                {where_clause}
        """
        
        result = frappe.db.sql(query, values=values, as_dict=True)
        return result[0]['count'] if result else 0
        
    except Exception as e:
        frappe.log_error(f"Error getting church attendance: {str(e)}")
        return 0


def get_sunday_school_count(from_date, to_date, demographic_group, branch=None):
    """Get Sunday School attendance count"""
    
    try:
        if not frappe.db.table_exists('Sunday School Attendance'):
            return 0
        
        filters_dict = {
            'service_date': ['between', [from_date, to_date]],
            'present': 1,
            'demographic_group': demographic_group
        }
        
        if branch:
            filters_dict['branch'] = branch
        
        count = frappe.db.count('Sunday School Attendance', filters_dict)
        return count
        
    except Exception as e:
        frappe.log_error(f"Error getting Sunday School attendance: {str(e)}")
        return 0


def get_unique_members_count(from_date, to_date, demographic_group, branch=None):
    """Get unique member count"""
    
    try:
        unique_ids = set()
        
        # From Church Attendance
        conditions = []
        values = {
            'from_date': from_date,
            'to_date': to_date,
            'demographic_group': demographic_group
        }
        
        if branch:
            conditions.append("ca.branch = %(branch)s")
            values['branch'] = branch
        
        where_clause = " AND " + " AND ".join(conditions) if conditions else ""
        
        query = f"""
            SELECT DISTINCT ca.member_id
            FROM `tabChurch Attendance` ca
            INNER JOIN `tabMember` m ON ca.member_id = m.name
            WHERE ca.service_date BETWEEN %(from_date)s AND %(to_date)s
                AND ca.present = 1
                AND m.demographic_group = %(demographic_group)s
                {where_clause}
        """
        
        result = frappe.db.sql(query, values=values, pluck='member_id')
        unique_ids.update(result)
        
        return len(unique_ids)
        
    except Exception as e:
        frappe.log_error(f"Error getting unique members: {str(e)}")
        return 0


def get_visitor_data(from_date, to_date, demographic_group, branch=None):
    """Get visitor statistics"""
    
    try:
        filters_dict = {
            'date_of_visit': ['between', [from_date, to_date]],
            'demographic_group': demographic_group
        }
        
        if branch:
            filters_dict['branch'] = branch
        
        visitors = frappe.get_all('Visitor',
            filters=filters_dict,
            fields=['name', 'visit_type', 'conversion_status']
        )
        
        data = {
            'total': len(visitors),
            'first_time': 0,
            'converted': 0
        }
        
        for visitor in visitors:
            if visitor.visit_type == 'First Time Visitor':
                data['first_time'] += 1
            
            if visitor.conversion_status == 'Converted to Member':
                data['converted'] += 1
        
        return data
        
    except Exception as e:
        frappe.log_error(f"Error getting visitor data: {str(e)}")
        return {'total': 0, 'first_time': 0, 'converted': 0}


def get_service_count(from_date, to_date, branch=None):
    """Get number of services in period"""
    
    try:
        # Get unique service dates from Church Attendance
        filters_dict = {
            'service_date': ['between', [from_date, to_date]],
            'present': 1
        }
        
        if branch:
            filters_dict['branch'] = branch
        
        # Get unique dates
        dates = frappe.get_all('Church Attendance',
            filters=filters_dict,
            fields=['service_date'],
            distinct=True
        )
        
        return len(dates) if dates else 1  # Return 1 to avoid division by zero
        
    except Exception as e:
        frappe.log_error(f"Error getting service count: {str(e)}")
        return 1


def calculate_trend(from_date, to_date, demographic_group, branch=None):
    """Calculate attendance trend"""
    
    try:
        # Get attendance over time
        conditions = []
        values = {
            'from_date': from_date,
            'to_date': to_date,
            'demographic_group': demographic_group
        }
        
        if branch:
            conditions.append("ca.branch = %(branch)s")
            values['branch'] = branch
        
        where_clause = " AND " + " AND ".join(conditions) if conditions else ""
        
        query = f"""
            SELECT ca.service_date, COUNT(*) as count
            FROM `tabChurch Attendance` ca
            INNER JOIN `tabMember` m ON ca.member_id = m.name
            WHERE ca.service_date BETWEEN %(from_date)s AND %(to_date)s
                AND ca.present = 1
                AND m.demographic_group = %(demographic_group)s
                {where_clause}
            GROUP BY ca.service_date
            ORDER BY ca.service_date
        """
        
        result = frappe.db.sql(query, values=values, as_dict=True)
        
        if len(result) < 4:
            return '➡️ Stable'
        
        # Split into halves
        mid = len(result) // 2
        first_half = result[:mid]
        second_half = result[mid:]
        
        first_avg = sum(r['count'] for r in first_half) / len(first_half)
        second_avg = sum(r['count'] for r in second_half) / len(second_half)
        
        if second_avg > first_avg * 1.1:
            return '📈 Increasing'
        elif second_avg < first_avg * 0.9:
            return '📉 Decreasing'
        else:
            return '➡️ Stable'
            
    except Exception as e:
        frappe.log_error(f"Error calculating trend: {str(e)}")
        return '➡️ Stable'


def get_chart_data(data, filters):
    """
    Generate beautiful chart configuration
    """
    
    if not data:
        return None
    
    # Prepare labels and datasets
    labels = [row['demographic_group'] for row in data]
    
    datasets = [
        {
            'name': 'Church Attendance',
            'values': [row['church_attendance'] for row in data]
        },
        {
            'name': 'Sunday School',
            'values': [row['sunday_school'] for row in data]
        },
        {
            'name': 'Visitors',
            'values': [row['total_visitors'] for row in data]
        }
    ]
    
    chart = {
        'data': {
            'labels': labels,
            'datasets': datasets
        },
        'type': 'bar',
        'height': 300,
        'colors': ['#1976D2', '#4CAF50', '#FF9800']
    }
    
    return chart


def get_report_summary(data, filters):
    """
    Generate report summary cards
    """
    
    if not data:
        return []
    
    total_attendance = sum(row['total_attendance'] for row in data)
    total_visitors = sum(row['total_visitors'] for row in data)
    total_converted = sum(row['converted_visitors'] for row in data)
    avg_conversion = sum(row['conversion_rate'] for row in data) / len(data) if data else 0
    
    # Find top demographic
    top_demo = max(data, key=lambda x: x['total_attendance']) if data else None
    
    summary = [
        {
            'value': total_attendance,
            'label': _('Total Attendance'),
            'indicator': 'Blue',
            'datatype': 'Int'
        },
        {
            'value': total_visitors,
            'label': _('Total Visitors'),
            'indicator': 'Orange',
            'datatype': 'Int'
        },
        {
            'value': total_converted,
            'label': _('Converted to Members'),
            'indicator': 'Green',
            'datatype': 'Int'
        },
        {
            'value': avg_conversion,
            'label': _('Avg Conversion Rate'),
            'indicator': 'Purple',
            'datatype': 'Percent'
        }
    ]
    
    if top_demo:
        summary.append({
            'value': top_demo['demographic_group'],
            'label': _('Top Demographic'),
            'indicator': 'Blue',
            'datatype': 'Data'
        })
    
    return summary


# ============================================================================
# HTML REPORT GENERATION
# ============================================================================

def get_html(filters):
    """
    Generate beautiful HTML report
    Called when user clicks "Print" or "Export HTML"
    """
    
    columns, data, _, chart, summary = execute(filters)
    
    church_settings = frappe.get_single('Church Settings')
    church_name = church_settings.church_name if church_settings else 'Church'
    
    # Demographic colors
    colors = {
        'Men': '#1976D2',
        'Women': '#E91E63',
        'Youth': '#4CAF50',
        'Teens': '#9C27B0',
        'Children': '#FF9800'
    }
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Demographic Attendance Report</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 20px;
            }}
            .container {{
                max-width: 1400px;
                margin: 0 auto;
                background: white;
                border-radius: 15px;
                overflow: hidden;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px;
                color: white;
                text-align: center;
                position: relative;
            }}
            .header h1 {{
                font-size: 36px;
                margin-bottom: 10px;
                text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
            }}
            .header p {{
                opacity: 0.9;
                font-size: 16px;
            }}
            .summary-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                padding: 30px;
                background: #f8f9fa;
            }}
            .summary-card {{
                background: white;
                padding: 20px;
                border-radius: 12px;
                text-align: center;
                box-shadow: 0 4px 12px rgba(0,0,0,0.1);
                border-top: 4px solid #667eea;
            }}
            .summary-card .value {{
                font-size: 32px;
                font-weight: bold;
                color: #2c3e50;
                margin: 10px 0;
            }}
            .summary-card .label {{
                color: #7f8c8d;
                font-size: 14px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            .data-table {{
                width: 100%;
                padding: 30px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                background: white;
            }}
            th {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 15px;
                text-align: left;
                font-weight: 600;
                font-size: 13px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            td {{
                padding: 15px;
                border-bottom: 1px solid #ecf0f1;
                font-size: 14px;
            }}
            tr:hover {{
                background: #f8f9fa;
            }}
            .demographic-cell {{
                font-weight: bold;
                font-size: 16px;
            }}
            .trend-cell {{
                font-weight: 600;
            }}
            .footer {{
                background: #2c3e50;
                color: white;
                padding: 20px;
                text-align: center;
                font-size: 12px;
            }}
            @media print {{
                body {{ background: white; padding: 0; }}
                .container {{ box-shadow: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📊 Demographic Attendance Report</h1>
                <p>{church_name}</p>
                <p>{format_date(filters.get('from_date'))} - {format_date(filters.get('to_date'))}</p>
                {f"<p>Branch: {filters.get('branch')}</p>" if filters.get('branch') else "<p>All Branches</p>"}
            </div>
            
            <div class="summary-grid">
    """
    
    # Add summary cards
    for card in summary:
        indicator_colors = {
            'Blue': '#1976D2',
            'Orange': '#FF9800',
            'Green': '#4CAF50',
            'Purple': '#9C27B0'
        }
        color = indicator_colors.get(card['indicator'], '#667eea')
        
        value_display = card['value']
        if card['datatype'] == 'Percent':
            value_display = f"{card['value']:.1f}%"
        elif card['datatype'] == 'Int':
            value_display = f"{int(card['value']):,}"
        
        html += f"""
                <div class="summary-card" style="border-top-color: {color};">
                    <div class="label">{card['label']}</div>
                    <div class="value" style="color: {color};">{value_display}</div>
                </div>
        """
    
    html += """
            </div>
            
            <div class="data-table">
                <table>
                    <thead>
                        <tr>
    """
    
    # Add table headers
    for col in columns:
        html += f"<th>{col['label']}</th>"
    
    html += """
                        </tr>
                    </thead>
                    <tbody>
    """
    
    # Add table rows
    for row in data:
        demo = row['demographic_group']
        color = colors.get(demo, '#667eea')
        
        html += "<tr>"
        for col in columns:
            value = row.get(col['fieldname'], '')
            
            # Format value based on type
            if col['fieldtype'] == 'Percent':
                value = f"{float(value):.1f}%"
            elif col['fieldtype'] == 'Float':
                value = f"{float(value):.1f}"
            elif col['fieldtype'] == 'Int':
                value = f"{int(value):,}"
            
            # Special styling for demographic and trend cells
            cell_class = ''
            cell_style = ''
            if col['fieldname'] == 'demographic_group':
                cell_class = 'demographic-cell'
                cell_style = f'color: {color};'
            elif col['fieldname'] == 'trend':
                cell_class = 'trend-cell'
            
            html += f'<td class="{cell_class}" style="{cell_style}">{value}</td>'
        
        html += "</tr>"
    
    html += f"""
                    </tbody>
                </table>
            </div>
            
            <div class="footer">
                <p>Generated on {format_date(nowdate())} | {church_name}</p>
                <p>⛪ Building faith together through data-driven insights</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html


# ============================================================================
# EXCEL REPORT GENERATION
# ============================================================================

def get_excel(filters):
    """
    Generate beautifully colored Excel report
    Called when user clicks "Export Excel"
    """
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import base64
        from io import BytesIO
    except ImportError:
        frappe.throw(_("Excel export requires openpyxl library"))
    
    columns, data, _, chart, summary = execute(filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create main sheet
    ws = wb.create_sheet("Demographic Report")
    
    # Define colors
    colors = {
        'Men': 'ADD8E6',        # Light Blue
        'Women': 'FFB6C1',      # Light Pink
        'Youth': '90EE90',      # Light Green
        'Teens': 'DDA0DD',      # Plum
        'Children': 'FFD700'    # Gold
    }
    
    # Header
    ws.merge_cells('A1:L1')
    ws['A1'] = '📊 Demographic Attendance Report'
    ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # Report details
    ws.merge_cells('A2:L2')
    church_settings = frappe.get_single('Church Settings')
    church_name = church_settings.church_name if church_settings else 'Church'
    ws['A2'] = f"{church_name} | {format_date(filters.get('from_date'))} - {format_date(filters.get('to_date'))}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Column headers (row 4)
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = col['label']
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='764BA2', end_color='764BA2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
        
        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get('width', 100) / 7
    
    ws.row_dimensions[4].height = 30
    
    # Data rows
    for row_idx, row_data in enumerate(data, 5):
        demo = row_data['demographic_group']
        row_color = colors.get(demo, 'FFFFFF')
        
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col['fieldname'], '')
            
            # Format value
            if col['fieldtype'] == 'Percent':
                cell.value = float(value) / 100 if value else 0
                cell.number_format = '0.0%'
            elif col['fieldtype'] == 'Float':
                cell.value = float(value) if value else 0
                cell.number_format = '0.0'
            elif col['fieldtype'] == 'Int':
                cell.value = int(value) if value else 0
                cell.number_format = '#,##0'
            else:
                cell.value = str(value)
            
            # Styling
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Bold demographic name
            if col['fieldname'] == 'demographic_group':
                cell.font = Font(bold=True, size=12)
        
        ws.row_dimensions[row_idx].height = 25
    
    # Summary section
    summary_row = len(data) + 6
    ws.merge_cells(f'A{summary_row}:L{summary_row}')
    ws[f'A{summary_row}'] = '📈 Report Summary'
    ws[f'A{summary_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{summary_row}'].fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    ws[f'A{summary_row}'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[summary_row].height = 30
    
    # Summary cards
    summary_row += 1
    for idx, card in enumerate(summary, summary_row):
        ws[f'A{idx}'] = card['label']
        ws[f'A{idx}'].font = Font(bold=True)
        
        value_cell = ws[f'B{idx}']
        if card['datatype'] == 'Percent':
            value_cell.value = card['value'] / 100
            value_cell.number_format = '0.0%'
        elif card['datatype'] == 'Int':
            value_cell.value = int(card['value'])
            value_cell.number_format = '#,##0'
        else:
            value_cell.value = card['value']
        
        value_cell.font = Font(bold=True, size=12)
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()