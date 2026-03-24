# Copyright (c) 2026, kunle and contributors
# For license information, please see license.txt

# -*- coding: utf-8 -*-
# Copyright (c) 2026, Church Management and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe import _
from frappe.utils import getdate, formatdate, format_date, format_datetime, now_datetime, flt, add_days
import json


def execute(filters=None):
    """
    Main execution function for Script Report
    Returns columns and data
    """
    if not filters:
        filters = {}
    
    # Validate filters
    validate_filters(filters)
    
    # Get data
    columns = get_columns()
    data = get_data(filters)
    
    # Add charts
    chart = get_chart_data(filters, data)
    
    # Add summary
    report_summary = get_report_summary(filters, data)
    
    return columns, data, None, chart, report_summary


def validate_filters(filters):
    """Validate filter combinations"""
    
    # Check if at least one filter is provided
    has_sheet = filters.get("attendance_sheet")
    has_date_range = filters.get("from_date") and filters.get("to_date")
    has_branch = filters.get("branch")
    
    if not has_sheet and not has_date_range and not has_branch:
        frappe.msgprint(_("Please select either an Attendance Sheet, Date Range, or Branch"))
        return False
    
    # Validate date range
    if filters.get("from_date") and filters.get("to_date"):
        if getdate(filters.get("from_date")) > getdate(filters.get("to_date")):
            frappe.throw(_("From Date cannot be greater than To Date"))
    
    return True


def get_columns():
    """Define report columns"""
    return [
        {
            "fieldname": "attendance_sheet",
            "label": _("Sheet"),
            "fieldtype": "Link",
            "options": "Attendance Sheet",
            "width": 150
        },
        {
            "fieldname": "branch",
            "label": _("Branch"),
            "fieldtype": "Link",
            "options": "Branch",
            "width": 120
        },
        {
            "fieldname": "date",
            "label": _("Date"),
            "fieldtype": "Date",
            "width": 100
        },
        {
            "fieldname": "day",
            "label": _("Day"),
            "fieldtype": "Data",
            "width": 80
        },
        {
            "fieldname": "programme",
            "label": _("Programme"),
            "fieldtype": "Data",
            "width": 180
        },
        {
            "fieldname": "men",
            "label": _("Men"),
            "fieldtype": "Int",
            "width": 70
        },
        {
            "fieldname": "women",
            "label": _("Women"),
            "fieldtype": "Int",
            "width": 80
        },
        {
            "fieldname": "children",
            "label": _("Children"),
            "fieldtype": "Int",
            "width": 80
        },
        {
            "fieldname": "total",
            "label": _("Total"),
            "fieldtype": "Int",
            "width": 80
        },
        {
            "fieldname": "new_men",
            "label": _("New Men"),
            "fieldtype": "Int",
            "width": 90
        },
        {
            "fieldname": "new_women",
            "label": _("New Women"),
            "fieldtype": "Int",
            "width": 100
        },
        {
            "fieldname": "new_children",
            "label": _("New Children"),
            "fieldtype": "Int",
            "width": 110
        },
        {
            "fieldname": "visitors",
            "label": _("Visitors"),
            "fieldtype": "Int",
            "width": 90
        },
        {
            "fieldname": "visitor_percent",
            "label": _("Visitor %"),
            "fieldtype": "Percent",
            "width": 90
        }
    ]


def get_data(filters):
    """Get report data based on filters"""
    
    data = []
    
    # Scenario 1: Single Attendance Sheet
    if filters.get("attendance_sheet"):
        data = get_single_sheet_data(filters)
    
    # Scenario 2: Date Range (with optional branch filter)
    elif filters.get("from_date") and filters.get("to_date"):
        data = get_date_range_data(filters)
    
    # Scenario 3: Branch only
    elif filters.get("branch"):
        data = get_branch_data(filters)
    
    return data


def get_single_sheet_data(filters):
    """Get data for a single attendance sheet"""
    
    attendance_sheet = filters.get("attendance_sheet")
    doc = frappe.get_doc("Attendance Sheet", attendance_sheet)
    
    data = []
    
    for row in doc.church_attendance_analysis:
        visitor_percent = calculate_percentage(row.new_total or 0, row.total or 0)
        
        data.append({
            "attendance_sheet": doc.name,
            "branch": doc.branch,
            "date": row.date,
            "day": row.day,
            "programme": row.programme,
            "men": row.men or 0,
            "women": row.women or 0,
            "children": row.children or 0,
            "total": row.total or 0,
            "new_men": row.new_men or 0,
            "new_women": row.new_women or 0,
            "new_children": row.new_children or 0,
            "visitors": row.new_total or 0,
            "visitor_percent": visitor_percent
        })
    
    return data


def get_date_range_data(filters):
    """Get data for a date range across multiple sheets"""
    
    from_date = getdate(filters.get("from_date"))
    to_date = getdate(filters.get("to_date"))
    
    # Build query filters
    sheet_filters = {
        "reporting_date": ["between", [from_date, to_date]]
    }
    
    if filters.get("branch"):
        sheet_filters["branch"] = filters.get("branch")
    
    # Get all attendance sheets in the date range
    sheets = frappe.get_all(
        "Attendance Sheet",
        filters=sheet_filters,
        fields=["name", "branch", "reporting_date"],
        order_by="reporting_date asc"
    )
    
    data = []
    
    for sheet in sheets:
        doc = frappe.get_doc("Attendance Sheet", sheet.name)
        
        for row in doc.church_attendance_analysis:
            visitor_percent = calculate_percentage(row.new_total or 0, row.total or 0)
            
            data.append({
                "attendance_sheet": doc.name,
                "branch": doc.branch,
                "date": row.date,
                "day": row.day,
                "programme": row.programme,
                "men": row.men or 0,
                "women": row.women or 0,
                "children": row.children or 0,
                "total": row.total or 0,
                "new_men": row.new_men or 0,
                "new_women": row.new_women or 0,
                "new_children": row.new_children or 0,
                "visitors": row.new_total or 0,
                "visitor_percent": visitor_percent
            })
    
    return data


def get_branch_data(filters):
    """Get latest attendance data for a specific branch"""
    
    branch = filters.get("branch")
    limit = filters.get("limit") or 10
    
    sheets = frappe.get_all(
        "Attendance Sheet",
        filters={"branch": branch},
        fields=["name", "branch", "reporting_date"],
        order_by="reporting_date desc",
        limit=limit
    )
    
    data = []
    
    for sheet in sheets:
        doc = frappe.get_doc("Attendance Sheet", sheet.name)
        
        for row in doc.church_attendance_analysis:
            visitor_percent = calculate_percentage(row.new_total or 0, row.total or 0)
            
            data.append({
                "attendance_sheet": doc.name,
                "branch": doc.branch,
                "date": row.date,
                "day": row.day,
                "programme": row.programme,
                "men": row.men or 0,
                "women": row.women or 0,
                "children": row.children or 0,
                "total": row.total or 0,
                "new_men": row.new_men or 0,
                "new_women": row.new_women or 0,
                "new_children": row.new_children or 0,
                "visitors": row.new_total or 0,
                "visitor_percent": visitor_percent
            })
    
    return data


def get_chart_data(filters, data):
    """Generate chart data for visualization"""
    
    if not data:
        return None
    
    # Aggregate totals
    total_men = sum(row.get("men", 0) for row in data)
    total_women = sum(row.get("women", 0) for row in data)
    total_children = sum(row.get("children", 0) for row in data)
    
    # Pie chart for attendance distribution
    chart = {
        "data": {
            "labels": ["Men", "Women", "Children"],
            "datasets": [
                {
                    "name": "Attendance Distribution",
                    "values": [total_men, total_women, total_children]
                }
            ]
        },
        "type": "pie",
        "colors": ["#4facfe", "#fa709a", "#a8edea"]
    }
    
    # If date range, also show trend
    if filters.get("from_date") and filters.get("to_date"):
        chart = get_trend_chart(data)
    
    return chart


def get_trend_chart(data):
    """Generate trend chart for date range"""
    
    # Group by date
    date_totals = {}
    for row in data:
        date_key = row.get("date")
        if date_key not in date_totals:
            date_totals[date_key] = {"total": 0, "visitors": 0}
        
        date_totals[date_key]["total"] += row.get("total", 0)
        date_totals[date_key]["visitors"] += row.get("visitors", 0)
    
    # Sort by date
    sorted_dates = sorted(date_totals.keys())
    
    chart = {
        "data": {
            "labels": [format_date(d, "dd MMM") for d in sorted_dates],
            "datasets": [
                {
                    "name": "Total Attendance",
                    "values": [date_totals[d]["total"] for d in sorted_dates]
                },
                {
                    "name": "Visitors",
                    "values": [date_totals[d]["visitors"] for d in sorted_dates]
                }
            ]
        },
        "type": "line",
        "colors": ["#667eea", "#f5576c"]
    }
    
    return chart


def get_report_summary(filters, data):
    """Generate report summary cards"""
    
    if not data:
        return []
    
    # Aggregate totals
    total_attendance = sum(row.get("total", 0) for row in data)
    total_men = sum(row.get("men", 0) for row in data)
    total_women = sum(row.get("women", 0) for row in data)
    total_children = sum(row.get("children", 0) for row in data)
    total_visitors = sum(row.get("visitors", 0) for row in data)
    
    # Count unique dates/sheets
    unique_dates = len(set(row.get("date") for row in data if row.get("date")))
    unique_sheets = len(set(row.get("attendance_sheet") for row in data if row.get("attendance_sheet")))
    
    summary = [
        {
            "value": total_attendance,
            "label": _("Total Attendance"),
            "datatype": "Int",
            "indicator": "blue"
        },
        {
            "value": total_men,
            "label": _("Men"),
            "datatype": "Int",
            "indicator": "green"
        },
        {
            "value": total_women,
            "label": _("Women"),
            "datatype": "Int",
            "indicator": "orange"
        },
        {
            "value": total_children,
            "label": _("Children"),
            "datatype": "Int",
            "indicator": "purple"
        },
        {
            "value": total_visitors,
            "label": _("Visitors"),
            "datatype": "Int",
            "indicator": "red"
        },
        {
            "value": calculate_percentage(total_visitors, total_attendance),
            "label": _("Visitor %"),
            "datatype": "Percent",
            "indicator": "yellow"
        }
    ]
    
    # Add period info if date range
    if filters.get("from_date") and filters.get("to_date"):
        summary.extend([
            {
                "value": unique_dates,
                "label": _("Days Covered"),
                "datatype": "Int",
                "indicator": "gray"
            },
            {
                "value": unique_sheets,
                "label": _("Sheets Analyzed"),
                "datatype": "Int",
                "indicator": "gray"
            }
        ])
    
    return summary


def calculate_percentage(value, total):
    """Calculate percentage"""
    if not total or total == 0:
        return 0
    return round((value / total) * 100, 1)


# =============================================================================
# CUSTOM EXPORT FUNCTIONS (Updated for period support)
# =============================================================================

@frappe.whitelist()
def export_period_to_excel(filters):
    """Export date range data to Excel"""
    
    if isinstance(filters, str):
        filters = json.loads(filters)
    
    # Generate data
    columns, data, _, _, _ = execute(filters)
    
    # Create Excel file
    return create_period_excel(filters, data)


def create_period_excel(filters, data):
    """Create Excel for period analysis"""
    from frappe.utils import format_date, format_datetime
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.throw(_("openpyxl library required. Run: bench pip install openpyxl"))
    
    import io
    import base64
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Period Analysis"
    
    # Header
    ws['A1'] = "ATTENDANCE ANALYSIS - PERIOD REPORT"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws.merge_cells('A1:N1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Period info
    if filters.get("from_date") and filters.get("to_date"):
        period_text = f"{format_date(filters['from_date'], 'dd MMM yyyy')} to {format_date(filters['to_date'], 'dd MMM yyyy')}"
    else:
        period_text = "All Data"
    
    ws['A2'] = period_text
    ws['A2'].font = Font(size=12, bold=True)
    ws.merge_cells('A2:N2')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25
    
    # Column headers
    headers = ["Sheet", "Branch", "Date", "Day", "Programme", "Men", "Women", "Children", 
               "Total", "New Men", "New Women", "New Children", "Visitors", "Visitor %"]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4FACFE", end_color="4FACFE", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, row_data in enumerate(data, 5):
        ws.cell(row=row_idx, column=1).value = row_data.get("attendance_sheet")
        ws.cell(row=row_idx, column=2).value = row_data.get("branch")
        ws.cell(row=row_idx, column=3).value = row_data.get("date")
        ws.cell(row=row_idx, column=4).value = row_data.get("day")
        ws.cell(row=row_idx, column=5).value = row_data.get("programme")
        ws.cell(row=row_idx, column=6).value = row_data.get("men")
        ws.cell(row=row_idx, column=7).value = row_data.get("women")
        ws.cell(row=row_idx, column=8).value = row_data.get("children")
        ws.cell(row=row_idx, column=9).value = row_data.get("total")
        ws.cell(row=row_idx, column=10).value = row_data.get("new_men")
        ws.cell(row=row_idx, column=11).value = row_data.get("new_women")
        ws.cell(row=row_idx, column=12).value = row_data.get("new_children")
        ws.cell(row=row_idx, column=13).value = row_data.get("visitors")
        ws.cell(row=row_idx, column=14).value = f"{row_data.get('visitor_percent')}%"
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    file_content = base64.b64encode(excel_file.read()).decode('utf-8')
    filename = f"Attendance_Period_{filters.get('from_date', 'all')}_{filters.get('to_date', 'all')}.xlsx"
    
    return {
        'success': True,
        'file_content': file_content,
        'filename': filename
    }


@frappe.whitelist()
def export_to_beautiful_excel(attendance_sheet):
    """Export to beautifully formatted Excel (single sheet)"""
    from church.church.doctype.attendance_sheet.attendance_excel_export import export_attendance_to_excel
    
    result = export_attendance_to_excel(attendance_sheet)
    return result


@frappe.whitelist()
def generate_html_report(attendance_sheet):
    """Generate beautiful HTML report (single sheet)"""
    from church.church.doctype.attendance_sheet.attendance_report import preview_report
    
    html = preview_report(attendance_sheet)
    return html


@frappe.whitelist()
def email_report(attendance_sheet, recipients):
    """Email the report (single sheet)"""
    from church.church.doctype.attendance_sheet.attendance_report import email_report as send_email
    
    result = send_email(attendance_sheet, recipients)
    return result