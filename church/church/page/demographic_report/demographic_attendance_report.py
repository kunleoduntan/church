# -*- coding: utf-8 -*-
"""
Demographic Attendance Report System - OPTIMIZED WITH VISITORS

Generates comprehensive attendance reports grouped by demographic:
- Men, Women, Youth, Teens, Children
- Multiple data sources: Church Attendance, Sunday School, Visitors
- Beautiful HTML and Excel exports
- Performance optimized with proper error handling
"""

import frappe
from frappe import _
from frappe.utils import getdate, format_date, fmt_money, flt, add_days, nowdate, cint
from collections import defaultdict
import json


@frappe.whitelist()
def generate_demographic_attendance_report(from_date, to_date, branch=None, report_type='summary'):
    """
    Generate comprehensive demographic attendance report with VISITORS
    
    Args:
        from_date: Start date for report
        to_date: End date for report
        branch: Optional branch filter
        report_type: 'summary' or 'detailed'
    
    Returns:
        Report data with statistics by demographic group
    """
    try:
        # Validate dates
        if not from_date or not to_date:
            return {
                'success': False,
                'error': 'Please provide both From Date and To Date'
            }
        
        from_date = getdate(from_date)
        to_date = getdate(to_date)
        
        if from_date > to_date:
            return {
                'success': False,
                'error': 'From Date cannot be after To Date'
            }
        
        # Initialize report data structure
        report_data = {
            'from_date': str(from_date),
            'to_date': str(to_date),
            'branch': branch or 'All Branches',
            'generated_at': nowdate(),
            'demographics': {},
            'has_data': False
        }
        
        # Get demographic groups
        demographic_groups = ['Men', 'Women', 'Youth', 'Teens', 'Children']
        
        # Collect data from all sources
        frappe.logger().info(f"Generating report from {from_date} to {to_date}")
        
        for demo_group in demographic_groups:
            frappe.logger().info(f"Fetching data for {demo_group}")
            
            church_att = get_church_attendance_by_demo(from_date, to_date, demo_group, branch)
            ss_att = get_sunday_school_attendance_by_demo(from_date, to_date, demo_group, branch)
            visitors = get_visitors_by_demo(from_date, to_date, demo_group, branch)
            unique_members = get_unique_members_by_demo(from_date, to_date, demo_group, branch)
            
            report_data['demographics'][demo_group] = {
                'church_attendance': church_att,
                'sunday_school': ss_att,
                'visitors': visitors,
                'unique_members': unique_members,
                'statistics': {}
            }
            
            # Check if we have any data
            if church_att or ss_att or visitors:
                report_data['has_data'] = True
        
        # Calculate statistics for each demographic
        for demo_group in demographic_groups:
            demo_data = report_data['demographics'][demo_group]
            demo_data['statistics'] = calculate_demographic_statistics(demo_data)
        
        # Calculate overall statistics
        report_data['overall_statistics'] = calculate_overall_statistics(report_data)
        
        frappe.logger().info(f"Report generated. Has data: {report_data['has_data']}")
        
        return {
            'success': True,
            'data': report_data
        }
        
    except Exception as e:
        frappe.log_error(
            message=frappe.get_traceback(),
            title="Demographic Report Generation Error"
        )
        return {
            'success': False,
            'error': str(e)
        }


def get_church_attendance_by_demo(from_date, to_date, demographic_group, branch=None):
    """Get Church Attendance records by demographic group - OPTIMIZED"""
    
    try:
        # Build base query
        conditions = []
        values = {
            'from_date': from_date,
            'to_date': to_date,
            'demographic_group': demographic_group
        }
        
        if branch:
            conditions.append("ca.branch = %(branch)s")
            values['branch'] = branch
        
        where_clause = " AND " + " AND ".join(conditions) if conditions else ""
        
        # Use SQL for better performance
        query = f"""
            SELECT DISTINCT
                ca.name,
                ca.member_id,
                ca.full_name,
                ca.service_date,
                ca.service_type,
                ca.branch,
                ca.time_in,
                ca.gender,
                ca.age
            FROM `tabChurch Attendance` ca
            INNER JOIN `tabMember` m ON ca.member_id = m.name
            WHERE ca.service_date BETWEEN %(from_date)s AND %(to_date)s
                AND ca.present = 1
                AND m.demographic_group = %(demographic_group)s
                AND m.member_status = 'Active'
                {where_clause}
            ORDER BY ca.service_date DESC
        """
        
        attendance = frappe.db.sql(query, values=values, as_dict=True)
        frappe.logger().info(f"Church Attendance for {demographic_group}: {len(attendance)} records")
        
        return attendance
        
    except Exception as e:
        frappe.log_error(f"Error fetching church attendance for {demographic_group}: {str(e)}")
        return []


def get_sunday_school_attendance_by_demo(from_date, to_date, demographic_group, branch=None):
    """Get Sunday School Attendance by demographic group - OPTIMIZED"""
    
    try:
        filters = {
            'service_date': ['between', [from_date, to_date]],
            'present': 1,
            'demographic_group': demographic_group
        }
        
        if branch:
            filters['branch'] = branch
        
        # Check if Sunday School Attendance table exists
        if not frappe.db.table_exists('Sunday School Attendance'):
            frappe.logger().info("Sunday School Attendance table does not exist")
            return []
        
        attendance = frappe.get_all('Sunday School Attendance',
            filters=filters,
            fields=[
                'name',
                'member_id',
                'full_name',
                'service_date',
                'sunday_school_class',
                'branch',
                'demographic_group'
            ],
            order_by='service_date desc'
        )
        
        frappe.logger().info(f"Sunday School for {demographic_group}: {len(attendance)} records")
        
        return attendance
        
    except Exception as e:
        frappe.log_error(f"Error fetching Sunday School attendance for {demographic_group}: {str(e)}")
        return []


def get_visitors_by_demo(from_date, to_date, demographic_group, branch=None):
    """Get Visitors by demographic group - NEW FEATURE"""
    
    try:
        filters = {
            'date_of_visit': ['between', [from_date, to_date]],
            'demographic_group': demographic_group
        }
        
        if branch:
            filters['branch'] = branch
        
        visitors = frappe.get_all('Visitor',
            filters=filters,
            fields=[
                'name',
                'full_name',
                'date_of_visit',
                'service_name',
                'branch',
                'demographic_group',
                'gender',
                'age',
                'visit_type',
                'conversion_status'
            ],
            order_by='date_of_visit desc'
        )
        
        frappe.logger().info(f"Visitors for {demographic_group}: {len(visitors)} records")
        
        return visitors
        
    except Exception as e:
        frappe.log_error(f"Error fetching visitors for {demographic_group}: {str(e)}")
        return []


def get_unique_members_by_demo(from_date, to_date, demographic_group, branch=None):
    """Get list of unique members who attended in this period"""
    
    unique_members = set()
    
    try:
        # From Church Attendance
        church_att = get_church_attendance_by_demo(from_date, to_date, demographic_group, branch)
        unique_members.update([att['member_id'] for att in church_att if att.get('member_id')])
        
        # From Sunday School
        ss_att = get_sunday_school_attendance_by_demo(from_date, to_date, demographic_group, branch)
        unique_members.update([att['member_id'] for att in ss_att if att.get('member_id')])
        
    except Exception as e:
        frappe.log_error(f"Error getting unique members for {demographic_group}: {str(e)}")
    
    return list(unique_members)


def calculate_demographic_statistics(demo_data):
    """Calculate statistics for a demographic group - ENHANCED WITH VISITORS"""
    
    try:
        # Get unique service dates from all attendance sources
        service_dates = set()
        
        for att in demo_data.get('church_attendance', []):
            if att.get('service_date'):
                service_dates.add(str(att['service_date']))
        
        for att in demo_data.get('sunday_school', []):
            if att.get('service_date'):
                service_dates.add(str(att['service_date']))
        
        # Visitor dates
        visitor_dates = set()
        for visitor in demo_data.get('visitors', []):
            if visitor.get('date_of_visit'):
                visitor_dates.add(str(visitor['date_of_visit']))
        
        total_services = len(service_dates)
        total_church = len(demo_data.get('church_attendance', []))
        total_ss = len(demo_data.get('sunday_school', []))
        total_visitors = len(demo_data.get('visitors', []))
        total_attendance = total_church + total_ss
        
        # Visitor analysis
        visitor_stats = analyze_visitors(demo_data.get('visitors', []))
        
        stats = {
            'total_church_attendance': total_church,
            'total_sunday_school_attendance': total_ss,
            'total_visitors': total_visitors,
            'unique_members_attended': len(demo_data.get('unique_members', [])),
            'total_services': total_services,
            'average_per_service': round(total_attendance / total_services, 1) if total_services > 0 else 0,
            'attendance_trend': 'stable',
            'visitor_stats': visitor_stats
        }
        
        # Calculate trend
        if total_church >= 4:
            sorted_attendance = sorted(
                demo_data.get('church_attendance', []), 
                key=lambda x: x.get('service_date', '')
            )
            
            if len(sorted_attendance) >= 4:
                mid = len(sorted_attendance) // 2
                first_half = sorted_attendance[:mid]
                second_half = sorted_attendance[mid:]
                
                if len(second_half) > len(first_half) * 1.1:
                    stats['attendance_trend'] = 'increasing'
                elif len(second_half) < len(first_half) * 0.9:
                    stats['attendance_trend'] = 'decreasing'
        
        return stats
        
    except Exception as e:
        frappe.log_error(f"Error calculating statistics: {str(e)}")
        return {
            'total_church_attendance': 0,
            'total_sunday_school_attendance': 0,
            'total_visitors': 0,
            'unique_members_attended': 0,
            'total_services': 0,
            'average_per_service': 0,
            'attendance_trend': 'stable',
            'visitor_stats': {}
        }


def analyze_visitors(visitors):
    """Analyze visitor data - NEW FEATURE"""
    
    stats = {
        'total': len(visitors),
        'first_time': 0,
        'return_visitors': 0,
        'converted': 0,
        'in_followup': 0,
        'by_visit_type': defaultdict(int),
        'by_conversion_status': defaultdict(int)
    }
    
    for visitor in visitors:
        visit_type = visitor.get('visit_type', 'Unknown')
        stats['by_visit_type'][visit_type] += 1
        
        if visit_type == 'First Time Visitor':
            stats['first_time'] += 1
        elif visit_type == 'Return Visitor':
            stats['return_visitors'] += 1
        
        conversion_status = visitor.get('conversion_status', 'New Visitor')
        stats['by_conversion_status'][conversion_status] += 1
        
        if conversion_status == 'Converted to Member':
            stats['converted'] += 1
        elif conversion_status == 'In Follow-up':
            stats['in_followup'] += 1
    
    # Convert defaultdict to dict
    stats['by_visit_type'] = dict(stats['by_visit_type'])
    stats['by_conversion_status'] = dict(stats['by_conversion_status'])
    
    return stats


def calculate_overall_statistics(report_data):
    """Calculate overall statistics across all demographics - ENHANCED"""
    
    overall = {
        'total_attendance_records': 0,
        'total_unique_members': 0,
        'total_visitors': 0,
        'demographic_breakdown': {},
        'visitor_breakdown': {},
        'top_demographic': None,
        'lowest_demographic': None,
        'visitor_conversion_rate': 0
    }
    
    total_converted = 0
    total_visitors_count = 0
    
    for demo_group, demo_data in report_data['demographics'].items():
        stats = demo_data.get('statistics', {})
        
        total_attendance = stats.get('total_church_attendance', 0) + \
                          stats.get('total_sunday_school_attendance', 0)
        visitor_count = stats.get('total_visitors', 0)
        
        overall['total_attendance_records'] += total_attendance
        overall['total_unique_members'] += stats.get('unique_members_attended', 0)
        overall['total_visitors'] += visitor_count
        overall['demographic_breakdown'][demo_group] = total_attendance
        overall['visitor_breakdown'][demo_group] = visitor_count
        
        # Track conversions
        visitor_stats = stats.get('visitor_stats', {})
        total_converted += visitor_stats.get('converted', 0)
        total_visitors_count += visitor_stats.get('total', 0)
    
    # Calculate conversion rate
    if total_visitors_count > 0:
        overall['visitor_conversion_rate'] = round(
            (total_converted / total_visitors_count) * 100, 1
        )
    
    # Find top and lowest demographics
    if overall['demographic_breakdown']:
        sorted_demos = sorted(
            overall['demographic_breakdown'].items(), 
            key=lambda x: x[1], 
            reverse=True
        )
        overall['top_demographic'] = sorted_demos[0][0] if sorted_demos else None
        overall['lowest_demographic'] = sorted_demos[-1][0] if sorted_demos else None
    
    return overall


@frappe.whitelist()
def generate_demographic_html_report(from_date, to_date, branch=None):
    """Generate beautiful HTML report - ENHANCED WITH VISITORS"""
    
    try:
        # Get report data
        result = generate_demographic_attendance_report(from_date, to_date, branch, 'summary')
        
        if not result.get('success'):
            return {
                'success': False,
                'error': result.get('error', 'Failed to generate report'),
                'html': generate_error_html(result.get('error', 'Unknown error'))
            }
        
        data = result['data']
        
        # Check if we have data
        if not data.get('has_data'):
            return {
                'success': True,
                'html': generate_no_data_html(data)
            }
        
        # Generate HTML
        html = generate_demographic_html(data)
        
        return {
            'success': True,
            'html': html
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Demographic HTML Report Error")
        return {
            'success': False,
            'error': str(e),
            'html': generate_error_html(str(e))
        }


def generate_error_html(error_message):
    """Generate error HTML"""
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Report Error</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                padding: 40px;
                background: #f5f5f5;
            }}
            .error-box {{
                max-width: 600px;
                margin: 0 auto;
                background: white;
                padding: 30px;
                border-radius: 8px;
                border-left: 4px solid #f44336;
            }}
            h2 {{ color: #f44336; }}
        </style>
    </head>
    <body>
        <div class="error-box">
            <h2>❌ Error Generating Report</h2>
            <p><strong>Error:</strong> {error_message}</p>
            <p>Please check your date range and filters, then try again.</p>
        </div>
    </body>
    </html>
    """


def generate_no_data_html(data):
    """Generate no data HTML"""
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>No Data</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                padding: 40px;
                background: #f5f5f5;
            }}
            .info-box {{
                max-width: 600px;
                margin: 0 auto;
                background: white;
                padding: 30px;
                border-radius: 8px;
                border-left: 4px solid #ff9800;
                text-align: center;
            }}
            h2 {{ color: #ff9800; }}
        </style>
    </head>
    <body>
        <div class="info-box">
            <h2>📊 No Data Found</h2>
            <p>No attendance records found for the selected period:</p>
            <p><strong>{format_date(data['from_date'])} - {format_date(data['to_date'])}</strong></p>
            <p><strong>Branch:</strong> {data['branch']}</p>
            <p>Try expanding your date range or selecting a different branch.</p>
        </div>
    </body>
    </html>
    """


def generate_demographic_html(data):
    """Generate beautiful HTML report - ENHANCED WITH VISITORS"""
    
    church_settings = frappe.get_single('Church Settings')
    church_name = church_settings.church_name if church_settings else 'Church'
    
    # Demographic colors
    demo_colors = {
        'Men': '#1976D2',
        'Women': '#E91E63',
        'Youth': '#4CAF50',
        'Teens': '#9C27B0',
        'Children': '#FF9800'
    }
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Demographic Attendance Report</title>
        <style>
            body {{
                font-family: 'Segoe UI', Arial, sans-serif;
                margin: 0;
                padding: 20px;
                background: #f5f5f5;
            }}
            .container {{
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px;
                color: white;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 32px;
                font-weight: 300;
            }}
            .header p {{
                margin: 10px 0 0 0;
                opacity: 0.9;
            }}
            .summary-cards {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                padding: 30px;
            }}
            .summary-card {{
                background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
                padding: 25px;
                border-radius: 12px;
                border-left: 4px solid #667eea;
                box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            }}
            .summary-card h3 {{
                margin: 0 0 10px 0;
                font-size: 14px;
                color: #7f8c8d;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            .summary-card .value {{
                font-size: 36px;
                font-weight: bold;
                color: #2c3e50;
            }}
            .summary-card .subvalue {{
                font-size: 14px;
                color: #95a5a6;
                margin-top: 5px;
            }}
            .demographic-section {{
                padding: 30px;
            }}
            .demographic-card {{
                background: #f8f9fa;
                border-radius: 12px;
                padding: 25px;
                margin-bottom: 20px;
                border-left: 5px solid #667eea;
            }}
            .demographic-header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 20px;
            }}
            .demographic-title {{
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
            }}
            .trend-badge {{
                padding: 6px 15px;
                border-radius: 20px;
                font-size: 12px;
                font-weight: bold;
                text-transform: uppercase;
            }}
            .trend-increasing {{
                background: #e8f5e9;
                color: #2e7d32;
            }}
            .trend-decreasing {{
                background: #ffebee;
                color: #c62828;
            }}
            .trend-stable {{
                background: #e3f2fd;
                color: #1976d2;
            }}
            .stats-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                gap: 15px;
                margin-top: 15px;
            }}
            .stat-item {{
                background: white;
                padding: 15px;
                border-radius: 8px;
            }}
            .stat-label {{
                font-size: 12px;
                color: #7f8c8d;
                margin-bottom: 5px;
            }}
            .stat-value {{
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
            }}
            .progress-bar {{
                height: 30px;
                background: #e0e0e0;
                border-radius: 15px;
                overflow: hidden;
                margin: 10px 0;
            }}
            .progress-fill {{
                height: 100%;
                display: flex;
                align-items: center;
                justify-content: center;
                color: white;
                font-weight: bold;
                font-size: 12px;
                transition: width 0.3s ease;
            }}
            @media print {{
                body {{ background: white; }}
                .container {{ box-shadow: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <!-- Header -->
            <div class="header">
                <h1>📊 Demographic Attendance Report</h1>
                <p>{church_name}</p>
                <p>{format_date(data['from_date'])} - {format_date(data['to_date'])}</p>
                <p>{data['branch']}</p>
            </div>
            
            <!-- Summary Cards -->
            <div class="summary-cards">
                <div class="summary-card">
                    <h3>Total Attendance</h3>
                    <div class="value">{data['overall_statistics']['total_attendance_records']:,}</div>
                </div>
                <div class="summary-card">
                    <h3>Unique Members</h3>
                    <div class="value">{data['overall_statistics']['total_unique_members']:,}</div>
                </div>
                <div class="summary-card" style="border-left-color: #FF9800;">
                    <h3>Total Visitors</h3>
                    <div class="value">{data['overall_statistics']['total_visitors']:,}</div>
                    <div class="subvalue">Conversion: {data['overall_statistics']['visitor_conversion_rate']}%</div>
                </div>
                <div class="summary-card">
                    <h3>Top Demographic</h3>
                    <div class="value" style="font-size: 20px;">{data['overall_statistics']['top_demographic'] or 'N/A'}</div>
                </div>
            </div>
            
            <!-- Demographic Breakdown Chart -->
            <div class="demographic-section">
                <h2 style="margin-bottom: 20px;">Demographic Breakdown</h2>
    """
    
    # Add progress bars for each demographic
    total = data['overall_statistics']['total_attendance_records']
    if total > 0:
        for demo_group in ['Men', 'Women', 'Youth', 'Teens', 'Children']:
            count = data['overall_statistics']['demographic_breakdown'].get(demo_group, 0)
            visitor_count = data['overall_statistics']['visitor_breakdown'].get(demo_group, 0)
            percentage = (count / total * 100) if total > 0 else 0
            color = demo_colors.get(demo_group, '#667eea')
            
            html += f"""
                <div style="margin-bottom: 15px;">
                    <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                        <span style="font-weight: bold; color: {color};">{demo_group}</span>
                        <span style="color: #7f8c8d;">
                            Attendance: {count:,} ({percentage:.1f}%) | Visitors: {visitor_count}
                        </span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: {percentage}%; background: {color};">
                            {percentage:.1f}%
                        </div>
                    </div>
                </div>
            """
    
    html += """
            </div>
            
            <!-- Individual Demographic Details -->
            <div class="demographic-section">
                <h2 style="margin-bottom: 20px;">Detailed Statistics</h2>
    """
    
    # Add detailed cards for each demographic
    for demo_group in ['Men', 'Women', 'Youth', 'Teens', 'Children']:
        demo_data = data['demographics'][demo_group]
        stats = demo_data['statistics']
        visitor_stats = stats.get('visitor_stats', {})
        color = demo_colors.get(demo_group, '#667eea')
        
        trend_class = f"trend-{stats['attendance_trend']}"
        trend_icon = '📈' if stats['attendance_trend'] == 'increasing' else ('📉' if stats['attendance_trend'] == 'decreasing' else '➡️')
        
        html += f"""
            <div class="demographic-card" style="border-left-color: {color};">
                <div class="demographic-header">
                    <div class="demographic-title" style="color: {color};">{demo_group}</div>
                    <div class="trend-badge {trend_class}">
                        {trend_icon} {stats['attendance_trend']}
                    </div>
                </div>
                
                <div class="stats-grid">
                    <div class="stat-item">
                        <div class="stat-label">Church Attendance</div>
                        <div class="stat-value">{stats['total_church_attendance']:,}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Sunday School</div>
                        <div class="stat-value">{stats['total_sunday_school_attendance']:,}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Unique Members</div>
                        <div class="stat-value">{stats['unique_members_attended']:,}</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-label">Avg Per Service</div>
                        <div class="stat-value">{stats['average_per_service']}</div>
                    </div>
                    <div class="stat-item" style="background: #fff3cd;">
                        <div class="stat-label">Total Visitors</div>
                        <div class="stat-value" style="color: #FF9800;">{visitor_stats.get('total', 0):,}</div>
                    </div>
                    <div class="stat-item" style="background: #d4edda;">
                        <div class="stat-label">First Time</div>
                        <div class="stat-value" style="font-size: 18px;">{visitor_stats.get('first_time', 0)}</div>
                    </div>
                    <div class="stat-item" style="background: #d1ecf1;">
                        <div class="stat-label">Return Visitors</div>
                        <div class="stat-value" style="font-size: 18px;">{visitor_stats.get('return_visitors', 0)}</div>
                    </div>
                    <div class="stat-item" style="background: #d4edda;">
                        <div class="stat-label">Converted</div>
                        <div class="stat-value" style="font-size: 18px; color: #28a745;">{visitor_stats.get('converted', 0)}</div>
                    </div>
                </div>
            </div>
        """
    
    html += """
            </div>
            
            <!-- Footer -->
            <div style="padding: 30px; text-align: center; background: #f8f9fa; color: #7f8c8d; font-size: 12px;">
                <p>Generated on """ + format_date(nowdate()) + """</p>
                <p>⛪ """ + church_name + """ - Building faith together</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html


@frappe.whitelist()
def generate_demographic_excel_report(from_date, to_date, branch=None):
    """Generate Excel report - ENHANCED WITH VISITORS"""
    
    try:
        # Check if openpyxl is installed
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import base64
            from io import BytesIO
        except ImportError as ie:
            return {
                'success': False,
                'error': f'Missing library: {str(ie)}. Please run: bench pip install openpyxl'
            }
        
        # Get report data
        result = generate_demographic_attendance_report(from_date, to_date, branch, 'detailed')
        
        if not result.get('success'):
            return {
                'success': False,
                'error': result.get('error', 'Failed to generate report')
            }
        
        data = result['data']
        
        # Check if we have data
        if not data.get('has_data'):
            return {
                'success': False,
                'error': 'No data found for the selected period. Please try a different date range.'
            }
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create sheets
        create_summary_sheet(wb, data)
        create_demographic_details_sheet(wb, data)
        create_visitor_details_sheet(wb, data)
        
        # Save to bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Encode to base64
        file_content = base64.b64encode(excel_file.read()).decode('utf-8')
        
        return {
            'success': True,
            'file_content': file_content,
            'filename': f'Demographic_Report_{from_date}_{to_date}.xlsx'
        }
        
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "Excel Report Error")
        return {
            'success': False,
            'error': f'Excel generation failed: {str(e)}'
        }


def create_summary_sheet(wb, data):
    """Create summary sheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        frappe.throw(_("openpyxl library is required. Please install: bench pip install openpyxl"))
    
    ws = wb.create_sheet("Summary")
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Demographic Attendance Report'
    ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Report details
    ws['A3'] = 'Period:'
    ws['B3'] = f"{data['from_date']} to {data['to_date']}"
    ws['A4'] = 'Branch:'
    ws['B4'] = data['branch']
    
    # Summary statistics
    ws['A6'] = 'Overall Statistics'
    ws['A6'].font = Font(bold=True, size=14)
    
    ws['A7'] = 'Total Attendance Records'
    ws['B7'] = data['overall_statistics']['total_attendance_records']
    
    ws['A8'] = 'Total Unique Members'
    ws['B8'] = data['overall_statistics']['total_unique_members']
    
    ws['A9'] = 'Total Visitors'
    ws['B9'] = data['overall_statistics']['total_visitors']
    
    ws['A10'] = 'Visitor Conversion Rate'
    ws['B10'] = f"{data['overall_statistics']['visitor_conversion_rate']}%"
    
    # Demographic breakdown
    ws['A12'] = 'Demographic Breakdown'
    ws['A12'].font = Font(bold=True, size=14)
    
    ws['A13'] = 'Demographic'
    ws['B13'] = 'Attendance'
    ws['C13'] = 'Visitors'
    ws['D13'] = 'Percentage'
    
    row = 14
    total = data['overall_statistics']['total_attendance_records']
    for demo, count in data['overall_statistics']['demographic_breakdown'].items():
        visitor_count = data['overall_statistics']['visitor_breakdown'].get(demo, 0)
        ws[f'A{row}'] = demo
        ws[f'B{row}'] = count
        ws[f'C{row}'] = visitor_count
        ws[f'D{row}'] = f"{(count/total*100):.1f}%" if total > 0 else "0%"
        row += 1
    
    # Auto-size columns
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 25


def create_demographic_details_sheet(wb, data):
    """Create detailed demographics sheet"""
    try:
        from openpyxl.styles import Font
    except ImportError:
        frappe.throw(_("openpyxl library is required. Please install: bench pip install openpyxl"))
    
    ws = wb.create_sheet("Demographics Detail")
    
    # Headers
    headers = [
        'Demographic', 'Church Attendance', 'Sunday School', 'Unique Members',
        'Total Services', 'Avg/Service', 'Visitors', 'First Time', 'Converted', 'Trend'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Data
    row = 2
    for demo_group in ['Men', 'Women', 'Youth', 'Teens', 'Children']:
        stats = data['demographics'][demo_group]['statistics']
        visitor_stats = stats.get('visitor_stats', {})
        
        ws.cell(row=row, column=1).value = demo_group
        ws.cell(row=row, column=2).value = stats['total_church_attendance']
        ws.cell(row=row, column=3).value = stats['total_sunday_school_attendance']
        ws.cell(row=row, column=4).value = stats['unique_members_attended']
        ws.cell(row=row, column=5).value = stats['total_services']
        ws.cell(row=row, column=6).value = stats['average_per_service']
        ws.cell(row=row, column=7).value = visitor_stats.get('total', 0)
        ws.cell(row=row, column=8).value = visitor_stats.get('first_time', 0)
        ws.cell(row=row, column=9).value = visitor_stats.get('converted', 0)
        ws.cell(row=row, column=10).value = stats['attendance_trend'].upper()
        row += 1
    
    # Auto-size columns
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18


def create_visitor_details_sheet(wb, data):
    """Create visitor details sheet - NEW"""
    try:
        from openpyxl.styles import Font
    except ImportError:
        frappe.throw(_("openpyxl library is required. Please install: bench pip install openpyxl"))
    
    ws = wb.create_sheet("Visitor Details")
    
    # Headers
    headers = ['Name', 'Date of Visit', 'Demographic', 'Branch', 'Visit Type', 'Conversion Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Data
    row = 2
    for demo_group in ['Men', 'Women', 'Youth', 'Teens', 'Children']:
        visitors = data['demographics'][demo_group].get('visitors', [])
        for visitor in visitors:
            ws.cell(row=row, column=1).value = visitor.get('full_name', '')
            ws.cell(row=row, column=2).value = str(visitor.get('date_of_visit', ''))
            ws.cell(row=row, column=3).value = visitor.get('demographic_group', '')
            ws.cell(row=row, column=4).value = visitor.get('branch', '')
            ws.cell(row=row, column=5).value = visitor.get('visit_type', '')
            ws.cell(row=row, column=6).value = visitor.get('conversion_status', '')
            row += 1
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 25